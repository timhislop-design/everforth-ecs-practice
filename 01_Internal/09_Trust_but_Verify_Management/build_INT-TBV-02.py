"""
Build INT-TBV-02 — Engagement Health Dashboard (template)
Weekly Excel workbook for ECS Engagement Managers to score all 5 health vectors.
Tabs: Roll-Up | Process Adoption | Config Hygiene | Custom Variance | Adoption Readiness | Sentiment & Trust
"""
import sys, os
HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(HERE, "..", ".."))
sys.path.insert(0, os.path.join(REPO, "03_Shared", "00_Templates_and_Branding"))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(HERE, "INT-TBV-02_Engagement_Health_Dashboard.xlsx")

# ── Palette (aligned with accelerator_pack_builder) ─────────────────────────
NAVY    = "0B1F3A"
WHITE   = "FFFFFF"
TEAL    = "14B8A6"
CYAN_BG = "ECFEFF"
AMBER   = "FEF3C7"
AMBER_T = "92400E"
SLATE   = "1E293B"
ALT     = "F1F5F9"
BORDER  = "E2E8F0"
GREEN   = "D1FAE5"
GREEN_T = "065F46"
YELLOW  = "FEF9C3"
YELLOW_T= "713F12"
ORANGE  = "FFEDD5"
ORANGE_T= "9A3412"
RED_BG  = "FEE2E2"
RED_T   = "991B1B"

F = "Arial"

def font(color=SLATE, bold=False, size=10): return Font(name=F, color=color, bold=bold, size=size)
def fill(c): return PatternFill(fill_type="solid", start_color=c, end_color=c)
def border(): return Border(
    left=Side(style="thin", color=BORDER), right=Side(style="thin", color=BORDER),
    top=Side(style="thin", color=BORDER), bottom=Side(style="thin", color=BORDER))
def al(h="left", v="top", wrap=True): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def banner(ws, row, text, span=8, subtitle=False):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(NAVY); c.font = font(WHITE, bold=not subtitle, size=13 if not subtitle else 9)
    c.alignment = al("left", "center"); ws.row_dimensions[row].height = 26 if not subtitle else 18

def hdr(ws, row, cols):
    for i, h in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = fill(NAVY); c.font = font(WHITE, bold=True); c.border = border()
        c.alignment = al("center", "center"); ws.row_dimensions[row].height = 22

def sec(ws, row, text, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(CYAN_BG); c.font = font(NAVY, bold=True); c.alignment = al()
    ws.row_dimensions[row].height = 20

def cell(ws, row, col, value, *, bold=False, bg=None, fg=SLATE, center=False, ht=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(fg, bold=bold); c.border = border()
    c.alignment = al("center" if center else "left", "center")
    if bg: c.fill = fill(bg)
    if ht: ws.row_dimensions[row].height = ht
    return c

def col_w(ws, widths):
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, row=5): ws.freeze_panes = ws.cell(row=row, column=1)

VECTORS = ["Process Adoption", "Config Hygiene", "Custom Variance", "Adoption Readiness", "Sentiment & Trust"]
SPRINTS = [f"Sprint {i}" for i in range(1, 7)]

# ── TAB 1: Roll-Up ───────────────────────────────────────────────────────────
def build_rollup(wb):
    ws = wb.create_sheet("Roll-Up")
    banner(ws, 1, "INT-TBV-02 · Engagement Health Dashboard — Practice Roll-Up", span=9)
    banner(ws, 2, "Practice Lead review · Updated monthly · One row per active engagement", span=9, subtitle=True)
    ws.row_dimensions[3].height = 8
    hdr(ws, 4, ["Engagement", "Process\nAdoption", "Config\nHygiene", "Custom\nVariance", "Adoption\nReadiness", "Sentiment\n& Trust", "Overall\nBand", "Last Updated", "Practice Lead Notes"])
    r = 5
    sample = [
        ("— Enter engagement name —", "Green", "Green", "Yellow", "Green", "Green", "Yellow", "Enter date", ""),
        ("", "", "", "", "", "", "", "", ""),
        ("", "", "", "", "", "", "", "", ""),
        ("", "", "", "", "", "", "", "", ""),
        ("", "", "", "", "", "", "", "", ""),
    ]
    band_color = {"Green": (GREEN, GREEN_T), "Yellow": (YELLOW, YELLOW_T), "Orange": (ORANGE, ORANGE_T), "Red": (RED_BG, RED_T)}
    for eng, pa, ch, cv, ar, st, band, upd, notes in sample:
        for col, val in enumerate([eng, pa, ch, cv, ar, st, band, upd, notes], 1):
            bg, fg = (None, SLATE)
            if val in band_color: bg, fg = band_color[val]
            cell(ws, r, col, val, bg=bg, fg=fg, center=(col in [2,3,4,5,6,7]))
        ws.row_dimensions[r].height = 22; r += 1
    r += 1
    sec(ws, r, "Band Definitions", span=9); r += 1
    defs = [("Green","0–5% custom variance · all vectors passing",""), ("Yellow","5–10% variance OR one vector at threshold","Raise in next Sponsor Sync"),
            ("Orange","10–15% variance OR two vectors yellow","Practice Lead joins Sponsor Sync · Council reviews backlog"),
            ("Red",">15% variance OR unannounced build detected","Course-correction protocol invoked (INT-TBV-08)")]
    hdr(ws, r, ["Band","Definition","Action Required"]); r += 1
    for band, defn, action in defs:
        bg, fg = band_color.get(band, (None, SLATE))
        cell(ws, r, 1, band, bg=bg, fg=fg, bold=True, center=True)
        cell(ws, r, 2, defn); cell(ws, r, 3, action); ws.row_dimensions[r].height = 22; r += 1
    col_w(ws, [32,12,12,12,12,12,10,14,42])
    freeze(ws)

# ── TAB 2: Process Adoption ──────────────────────────────────────────────────
def build_process_adoption(wb):
    ws = wb.create_sheet("Process Adoption")
    banner(ws, 1, "Process Adoption — Weekly Scoring", span=8)
    banner(ws, 2, "Measures: configured processes running through OOTB tables, states, and workflows without custom routing", span=8, subtitle=True)
    ws.row_dimensions[3].height = 8
    hdr(ws, 4, ["Week / Sprint", "Custom UI Policies\n(Yellow ≥1)", "Custom Client Scripts\n(Yellow ≥1)", "Custom Tables\n(Red ≥1)", "Custom Workflows\n(Red ≥1)", "Custom Routing\n(Red ≥1)", "Vector Score\n(Green/Yellow/Red)", "Evidence / Notes"])
    r = 5
    for sp in SPRINTS:
        for wk in ["Week A", "Week B"]:
            cell(ws, r, 1, f"{sp} – {wk}", bold=True)
            for col in range(2, 7): cell(ws, r, col, "0", center=True)
            cell(ws, r, 7, "Green", bg=GREEN, fg=GREEN_T, center=True)
            cell(ws, r, 8, "")
            ws.row_dimensions[r].height = 22; r += 1
        ws.row_dimensions[r].height = 6; r += 1
    r += 1
    sec(ws, r, "Thresholds", span=8); r += 1
    hdr(ws, r, ["Signal","Threshold","Band"]); r += 1
    band_map = {"Green":(GREEN,GREEN_T),"Yellow":(YELLOW,YELLOW_T),"Orange":(ORANGE,ORANGE_T),"Red":(RED_BG,RED_T)}
    for sig, thr, band in [
        ("Custom UI Policy or Client Script added","≥ 1 in sprint","Yellow"),
        ("Custom Table created","Any","Red"),
        ("Custom Workflow mirroring OOTB workflow","Any","Red"),
        ("Custom Routing Engine replacing Assignment Rules","Any","Red"),
    ]:
        bg2, fg2 = band_map[band]
        cell(ws, r, 1, sig); cell(ws, r, 2, thr)
        cell(ws, r, 3, band, bg=bg2, fg=fg2, center=True)
        ws.row_dimensions[r].height = 22; r += 1
    col_w(ws, [18,16,16,14,14,14,14,50])
    freeze(ws)

# ── TAB 3: Configuration Hygiene ─────────────────────────────────────────────
def build_config_hygiene(wb):
    ws = wb.create_sheet("Config Hygiene")
    banner(ws, 1, "Configuration Hygiene — Weekly Scoring", span=7)
    banner(ws, 2, "Measures: object counts and patterns within OOTB-defensible bounds", span=7, subtitle=True)
    ws.row_dimensions[3].height = 8
    hdr(ws, 4, ["Week / Sprint", "Catalog Item Count", "Category Depth\n(max levels)", "Assignment Rule\nConditions (max)", "SLA Schedule\nCount", "Vector Score\n(Green/Yellow/Red)", "Evidence / Notes"])
    r = 5
    for sp in SPRINTS:
        for wk in ["Week A", "Week B"]:
            cell(ws, r, 1, f"{sp} – {wk}", bold=True)
            for col in [2,3,4,5]: cell(ws, r, col, "—", center=True)
            cell(ws, r, 6, "Green", bg=GREEN, fg=GREEN_T, center=True)
            cell(ws, r, 7, "")
            ws.row_dimensions[r].height = 22; r += 1
        ws.row_dimensions[r].height = 6; r += 1
    r += 1
    sec(ws, r, "Reference Counts — Yellow / Red Thresholds", span=7); r += 1
    hdr(ws, r, ["Metric","Green (OOTB-defensible)","Yellow Threshold","Red Threshold"]); r += 1
    refs = [
        ("Catalog item count at Sprint 1","< 100","100–200","≥ 200 with no rationalization plan"),
        ("Catalog item count at Sprint 3","< 150","150–300","≥ 300 at Sprint 3"),
        ("Final catalog item count (Sprint 6)","≤ 80 rationalized","81–120","≥ 121 without written exception"),
        ("Category nesting depth","≤ 3 levels","4 levels","≥ 5 levels"),
        ("Assignment rule conditions per rule","≤ 3","4","≥ 5"),
        ("SLA schedule count","≤ 3","4–5","≥ 6"),
    ]
    for metric, g, y, r2 in refs:
        cell(ws, r, 1, metric, bold=True); cell(ws, r, 2, g, bg=GREEN, fg=GREEN_T)
        cell(ws, r, 3, y, bg=YELLOW, fg=YELLOW_T); cell(ws, r, 4, r2, bg=RED_BG, fg=RED_T)
        ws.row_dimensions[r].height = 28; r += 1
    col_w(ws, [20,14,14,12,12,16,50])
    freeze(ws)

# ── TAB 4: Customization Variance ────────────────────────────────────────────
def build_custom_variance(wb):
    ws = wb.create_sheet("Custom Variance")
    banner(ws, 1, "Customization Variance — Weekly Scoring", span=7)
    banner(ws, 2, "Links to INT-TBV-03 (Customization Variance Tracker) as system of record. Enter aggregate totals here weekly.", span=7, subtitle=True)
    ws.row_dimensions[3].height = 8
    hdr(ws, 4, ["Week / Sprint", "Cumulative Approved\nCustomizations (#)", "Cumulative Approved\nEffort (hrs)", "Sprint Capacity\n(hrs)", "Variance %\n(Effort / Capacity)", "Vector Score\n(Green/Yellow/Red)", "Notes"])
    r = 5
    for sp in SPRINTS:
        for wk in ["Week A", "Week B"]:
            cell(ws, r, 1, f"{sp} – {wk}", bold=True)
            for col in [2,3,4]: cell(ws, r, col, 0, center=True)
            cell(ws, r, 5, "=IF(D{0}=0,\"—\",TEXT(C{0}/D{0},\"0%\"))".format(r), center=True)
            cell(ws, r, 6, "Green", bg=GREEN, fg=GREEN_T, center=True)
            cell(ws, r, 7, "")
            ws.row_dimensions[r].height = 22; r += 1
        ws.row_dimensions[r].height = 6; r += 1
    r += 1
    sec(ws, r, "Variance Band Reference", span=7); r += 1
    hdr(ws, r, ["Band","Aggregate Variance vs Sprint Capacity","Action"]); r += 1
    bands = [
        ("Green","0–5%","No intervention. Continue weekly scan."),
        ("Yellow","5–10%","EM raises at next Sponsor Sync. Add to Practice Lead monthly review."),
        ("Orange","10–15%","Practice Lead joins Sponsor Sync. Council reviews backlog."),
        ("Red",">15% OR unannounced build","Course-correction protocol invoked (INT-TBV-08). Possible PCR."),
    ]
    colors = {"Green":(GREEN,GREEN_T),"Yellow":(YELLOW,YELLOW_T),"Orange":(ORANGE,ORANGE_T),"Red":(RED_BG,RED_T)}
    for band, var, action in bands:
        bg, fg = colors[band]
        cell(ws, r, 1, band, bg=bg, fg=fg, bold=True, center=True)
        cell(ws, r, 2, var); cell(ws, r, 3, action)
        ws.row_dimensions[r].height = 26; r += 1
    col_w(ws, [20,20,20,14,14,16,44])
    freeze(ws)

# ── TAB 5: Adoption Readiness ─────────────────────────────────────────────────
def build_adoption_readiness(wb):
    ws = wb.create_sheet("Adoption Readiness")
    banner(ws, 1, "Adoption Readiness — Weekly Scoring", span=6)
    banner(ws, 2, "Measures: customer team preparing to operate OOTB-aligned model, not custom workarounds", span=6, subtitle=True)
    ws.row_dimensions[3].height = 8
    hdr(ws, 4, ["Week / Sprint", "Customer SOPs reference\ncustom workarounds?", "UAT scenarios written\nagainst custom paths?", "Training material\nshows old-system flow?", "Vector Score\n(Green/Yellow/Red)", "Evidence / Notes"])
    r = 5
    for sp in SPRINTS:
        for wk in ["Week A", "Week B"]:
            cell(ws, r, 1, f"{sp} – {wk}", bold=True)
            for col in [2,3,4]: cell(ws, r, col, "No", bg=GREEN, fg=GREEN_T, center=True)
            cell(ws, r, 5, "Green", bg=GREEN, fg=GREEN_T, center=True)
            cell(ws, r, 6, "")
            ws.row_dimensions[r].height = 22; r += 1
        ws.row_dimensions[r].height = 6; r += 1
    r += 1
    sec(ws, r, "Signal Reference", span=6); r += 1
    hdr(ws, r, ["Signal","Yellow","Red"]); r += 1
    for sig, y, rd in [
        ("Customer SOP content","Draft SOP references a workaround","SOP formally documents custom path"),
        ("UAT scenarios","Some scenarios test custom features","UAT plan written entirely against custom build"),
        ("Training material","Screenshots show old-system UI","Training deck describes old-system flow as target"),
        ("Customer leadership","Questioning OOTB rationale repeatedly","Stating 'we will customize after go-live'"),
    ]:
        cell(ws, r, 1, sig, bold=True); cell(ws, r, 2, y, bg=YELLOW, fg=YELLOW_T); cell(ws, r, 3, rd, bg=RED_BG, fg=RED_T)
        ws.row_dimensions[r].height = 28; r += 1
    col_w(ws, [20,20,20,20,16,44])
    freeze(ws)

# ── TAB 6: Sentiment & Trust ──────────────────────────────────────────────────
def build_sentiment(wb):
    ws = wb.create_sheet("Sentiment & Trust")
    banner(ws, 1, "Sentiment & Trust — Bi-Weekly Scoring", span=6)
    banner(ws, 2, "Measures: sponsor and SME alignment with OOTB-first principle. Tracked qualitatively in Sponsor Sync notes (INT-TBV-04).", span=6, subtitle=True)
    ws.row_dimensions[3].height = 8
    hdr(ws, 4, ["Sync Date", "Sponsor Alignment\n(Green/Yellow/Red)", "SME Alignment\n(Green/Yellow/Red)", "Key Concern Raised", "EM Response", "Vector Score"])
    r = 5
    for _ in range(9):  # 9 bi-weekly syncs across 18 weeks
        cell(ws, r, 1, "Enter date"); cell(ws, r, 2, "Green", bg=GREEN, fg=GREEN_T, center=True)
        cell(ws, r, 3, "Green", bg=GREEN, fg=GREEN_T, center=True)
        cell(ws, r, 4, ""); cell(ws, r, 5, ""); cell(ws, r, 6, "Green", bg=GREEN, fg=GREEN_T, center=True)
        ws.row_dimensions[r].height = 22; r += 1
    r += 1
    sec(ws, r, "Sentiment Signal Reference", span=6); r += 1
    hdr(ws, r, ["Stakeholder","Yellow Signal","Red Signal"]); r += 1
    for who, y, rd in [
        ("Sponsor","Stops asking 'why OOTB' and starts asking 'why not custom'","Bypasses consultant to lobby ECS leadership for a customization"),
        ("Customer SME","Repeatedly cites a legacy system capability as a requirement","Formally escalates a customization refusal to their management"),
        ("Customer Developer","Asks to 'help configure' OOTB areas","Starts writing code against the instance without authorization"),
    ]:
        cell(ws, r, 1, who, bold=True); cell(ws, r, 2, y, bg=YELLOW, fg=YELLOW_T); cell(ws, r, 3, rd, bg=RED_BG, fg=RED_T)
        ws.row_dimensions[r].height = 32; r += 1
    col_w(ws, [14,16,16,36,36,14])
    freeze(ws)

# ── ASSEMBLE ─────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)
build_rollup(wb)
build_process_adoption(wb)
build_config_hygiene(wb)
build_custom_variance(wb)
build_adoption_readiness(wb)
build_sentiment(wb)
wb.save(OUT)
print(f"Saved: {OUT}")
