"""
Build INT-TBV-03 — Customization Variance Tracker
System of record for all customization requests through the deviation lifecycle.
Tabs: Instructions | Variance Log | Decision Log | Sprint Capacity Reference
"""
import sys, os
HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(HERE, "..", ".."))
sys.path.insert(0, os.path.join(REPO, "03_Shared", "00_Templates_and_Branding"))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(HERE, "INT-TBV-03_Customization_Variance_Tracker.xlsx")

NAVY    = "0B1F3A"; WHITE = "FFFFFF"; CYAN_BG = "ECFEFF"; AMBER = "FEF3C7"
AMBER_T = "92400E"; SLATE = "1E293B"; ALT = "F1F5F9"; BORDER_C = "E2E8F0"
GREEN   = "D1FAE5"; GREEN_T = "065F46"; YELLOW = "FEF9C3"; YELLOW_T = "713F12"
ORANGE  = "FFEDD5"; ORANGE_T = "9A3412"; RED_BG = "FEE2E2"; RED_T = "991B1B"
F = "Arial"

def font(color=SLATE, bold=False, size=10): return Font(name=F, color=color, bold=bold, size=size)
def fill(c): return PatternFill(fill_type="solid", start_color=c, end_color=c)
def bdr(): return Border(left=Side(style="thin", color=BORDER_C), right=Side(style="thin", color=BORDER_C),
    top=Side(style="thin", color=BORDER_C), bottom=Side(style="thin", color=BORDER_C))
def al(h="left", v="top", wrap=True): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def banner(ws, row, text, span=10, subtitle=False):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(NAVY); c.font = font(WHITE, bold=not subtitle, size=13 if not subtitle else 9)
    c.alignment = al("left", "center"); ws.row_dimensions[row].height = 26 if not subtitle else 18

def hdr(ws, row, cols):
    for i, h in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = fill(NAVY); c.font = font(WHITE, bold=True, size=10)
        c.border = bdr(); c.alignment = al("center", "center")
    ws.row_dimensions[row].height = 24

def sec(ws, row, text, span=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(CYAN_BG); c.font = font(NAVY, bold=True); c.alignment = al()
    ws.row_dimensions[row].height = 20

def cell(ws, row, col, value="", *, bold=False, bg=None, fg=SLATE, center=False, ht=22):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(fg, bold=bold); c.border = bdr()
    c.alignment = al("center" if center else "left", "top")
    if bg: c.fill = fill(bg)
    ws.row_dimensions[row].height = ht
    return c

def col_w(ws, widths):
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, row=5): ws.freeze_panes = ws.cell(row=row, column=1)

# ── TAB 1: Instructions ───────────────────────────────────────────────────────
def build_instructions(wb):
    ws = wb.create_sheet("Instructions")
    banner(ws, 1, "INT-TBV-03 · Customization Variance Tracker", span=3)
    banner(ws, 2, "ECS Federal · ServiceNow Practice · Internal Use Only", span=3, subtitle=True)
    ws.row_dimensions[3].height = 8

    rows = [
        ("Purpose",
         "System of record for every customization request raised during the engagement. "
         "Tracks each request from initial raise through Council decision and build disposition. "
         "Feeds the Customization Variance vector in the Engagement Health Dashboard (INT-TBV-02)."),
        ("Who maintains this",
         "Engagement Manager (primary). Solution Architect updates the OOTB Alternative Analysis column. "
         "Practice Lead reviews monthly in the cross-engagement roll-up."),
        ("Update cadence",
         "Real-time as requests are raised. Reviewed every Friday during the weekly variance scan. "
         "Summarized at bi-weekly Sponsor Sync (INT-TBV-04) and monthly Practice Review (INT-TBV-07)."),
        ("Key rule",
         "Every customization that is council-approved and built adds to variance. "
         "Requests rejected by the Council do NOT add to variance. "
         "Customizations discovered post-build (unannounced) add to variance AND trigger the Red band."),
        ("Related artifacts",
         "INT-TBV-01 (Manager Playbook) · INT-TBV-02 (Health Dashboard) · INT-TBV-04 (Sponsor Sync) "
         "· INT-TBV-05 (Council Pre-Read) · INT-TBV-08 (Course-Correction Playbook)"),
    ]
    r = 4
    for label, val in rows:
        ws.cell(row=r, column=1, value=label).font = font(NAVY, bold=True)
        c = ws.cell(row=r, column=2, value=val); c.font = font(SLATE)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = max(24, 16 + 5 * (len(val) // 60))
        r += 1

    r += 1
    sec(ws, r, "Deviation Lifecycle Stages", span=3); r += 1
    ws.row_dimensions[r].height = 22
    ws.cell(row=r, column=1, value="Stage").font = font(WHITE, bold=True); ws.cell(row=r, column=1).fill = fill(NAVY); ws.cell(row=r, column=1).border = bdr()
    ws.cell(row=r, column=2, value="Description").font = font(WHITE, bold=True); ws.cell(row=r, column=2).fill = fill(NAVY); ws.cell(row=r, column=2).border = bdr()
    ws.cell(row=r, column=3, value="Who Advances").font = font(WHITE, bold=True); ws.cell(row=r, column=3).fill = fill(NAVY); ws.cell(row=r, column=3).border = bdr()
    r += 1
    stages = [
        ("1 – Raise","Customer or consultant identifies a need that appears non-OOTB. Logged in Variance Log immediately.","Consultant / EM"),
        ("2 – Analyze","Solution Architect documents OOTB Alternative Analysis. Business outcome alignment assessed.","Solution Architect"),
        ("3 – Classify","EM classifies: configuration (within OOTB), customization (code/custom table), or PCR.","Engagement Manager"),
        ("4 – Recommend","SA prepares Council Pre-Read (INT-TBV-05). EM circulates 48 hrs before Council.","SA + EM"),
        ("5 – Decide","Customization Council issues one of four decisions: approve-sprint, approve-backlog, reject, PCR.","Practice Lead + Sponsor"),
        ("6 – Disposition","Decision executed: backlog item added, Council rejection documented, PCR drafted, or build planned.","EM / Architect"),
    ]
    for stage, desc, who in stages:
        ws.cell(row=r, column=1, value=stage).font = font(NAVY, bold=True); ws.cell(row=r, column=1).border = bdr()
        c2 = ws.cell(row=r, column=2, value=desc); c2.font = font(SLATE); c2.border = bdr(); c2.alignment = Alignment(wrap_text=True, vertical="top")
        c3 = ws.cell(row=r, column=3, value=who); c3.font = font(SLATE); c3.border = bdr()
        ws.row_dimensions[r].height = max(24, 14 + 5 * (len(desc) // 50)); r += 1

    col_w(ws, [24, 88, 22])

# ── TAB 2: Variance Log ───────────────────────────────────────────────────────
def build_variance_log(wb):
    ws = wb.create_sheet("Variance Log")
    banner(ws, 1, "Customization Variance Log — One Row Per Request", span=12)
    banner(ws, 2, "Add a row each time a customization request is raised. Yellow cells = customer input / Council decision", span=12, subtitle=True)
    ws.row_dimensions[3].height = 8
    hdr(ws, 4, [
        "Req ID", "Sprint\nRaised", "Requester", "Description of Request",
        "OOTB Alternative\n(SA Analysis)", "Classification\n(Config/Custom/PCR)",
        "Council\nDecision", "Decision\nDate", "Sprint\nBuilt",
        "Est. Effort\n(hrs)", "Actual\nEffort (hrs)", "Post-GoLive\nOwner"
    ])
    # Sample rows (user fills in)
    r = 5
    for i in range(1, 16):
        alt = i % 2 == 0
        req_bg = ALT if alt else None
        cell(ws, r, 1, f"CVT-{i:03d}", bg=req_bg)
        cell(ws, r, 2, "", bg=req_bg)
        cell(ws, r, 3, "", bg=req_bg)
        cell(ws, r, 4, "", bg=req_bg)
        cell(ws, r, 5, "", bg=req_bg)  # SA fills
        cell(ws, r, 6, "", bg=AMBER, fg=AMBER_T)  # EM classifies
        cell(ws, r, 7, "", bg=AMBER, fg=AMBER_T)  # Council decides
        cell(ws, r, 8, "", bg=req_bg)
        cell(ws, r, 9, "", bg=req_bg)
        cell(ws, r, 10, 0, bg=req_bg, center=True)
        cell(ws, r, 11, 0, bg=req_bg, center=True)
        cell(ws, r, 12, "", bg=req_bg)
        ws.row_dimensions[r].height = 28; r += 1

    r += 1
    sec(ws, r, "Valid Classification Values", span=12); r += 1
    ws.cell(row=r, column=1, value="Configuration").font = font(GREEN_T, bold=True); ws.cell(row=r, column=1).fill = fill(GREEN); ws.cell(row=r, column=1).border = bdr()
    ws.cell(row=r, column=2, value="Within OOTB — a setting, not a code change. Does NOT add to variance.").font = font(SLATE); ws.cell(row=r, column=2).border = bdr(); ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 22; r += 1
    ws.cell(row=r, column=1, value="Customization").font = font(YELLOW_T, bold=True); ws.cell(row=r, column=1).fill = fill(YELLOW); ws.cell(row=r, column=1).border = bdr()
    ws.cell(row=r, column=2, value="Code change or custom table. Council decision required. ADDS to variance if approved.").font = font(SLATE); ws.cell(row=r, column=2).border = bdr(); ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 22; r += 1
    ws.cell(row=r, column=1, value="PCR").font = font(RED_T, bold=True); ws.cell(row=r, column=1).fill = fill(RED_BG); ws.cell(row=r, column=1).border = bdr()
    ws.cell(row=r, column=2, value="Scope change exceeding SOW. Pause sprint — PCR drafted before any commitment.").font = font(SLATE); ws.cell(row=r, column=2).border = bdr(); ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 22

    col_w(ws, [9, 8, 14, 38, 38, 16, 16, 12, 10, 10, 12, 18])
    freeze(ws)

# ── TAB 3: Decision Log ───────────────────────────────────────────────────────
def build_decision_log(wb):
    ws = wb.create_sheet("Decision Log")
    banner(ws, 1, "Council Decision Log — Closed Requests", span=8)
    banner(ws, 2, "One row per request that has received a final Council decision. Reference for pattern-spotting.", span=8, subtitle=True)
    ws.row_dimensions[3].height = 8
    hdr(ws, 4, ["Req ID","Sprint","Decision","Decision Date","Rationale","Deviation Approved?","Effort Added (hrs)","Notes"])
    r = 5
    for i in range(1, 11):
        alt = i % 2 == 0
        bg = ALT if alt else None
        cell(ws, r, 1, f"CVT-{i:03d}", bg=bg)
        for col in [2,3,4,5,7,8]: cell(ws, r, col, "", bg=bg)
        cell(ws, r, 6, "No", bg=GREEN, fg=GREEN_T, center=True)
        ws.row_dimensions[r].height = 22; r += 1
    r += 1
    sec(ws, r, "Decision Options Reference", span=8); r += 1
    hdr(ws, r, ["Decision","Meaning","Variance Impact","Next Action"]); r += 1
    for dec, meaning, impact, action in [
        ("Approve – Sprint","Build in current or next sprint","Adds effort to variance","Add to sprint backlog · update Variance Log"),
        ("Approve – Backlog","Defer to post-go-live","No current variance impact · tracked separately","Add to product backlog · inform sponsor"),
        ("Reject","OOTB alternative sufficient","No variance impact","Document rationale · close ticket"),
        ("PCR Required","Scope exceeds SOW","Pause build · commercial discussion","Draft PCR · Practice Lead + Sales lead discussion"),
    ]:
        cell(ws, r, 1, dec, bold=True, bg=CYAN_BG, fg=NAVY)
        cell(ws, r, 2, meaning); cell(ws, r, 3, impact); cell(ws, r, 4, action)
        ws.row_dimensions[r].height = 28; r += 1
    col_w(ws, [10, 8, 20, 14, 44, 16, 14, 32])
    freeze(ws)

# ── TAB 4: Sprint Capacity Reference ─────────────────────────────────────────
def build_capacity(wb):
    ws = wb.create_sheet("Sprint Capacity Reference")
    banner(ws, 1, "Sprint Capacity Reference — Variance Band Calculator", span=6)
    banner(ws, 2, "Enter team size and daily hours to auto-calculate sprint capacity and band thresholds", span=6, subtitle=True)
    ws.row_dimensions[3].height = 8

    r = 4
    inputs = [
        ("Consultants on engagement (FTE)", 3),
        ("Billable hours per consultant per day", 6),
        ("Working days per sprint (2 weeks)", 10),
    ]
    sec(ws, r, "Inputs — Update these values for your engagement", span=6); r += 1
    for label, default in inputs:
        ws.cell(row=r, column=1, value=label).font = font(NAVY, bold=True); ws.cell(row=r, column=1).border = bdr()
        c = ws.cell(row=r, column=2, value=default); c.fill = fill(AMBER); c.font = font(AMBER_T, bold=True)
        c.border = bdr(); c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[r].height = 22; r += 1

    r += 1
    sec(ws, r, "Calculated Thresholds (auto-updates when inputs change)", span=6); r += 1
    hdr(ws, r, ["Metric","Formula","Value","Notes"]); r += 1

    # Sprint capacity = B4 * B5 * B6
    formulas = [
        ("Sprint Capacity (hrs)","=B5*B6*B7","","Total available consultant hours per sprint"),
        ("Green ceiling (5%)","=B5*B6*B7*0.05","","Max approved custom effort before Yellow"),
        ("Yellow ceiling (10%)","=B5*B6*B7*0.10","","Max before Orange band"),
        ("Orange ceiling (15%)","=B5*B6*B7*0.15","","Max before Red band"),
        ("18-week total capacity (9 sprints)","=B5*B6*B7*9","","Full engagement capacity for context"),
    ]
    row_ref = 5  # row of first input
    for label, formula, val, notes in formulas:
        ws.cell(row=r, column=1, value=label).font = font(NAVY, bold=True); ws.cell(row=r, column=1).border = bdr()
        ws.cell(row=r, column=2, value=formula).font = font(SLATE); ws.cell(row=r, column=2).border = bdr()
        c = ws.cell(row=r, column=3, value=formula); c.fill = fill(GREEN); c.font = font(GREEN_T, bold=True)
        c.border = bdr(); c.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=4, value=notes).font = font(SLATE); ws.cell(row=r, column=4).border = bdr()
        ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 22; r += 1

    r += 1
    sec(ws, r, "Per-Sprint Variance Tracker — Paste cumulative approved effort from Variance Log", span=6); r += 1
    hdr(ws, r, ["Sprint","Capacity (hrs)","Cumulative Approved\nCustom Effort (hrs)","Variance %","Band","EM Action"]); r += 1
    for sp in [f"Sprint {i}" for i in range(1, 7)]:
        cell(ws, r, 1, sp, bold=True)
        cell(ws, r, 2, "=B5*B6*B7", center=True)
        cell(ws, r, 3, 0, bg=AMBER, fg=AMBER_T, center=True)
        cell(ws, r, 4, f"=IF(B{r}=0,\"—\",TEXT(C{r}/B{r},\"0.0%\"))", center=True)
        cell(ws, r, 5, "Green", bg=GREEN, fg=GREEN_T, center=True)
        cell(ws, r, 6, "Continue weekly scan")
        ws.row_dimensions[r].height = 22; r += 1

    col_w(ws, [36, 14, 22, 12, 10, 44])
    freeze(ws)

# ── ASSEMBLE ──────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)
build_instructions(wb)
build_variance_log(wb)
build_decision_log(wb)
build_capacity(wb)
wb.save(OUT)
print(f"Saved: {OUT}")
