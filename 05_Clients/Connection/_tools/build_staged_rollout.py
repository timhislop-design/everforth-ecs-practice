# -*- coding: utf-8 -*-
"""TRIGGER: build the JIT staged-rollout structure for client document releases.

Copies finalized client-appropriate artifacts from the static library
(06_Client_Upload/Connection) into 06_Client_Upload/Connection_Staged_Rollout/,
organized into five sequential drops aligned to the Phase 1 stage plan. Each
drop is released to Connection when its phase approaches — just-in-time, so the
client is never overwhelmed and never walks into a workshop cold.

Drop 1 = the initial engagement package (orientation + plans + obligations).
Drops 2-5 = workshop pre-reads, decks, and data-collection packs per stage,
then testing / go-live / handoff material.

Also writes Staged_Rollout_Guide.xlsx (what's in each drop, when to send, why)
and a ready-to-send zip per drop.

Run AFTER export_execution_library.py (the static library is the source).
Re-runnable: overwrites in place. Usage: python build_staged_rollout.py
"""
import os, shutil, zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(HERE, "..", "..", ".."))
SRC = os.path.join(REPO, "06_Client_Upload", "Connection")
ROLLOUT = os.path.join(REPO, "06_Client_Upload", "Connection_Staged_Rollout")
DST = os.path.join(ROLLOUT, "Client_Drops")          # client-facing releases only
KIT = os.path.join(ROLLOUT, "Internal_Release_Kit")  # ECS-only: guide, emails, runbook

NAVY = "0B1F3A"; WHITE = "FFFFFF"; TEAL = "14B8A6"; ALT = "F8FAFC"
BORDER = "E2E8F0"; SLATE = "475569"
thin = Side(style="thin", color=BORDER)
box = Border(left=thin, right=thin, top=thin, bottom=thin)

WS = "02_Delivery/Workshops"
PR = "02_Delivery/Workshops/Pre_Reads"
AP = "02_Delivery/Accelerator_Packs"

# (drop folder, when to send, purpose, [(src rel path, note), ...])
DROPS = [
    ("Drop_01_Initial_Package",
     "Now — with / just ahead of the Sprint 0 kickoff",
     "Orientation: what the project is, how it will run, who does what, and what Connection owns. Enough to engage leadership without overwhelming anyone.",
     [
        ("01_Onboarding/Client_Facing/Connection_Client_Onboarding_Guide.docx",
         "Start here — what to expect, the 18-week journey, governance in plain terms"),
        ("01_Onboarding/Client_Facing/Connection_Document_Roadmap.docx",
         "The five-drop document roadmap — what arrives when, and why staged (JIT by design)"),
        ("01_Onboarding/Client_Facing/Connection_Communication_Plan.docx",
         "Every touchpoint, owner, and cadence — plus the contact roster to complete at kickoff"),
        ("01_Onboarding/Client_Facing/Customer_Governance_Charter.docx",
         "The operating model and two-key decision rights"),
        ("01_Onboarding/Client_Facing/Connection_Kickoff_Deck.pptx",
         "The kickoff presentation — share after the session as the standing reference"),
        ("01_Onboarding/Connection_Onboarding_Checklist.xlsx",
         "The Sprint 0 plan — 19 readiness tasks with owners (ECS / Connection / Joint)"),
        ("02_Delivery/Connection_18Week_Project_Plan.xlsx",
         "The overall project plan — Sprints 0-8 across four stages, Go-Live Week 16"),
        ("02_Delivery/Connection_Customer_Dependency_Tracker.xlsx",
         "Connection-owned obligations (data, access, decisions) with timing and impact-if-late"),
        ("02_Delivery/Connection_SOW_Deliverables_Matrix.xlsx",
         "The deliverable checklist — every SOW-committed deliverable and its status"),
     ]),
    ("Drop_02_Foundation_and_Data",
     "Stage 1 — Sprints 1-2 (Weeks 3-6), release ~1 week before the first workshops",
     "Platform foundation and data-model workshops plus the data-collection packs Connection fills. Pre-reads go out ahead of each session so no one walks in cold.",
     [
        (PR + "/Connection_WP_01_Platform_Foundation_CLIENT.docx", "Pre-read — Platform Foundation workshop"),
        (PR + "/Connection_WP_12_CSDM_CLIENT.docx", "Pre-read — CSDM workshop"),
        (PR + "/Connection_WP_13_CMDB_CLIENT.docx", "Pre-read — CMDB workshop"),
        (PR + "/Connection_WP_14_Discovery_CLIENT.docx", "Pre-read — Discovery workshop"),
        (PR + "/Connection_WP_15_Service_Graph_Connectors_CLIENT.docx", "Pre-read — Service Graph Connectors workshop"),
        (WS + "/Connection_Platform_Foundation_Workshop.pptx", "Workshop deck" + " — release AFTER the session", "Post_Workshop"),
        (WS + "/Connection_CSDM_Workshop.pptx", "Workshop deck" + " — release AFTER the session", "Post_Workshop"),
        (WS + "/Connection_CMDB_Workshop.pptx", "Workshop deck" + " — release AFTER the session", "Post_Workshop"),
        (WS + "/Connection_Discovery_Workshop.pptx", "Workshop deck" + " — release AFTER the session", "Post_Workshop"),
        (WS + "/Connection_Service_Graph_Connectors_Workshop.pptx", "Workshop deck" + " — release AFTER the session", "Post_Workshop"),
        (WS + "/Connection_Integrations_Workshop.pptx", "Workshop deck — AD/SSO, SCCM, Intune" + " — release AFTER the session", "Post_Workshop"),
        (AP + "/Foundation_Accelerator_Pack", "Data pack — users, locations, groups, SLAs (Connection fills)"),
        (AP + "/CMDB_CSDM_Accelerator_Pack", "Data pack — service taxonomy, CI classes, relationships"),
        (AP + "/Discovery_Accelerator_Pack", "Data pack — discovery scope, MID servers, credentials"),
        (AP + "/Integration_Accelerator_Pack", "Data pack — AD/SSO, SCCM, Intune, Vonage CTI"),
        ("02_Delivery/Connection_RACI_Matrix.xlsx", "Who owns what, by deliverable — baseline for the working rhythm"),
     ]),
    ("Drop_03_ITSM_Core",
     "Stage 2 — Sprints 3-4 (Weeks 7-10), release ~1 week before the ITSM workshops",
     "The ITSM Core build: Incident, Major Incident, Problem, Change, and Service Catalog workshops with their pre-reads and process-decision packs.",
     [
        (PR + "/Connection_WP_02_Incident_Management_CLIENT.docx", "Pre-read — Incident Management workshop"),
        (PR + "/Connection_WP_03_Major_Incident_Management_CLIENT.docx", "Pre-read — Major Incident Management workshop"),
        (PR + "/Connection_WP_04_Problem_Management_CLIENT.docx", "Pre-read — Problem Management workshop"),
        (PR + "/Connection_WP_05_Change_Management_CLIENT.docx", "Pre-read — Change Management workshop"),
        (PR + "/Connection_WP_06_Service_Catalog_Request_CLIENT.docx", "Pre-read — Service Catalog & Request workshop"),
        (WS + "/Connection_Incident_Workshop.pptx", "Workshop deck" + " — release AFTER the session", "Post_Workshop"),
        (WS + "/Connection_Major_Incident_Workshop.pptx", "Workshop deck" + " — release AFTER the session", "Post_Workshop"),
        (WS + "/Connection_Problem_Workshop.pptx", "Workshop deck" + " — release AFTER the session", "Post_Workshop"),
        (WS + "/Connection_Change_Workshop.pptx", "Workshop deck — modernized Change + CAB Workbench" + " — release AFTER the session", "Post_Workshop"),
        (WS + "/Connection_Service_Catalog_Workshop.pptx", "Workshop deck" + " — release AFTER the session", "Post_Workshop"),
        (WS + "/Connection_Interactions_Vonage_CTI_Workshop.pptx", "Workshop deck — inbound voice via Vonage CTI" + " — release AFTER the session", "Post_Workshop"),
        (AP + "/ITSM_Accelerator_Pack", "Process pack — incident, MIM, problem, change, request, knowledge decisions"),
        (AP + "/Service_Catalog_Accelerator_Pack", "Data pack — category rationalization, item inventory, approvals"),
     ]),
    ("Drop_04_Employee_Experience_and_Analytics",
     "Sprints 5-6 (Weeks 11-14), release ~1 week before the experience workshops",
     "The employee-experience and analytics build: Employee Center, Virtual Agent, Knowledge, Predictive Intelligence, Performance Analytics, and HAM foundations.",
     [
        (PR + "/Connection_WP_07_Knowledge_Management_CLIENT.docx", "Pre-read — Knowledge Management workshop"),
        (PR + "/Connection_WP_08_Employee_Center_CLIENT.docx", "Pre-read — Employee Center workshop"),
        (PR + "/Connection_WP_09_Virtual_Agent_CLIENT.docx", "Pre-read — Virtual Agent workshop"),
        (PR + "/Connection_WP_10_Predictive_Intelligence_CLIENT.docx", "Pre-read — Predictive Intelligence workshop"),
        (PR + "/Connection_WP_16_Hardware_Asset_Management_CLIENT.docx", "Pre-read — HAM foundations workshop"),
        (WS + "/Connection_Knowledge_Workshop.pptx", "Workshop deck" + " — release AFTER the session", "Post_Workshop"),
        (WS + "/Connection_Employee_Center_Workshop.pptx", "Workshop deck" + " — release AFTER the session", "Post_Workshop"),
        (WS + "/Connection_Virtual_Agent_Workshop.pptx", "Workshop deck — 5 baseline topics + AI Search" + " — release AFTER the session", "Post_Workshop"),
        (WS + "/Connection_Predictive_Intelligence_Workshop.pptx", "Workshop deck" + " — release AFTER the session", "Post_Workshop"),
        (WS + "/Connection_Performance_Analytics_Workshop.pptx", "Workshop deck" + " — release AFTER the session", "Post_Workshop"),
        (WS + "/Connection_HAM_Workshop.pptx", "Workshop deck — stockrooms + HAM foundations" + " — release AFTER the session", "Post_Workshop"),
        (AP + "/Knowledge_Accelerator_Pack", "Data pack — KB structure, article standards, workflow"),
        (AP + "/Employee_Center_Accelerator_Pack", "Data pack — portal design, taxonomy, search"),
        (AP + "/Virtual_Agent_Accelerator_Pack", "Data pack — topic design, handoff, go-live readiness"),
        (AP + "/Predictive_Intelligence_Accelerator_Pack", "Data pack — scope, classification, assignment intelligence"),
        (AP + "/Performance_Analytics_Accelerator_Pack", "Data pack — dashboards, indicators, scorecards"),
        (AP + "/ITAM_HAM_Foundations_Accelerator_Pack", "Data pack — stockrooms, asset classes, lifecycle"),
     ]),
    ("Drop_05_Testing_GoLive_and_Handoff",
     "Stage 3-4 — Sprints 6-8 (Weeks 13-18), release as UAT scheduling begins",
     "Everything Connection needs to test, approve go-live, and take ownership: UAT material, the gated go/no-go checklist, cutover plan, and knowledge-transfer package.",
     [
        ("02_Delivery/Connection_UAT_Guidebook_for_End_Users.docx",
         "How to test, step by step — written for first-time testers"),
        ("02_Delivery/Connection_UAT_End_to_End_Test_Scripts.xlsx",
         "18 end-to-end scripts across 10 suites, traced to stories, with defect log"),
        ("02_Delivery/Connection_Go_Live_Readiness_Checklist.xlsx",
         "The gated go/no-go criteria reviewed jointly before cutover"),
        ("02_Delivery/Connection_Cutover_Runbook.docx",
         "Cutover sequence, validation points, rollback plan"),
        ("02_Delivery/Knowledge_Transfer/Connection_Administrator_Guide_and_KT.docx",
         "Run the platform — admin reference + 4-session KT plan"),
        ("02_Delivery/Knowledge_Transfer/Connection_Train_the_Trainer_Toolkit.docx",
         "Enable Connection trainers ahead of go-live"),
        ("02_Delivery/Connection_Operational_Handoff_Pack.docx",
         "Ownership, support model, and escalation after Hypercare"),
     ]),
]


def copy_entry(rel, drop_dir, subdir=None):
    """Copy a file or a whole folder from SRC into the drop (optionally a subfolder).
    Returns list of copied rel paths."""
    src = os.path.join(SRC, rel.replace("/", os.sep))
    if subdir:
        drop_dir = os.path.join(drop_dir, subdir)
        os.makedirs(drop_dir, exist_ok=True)
    base = os.path.basename(rel.rstrip("/"))
    out = []
    if os.path.isdir(src):
        dest = os.path.join(drop_dir, base)
        os.makedirs(dest, exist_ok=True)
        for f in sorted(os.listdir(src)):
            sf = os.path.join(src, f)
            if os.path.isfile(sf):
                shutil.copy2(sf, os.path.join(dest, f))
                out.append(((subdir + "/") if subdir else "") + base + "/" + f)
    elif os.path.isfile(src):
        shutil.copy2(src, os.path.join(drop_dir, base))
        out.append((subdir + "/" + base) if subdir else base)
    else:
        print(f"  !! MISSING in static library: {rel}")
    return out


def main():
    if not os.path.isdir(SRC):
        raise SystemExit("Static library not found - run export_execution_library.py first.")
    shutil.rmtree(DST, ignore_errors=True)  # best-effort; overwrites in place where delete is blocked
    os.makedirs(DST, exist_ok=True)
    os.makedirs(KIT, exist_ok=True)

    manifest = []  # (drop, when, purpose, file, note)
    for drop, when, purpose, entries in DROPS:
        drop_dir = os.path.join(DST, drop)
        os.makedirs(drop_dir, exist_ok=True)
        n = 0
        for entry in entries:
            rel, note = entry[0], entry[1]
            subdir = entry[2] if len(entry) > 2 else None
            copied = copy_entry(rel, drop_dir, subdir)
            for c in copied:
                manifest.append((drop, when, purpose, c, note))
            n += len(copied)
        # ready-to-send zip per drop
        zpath = os.path.join(DST, drop + ".zip")
        with zipfile.ZipFile(zpath, "w", zipfile.ZIP_DEFLATED) as z:
            for root, dirs, files in os.walk(drop_dir):
                dirs.sort()
                for f in sorted(files):
                    full = os.path.join(root, f)
                    z.write(full, os.path.join(drop, os.path.relpath(full, drop_dir)))
        print(f"{drop}: {n} files + zip")

    # ---- Staged_Rollout_Guide.xlsx ----
    wb = Workbook(); ws = wb.active; ws.title = "Rollout Guide"
    ws["A1"] = "CONNECTION - STAGED DOCUMENT ROLLOUT (JUST-IN-TIME)"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=NAVY)
    ws["A2"] = ("Release each drop when its phase approaches - orientation first, then workshop material ~1 week "
                "ahead of each stage, then testing and handoff. Keeps Connection engaged and prepared without "
                "overwhelming anyone. Regenerate via _tools/build_staged_rollout.py (after the library export).")
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color=SLATE)
    hdr_row = 4
    for c, h in enumerate(["Drop", "When to Send", "File", "Why It's In This Drop", "Released On", "Sent To / Notes"], 1):
        cell = ws.cell(hdr_row, c, h)
        cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY); cell.border = box
    r = hdr_row + 1
    last_drop = None
    for drop, when, purpose, f, note in manifest:
        first_of_drop = drop != last_drop
        if first_of_drop:
            # drop banner row with purpose
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            cell = ws.cell(r, 1, f"{drop.replace('_', ' ')}  -  {when}  -  {purpose}")
            cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
            cell.fill = PatternFill("solid", fgColor=TEAL)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            for c in range(1, 7):
                ws.cell(r, c).border = box
            ws.row_dimensions[r].height = 30
            r += 1
            last_drop = drop; shade = False
        for c, v in enumerate([drop.split("_")[0] + " " + drop.split("_")[1], when.split(" — ")[0].split(" - ")[0], f, note, "", ""], 1):
            cell = ws.cell(r, c, v)
            cell.font = Font(name="Calibri", size=10, color="1A1A1A")
            cell.border = box
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if shade:
                cell.fill = PatternFill("solid", fgColor=ALT)
        shade = not shade
        r += 1
    for col, w in zip("ABCDEF", [10, 16, 52, 46, 12, 22]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    ws.sheet_view.showGridLines = False
    wb.save(os.path.join(KIT, "Staged_Rollout_Guide.xlsx"))
    print(f"Guide + {len(manifest)} manifest rows -> {KIT}")


if __name__ == "__main__":
    main()
