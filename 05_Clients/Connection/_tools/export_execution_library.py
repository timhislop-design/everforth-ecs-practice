# -*- coding: utf-8 -*-
"""TRIGGER: convert the working engagement framework into a static, upload-ready library.
Copies finalized artifacts (docx/pptx/xlsx) only - no .py/.js/.json/.md, no source inputs.
Front door = 00_START_HERE/Connection_Execution_Guide.docx (role + phase navigator).
Usage: python export_execution_library.py
"""
import os, shutil, subprocess, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
REPO="/sessions/eager-epic-davinci/mnt/everforth-ecs-practice"
SRC=os.path.join(REPO,"05_Clients","Connection")
DST=os.path.join(REPO,"06_Client_Upload","Connection")
ARTIFACT_EXT={".docx",".pptx",".xlsx"}
EXCLUDE_DIRS={"00_Source_Inputs","_tools","03_Internal","__pycache__"}
NAVY="0B1F3A";WHITE="FFFFFF";ALT="F8FAFC";BORDER="E2E8F0";SLATE="475569"
thin=Side(style="thin",color=BORDER);box=Border(left=thin,right=thin,top=thin,bottom=thin)
def audience_for(rel):
    p=rel.replace("\\","/").lower()
    if "client_facing" in p: return "Client-facing"
    if "internal_team" in p: return "ECS internal"
    if "knowledge_transfer" in p: return "Client-facing (KT)"
    if "/demo_scripts/" in p: return "ECS internal"
    if "/pre_reads/" in p: return "Client-facing"
    if "/workshops/" in p: return "Client-facing"
    if "/accelerator_packs/" in p: return "Client (data)"
    if "01_onboarding/" in p: return "Both"
    return "Delivery / Both"
def main():
    if os.path.exists(DST): shutil.rmtree(DST, ignore_errors=True)
    os.makedirs(DST, exist_ok=True)
    copied=[]
    for root,dirs,files in os.walk(SRC):
        dirs[:]=[d for d in dirs if d not in EXCLUDE_DIRS]
        for f in files:
            if os.path.splitext(f)[1].lower() not in ARTIFACT_EXT: continue
            rel=os.path.relpath(os.path.join(root,f), SRC)
            dest=os.path.join(DST, rel); os.makedirs(os.path.dirname(dest), exist_ok=True)
            shutil.copy2(os.path.join(root,f), dest); copied.append(rel)
    subprocess.run([sys.executable, os.path.join(SRC,"_tools","build_execution_guide.py")], check=True)
    nav=os.path.join(SRC,"_tools","Connection_Execution_Navigator.html")
    if os.path.exists(nav): shutil.copy2(nav, os.path.join(DST,"00_START_HERE","Connection_Execution_Navigator.html"))
    wb=Workbook();ws=wb.active;ws.title="Library Index"
    ws["A1"]="CONNECTION EXECUTION LIBRARY - INDEX"; ws["A1"].font=Font(name="Calibri",size=14,bold=True,color=NAVY)
    ws["A2"]=f"Auto-generated catalog of {len(copied)} artifacts. Static upload-ready copy of the ECS working framework (artifacts only). Start with the Execution Guide (role + phase)."
    ws["A2"].font=Font(name="Calibri",size=9,italic=True,color=SLATE)
    for c,h in enumerate(["#","Folder","Artifact","Type","Audience"],1):
        cell=ws.cell(4,c,h);cell.font=Font(name="Calibri",size=11,bold=True,color=WHITE);cell.fill=PatternFill("solid",fgColor=NAVY);cell.border=box
    r=5
    for i,rel in enumerate(sorted(copied),1):
        folder=os.path.dirname(rel) or "(root)"; name=os.path.basename(rel); ext=os.path.splitext(name)[1].lstrip(".")
        for c,v in enumerate([i,folder,name,ext,audience_for(rel)],1):
            cell=ws.cell(r,c,v);cell.font=Font(name="Calibri",size=10,color="1A1A1A");cell.border=box;cell.alignment=Alignment(wrap_text=True,vertical="top")
            if i%2==0: cell.fill=PatternFill("solid",fgColor=ALT)
        r+=1
    for col,w in zip("ABCDE",[5,40,52,8,18]): ws.column_dimensions[col].width=w
    ws.freeze_panes="A5"; ws.sheet_view.showGridLines=False
    wb.save(os.path.join(DST,"00_START_HERE","Connection_Library_Index.xlsx"))
    print(f"Exported {len(copied)} artifacts to {DST}")
if __name__=="__main__": main()
