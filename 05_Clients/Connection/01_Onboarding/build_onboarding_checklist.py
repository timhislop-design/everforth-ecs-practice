# -*- coding: utf-8 -*-
"""Build: Connection - Onboarding & Sprint 0 Readiness Tracker (xlsx).
Both sides (ECS + Connection). Brand-styled; status dropdown + live summary."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

OUT = "/sessions/eager-epic-davinci/mnt/everforth-ecs-practice/05_Clients/Connection/01_Onboarding/Connection_Onboarding_Checklist.xlsx"

NAVY="0B1F3A"; TEAL="14B8A6"; WHITE="FFFFFF"; ALT="F8FAFC"; BORDER="E2E8F0"; SLATE="475569"
GREEN="DCFCE7"; YELLOW="FEF9C3"; BLUE="DBEAFE"; RED="FEE2E2"

wb = Workbook(); ws = wb.active; ws.title = "Onboarding Tracker"
thin = Side(style="thin", color=BORDER)
box = Border(left=thin, right=thin, top=thin, bottom=thin)

# Title
ws["A1"] = "CONNECTION ENGAGEMENT - ONBOARDING & SPRINT 0 READINESS TRACKER"
ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=NAVY)
ws["A2"] = "Shared - ECS Team + Connection   |   Update Status as items complete; any open item at end of Sprint 0 triggers a dependency-slip conversation with the Sponsor."
ws["A2"].font = Font(name="Calibri", size=9, italic=True, color=SLATE)

headers = ["ID", "Task", "Side", "Owner (role)", "Stage / Sprint", "Status", "Due", "Notes"]
hr = 4
for c, h in enumerate(headers, 1):
    cell = ws.cell(hr, c, h)
    cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = box

rows = [
 # ECS internal
 ["Internal ECS kickoff - SOW review & consultant assignment", "ECS", "EM", "Sprint 0", "Not Started", "", ""],
 ["Team reads onboarding package (Vision, Guidelines, Role Quick-Ref)", "ECS", "All ECS", "Sprint 0", "Not Started", "", ""],
 ["Confirm role assignments & decision rights", "ECS", "EM", "Sprint 0", "Not Started", "", ""],
 ["Stand up backlog tooling + story template", "ECS", "PC", "Sprint 0", "Not Started", "", ""],
 ["Prepare Accelerator Pack workbooks for distribution", "ECS", "SA / PC", "Sprint 0", "Not Started", "", ""],
 ["Establish weekly health reporting to practice management", "ECS", "EM -> Practice Lead", "Sprint 0", "Not Started", "", ""],
 # Connection customer readiness
 ["Project Sponsor identified (avail. for Council + bi-weekly sync)", "Connection", "Connection", "Sprint 0", "Not Started", "", ""],
 ["ServiceNow instances provisioned (sub-prod + prod) w/ admin access", "Connection", "Tech Lead", "Sprint 0", "Not Started", "", ""],
 ["Access & credentials for ECS team", "Connection", "Tech Lead", "Sprint 0", "Not Started", "", ""],
 ["Foundation Data Pack completed (users, locations, groups, SLAs)", "Connection", "Connection SMEs", "Sprint 0-1", "Not Started", "", ""],
 ["Stakeholder & SME mapping completed", "Connection", "Connection PM", "Sprint 0", "Not Started", "", ""],
 # Joint / governance
 ["Customer kickoff meeting", "Joint", "EM + Sponsor", "Sprint 0", "Not Started", "", ""],
 ["CSDM reference model selected & pre-loaded for Sprint 1", "Joint", "SA + Connection", "Sprint 0", "Not Started", "", ""],
 ["Definition of Done published & acknowledged by Product Owner", "Joint", "EM + Product Owner", "Sprint 0", "Not Started", "", ""],
 ["Governance & decision-rights workshop", "Joint", "EM + Sponsor", "Sprint 0", "Not Started", "", ""],
 ["Customization Council charter signed", "Joint", "Sponsor + EM", "Sprint 0", "Not Started", "", ""],
 ["Communication plan & cadence established", "Joint", "EM + Connection PM", "Sprint 0", "Not Started", "", ""],
 ["Risk register initialized", "Joint", "EM + Connection PM", "Sprint 0", "Not Started", "", ""],
 ["Sprint 0 readiness validation (ECS checklist)", "Joint", "EM", "Sprint 0", "Not Started", "", ""],
]

r = hr + 1
for i, row in enumerate(rows):
    rid = f"OB-{i+1:02d}"
    vals = [rid] + row
    shade = (i % 2 == 0)
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v)
        cell.font = Font(name="Calibri", size=10, color="1A1A1A")
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = box
        if shade:
            cell.fill = PatternFill("solid", fgColor=ALT)
    r += 1
last = r - 1

# Status dropdown
dv = DataValidation(type="list", formula1='"Not Started,In Progress,Complete,Blocked"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"F{hr+1}:F{last}")

# Conditional formatting on Status (col F)
fmt = {"Complete": GREEN, "In Progress": BLUE, "Blocked": RED, "Not Started": YELLOW}
for val, color in fmt.items():
    ws.conditional_formatting.add(f"F{hr+1}:F{last}",
        CellIsRule(operator="equal", formula=[f'"{val}"'], fill=PatternFill("solid", fgColor=color)))

# Summary block
sr = last + 2
ws.cell(sr, 1, "SUMMARY").font = Font(name="Calibri", size=11, bold=True, color=NAVY)
summ = [("Total tasks", f"=COUNTA(B{hr+1}:B{last})"),
        ("Complete", f'=COUNTIF(F{hr+1}:F{last},"Complete")'),
        ("In Progress", f'=COUNTIF(F{hr+1}:F{last},"In Progress")'),
        ("Blocked", f'=COUNTIF(F{hr+1}:F{last},"Blocked")'),
        ("Not Started", f'=COUNTIF(F{hr+1}:F{last},"Not Started")')]
for i, (lbl, f) in enumerate(summ):
    rr = sr + 1 + i
    ws.cell(rr, 1, lbl).font = Font(name="Calibri", size=10, color=SLATE)
    cc = ws.cell(rr, 2, f); cc.font = Font(name="Calibri", size=10, bold=True, color=NAVY)

widths = [7, 52, 12, 20, 14, 14, 12, 30]
for c, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64+c)].width = w
ws.freeze_panes = f"A{hr+1}"
ws.sheet_view.showGridLines = False

wb.save(OUT)
print("Saved:", OUT)
