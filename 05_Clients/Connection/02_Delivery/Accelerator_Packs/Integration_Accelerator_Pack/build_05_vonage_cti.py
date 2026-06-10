# -*- coding: utf-8 -*-
"""Build: Connection Vonage CTI & Interactions Accelerator Pack (Phase 1 - inbound voice).
Matches the Integration pack 8-tab format (SCCM) using accelerator_pack_builder styling primitives."""
import sys, os
REPO="/sessions/eager-epic-davinci/mnt/everforth-ecs-practice"
sys.path.insert(0, os.path.join(REPO,"03_Shared","00_Templates_and_Branding"))
import accelerator_pack_builder as B
from openpyxl import Workbook
from openpyxl.styles import Alignment

OUT=os.path.join(REPO,"05_Clients","Connection","02_Delivery","Accelerator_Packs","Integration_Accelerator_Pack","05_vonage_cti_interactions.xlsx")
F=B._font; FILL=B._fill
def body_block(ws, start, rows, label_col=1, fill_cols=None, widths=None):
    r=start
    for i,row in enumerate(rows):
        B.set_body_row(ws,r,row,label_col=label_col,customer_fill_cols=fill_cols or [],alt_shade=(i%2==1))
        ws.row_dimensions[r].height=max(26,14+5*(max((len(str(x)) for x in row),default=20)//35))
        r+=1
    if widths: B.set_columns(ws,widths)
    return r

wb=Workbook(); wb.remove(wb.active)
TITLE="05 — Vonage CTI & Interactions"; PACK="Integration Accelerator Pack"

# 1 Instructions
ws=wb.create_sheet("Instructions")
B.set_banner(ws,1,TITLE,span_cols=2); B.set_banner(ws,2,f"{PACK} · ECS Federal · ServiceNow Practice",subtitle=True,span_cols=2)
rows=[("Purpose","Capture requirements, dependencies, configuration, and developer instructions to stand up the Vonage telephony (CTI) integration through ServiceNow Interactions. Phase 1 scope is inbound voice only (chat/email deferred)."),
("Who fills this out","Customer-side: Vonage Contact Center administrator + service desk lead. ECS-side: Solution Architect / senior developer."),
("Sprint window","Connection + tenant details by end of Sprint 1; build alongside Incident/Request in Stage 2 (Sprints 3-4); validate before Go-Live."),
("Estimated effort","~8-12 hours across Sprints 1-4, including porting the legacy Vonage config and test calls."),
("Where the data comes from","Vonage Contact Center admin (tenant, queues, agents) and the existing Vonage-ServiceNow integration on the legacy instance (for porting)."),
("Related workbooks","01 Active Directory · 02 Single Sign-On · 03 SCCM · 04 Intune")]
r=4
for lbl,val in rows:
    ws.cell(r,1,lbl).font=F(color=B.APColor.NAVY,bold=True)
    c=ws.cell(r,2,val); c.font=F(color=B.APColor.SLATE); c.alignment=Alignment(wrap_text=True,vertical="top")
    ws.row_dimensions[r].height=max(24,16+6*(len(val)//70)); r+=1
r+=1; B.set_section_header(ws,r,"Success criteria",span_cols=2); r+=1
for crit in ["Inbound Vonage call opens/creates an Interaction with caller context (screen-pop).",
 "Agent creates or links an Incident/Request from the Interaction (OOTB).",
 "Vonage queues map to ServiceNow assignment groups for routing context.",
 "No custom telephony middleware - OOTB OpenFrame + Vonage connector only.",
 "Parity with the legacy phone experience validated before Go-Live."]:
    ws.cell(r,1,"•").font=F(color=B.APColor.NAVY,bold=True); ws.cell(r,2,crit).font=F(color=B.APColor.SLATE)
    ws.cell(r,2).alignment=Alignment(wrap_text=True,vertical="top"); ws.row_dimensions[r].height=22; r+=1
B.set_columns(ws,[24,96])

# 2 Requirements
ws=wb.create_sheet("Requirements")
B.set_banner(ws,1,"Vonage CTI — Integration Requirements",span_cols=4); B.set_banner(ws,2,"Customer fills the Answer column. Phase 1 = inbound voice only.",subtitle=True,span_cols=4)
B.set_header_row(ws,4,["#","Requirement Question","Answer","Notes"])
reqs=[["Which Vonage product is in use (Vonage Contact Center / VCC, or VBC)? Region/data center?","",""],
["Approximate number of agents and peak concurrent inbound calls in scope?","",""],
["Confirm Phase 1 channels = inbound voice only (chat, email, SMS deferred)?","",""],
["CTI mode: OpenFrame softphone embedded in the Agent / Service Operations Workspace?","",""],
["Screen-pop behavior on inbound call: open/create an Interaction with caller match?","",""],
["Caller identification: match on sys_user phone/mobile; behavior when no match (guest)?","",""],
["From the Interaction, agents create or link an Incident and/or Request (RITM)?","",""],
["Vonage queues/skills - how should they map to ServiceNow assignment groups?","",""],
["Is click-to-dial (outbound) in scope for Phase 1, or inbound only?","",""],
["Call recording in scope? Where stored (Vonage vs. link in ServiceNow)? Retention?","",""],
["Compliance constraints (call recording consent, PII, PCI on IVR)?","",""],
["Is there an existing Vonage-ServiceNow integration on the legacy instance to port?","",""]]
r=5
for i,row in enumerate(reqs,1):
    B.set_body_row(ws,r,[i]+row,customer_fill_cols=[3],alt_shade=(i%2==0))
    ws.row_dimensions[r].height=max(26,14+5*(len(row[0])//35)); r+=1
B.set_columns(ws,[4,64,30,34]); B.freeze_top(ws,4)

# 3 Dependencies
ws=wb.create_sheet("Dependencies")
B.set_banner(ws,1,"Vonage CTI — Dependencies & Prerequisites",span_cols=6); B.set_banner(ws,2,"Confirm before integration build begins.",subtitle=True,span_cols=6)
B.set_header_row(ws,4,["#","Dependency","Status","Owner","Due","Notes"])
deps=[["ITSM Core (Incident/Request) available in sub-production","In Progress","ECS","","Interactions link to incident / sc_req_item"],
["Interaction capability + Agent/Service Operations Workspace available","Not Started","ECS","",""],
["OpenFrame plugin (com.glide.openframe) activated","Not Started","ECS","",""],
["Vonage CTI / OpenFrame adapter installed (Vonage / ServiceNow Store)","Not Started","ECS","",""],
["Vonage tenant admin access + API credentials","Not Started","Customer","","Vaulted in ServiceNow Credentials by ECS"],
["Agent roles assigned (sn_openframe_user / interaction roles)","Not Started","ECS","",""],
["SSO configured for agents (see 02 Single Sign-On)","In Progress","ECS","",""],
["Network allowlist for Vonage CTI domain from agent browsers","Not Started","Customer","",""],
["Test phone number + test agent provisioned","Not Started","Customer","",""],
["Legacy Vonage integration inventory captured (for porting)","Not Started","Customer","","See Port from Legacy tab"]]
r=5
for i,row in enumerate(deps,1):
    B.set_body_row(ws,r,[i]+row,alt_shade=(i%2==0)); ws.row_dimensions[r].height=26; r+=1
B.set_columns(ws,[4,52,12,12,10,40]); B.freeze_top(ws,4)

# 4 Configuration Data
ws=wb.create_sheet("Configuration Data")
B.set_banner(ws,1,"Vonage CTI — Configuration Data",span_cols=3); B.set_banner(ws,2,"Customer fills the Value column. Yellow examples = replace with your values.",subtitle=True,span_cols=3)
B.set_header_row(ws,4,["Field","Value","Notes"])
sections=[("1. Vonage Tenant",[("Vonage product / edition","Vonage Contact Center (VCC)","VBC if business comms only"),
 ("Region / data center","US","Match agent location"),
 ("API key / application ID","[customer]","Vaulted by ECS"),
 ("Credential reference","[ECS will populate]","ServiceNow Credential store")]),
("2. OpenFrame",[("OpenFrame config name","Vonage CTI","Appears in Agent Workspace"),
 ("Adapter URL","https://<vonage-cti-adapter>","From Vonage/Store adapter"),
 ("Frame width x height","400 x 600","px"),
 ("Roles with access","sn_openframe_user","Assigned to service desk agents")]),
("3. Agent Mapping",[("Agent -> Vonage extension/agent ID","[customer list]","One row per agent at build"),
 ("Default agent state on login","Available","")]),
("4. Queue / Routing Mapping",[("Vonage queue/skill -> assignment group","[customer mapping]","Drives routing context on the Interaction"),
 ("Unknown-caller handling","Open Interaction as guest","No sys_user match")]),
("5. Recording",[("Recording in scope (Y/N)","[customer]",""),
 ("Storage / link location","Vonage; link on Interaction","Avoid storing media in ServiceNow"),
 ("Retention","[customer policy]","Compliance-driven")])]
r=5
for title,fields in sections:
    B.set_section_header(ws,r,title,span_cols=3); r+=1
    for i,(fld,val,note) in enumerate(fields):
        cust=val.startswith("[customer") or "customer" in val.lower()
        B.set_body_row(ws,r,[fld,val,note],customer_fill_cols=[2] if cust else [],alt_shade=(i%2==1)); ws.row_dimensions[r].height=24; r+=1
B.set_columns(ws,[40,34,44]); B.freeze_top(ws,4)

# 5 R&R
ws=wb.create_sheet("R&R")
B.set_banner(ws,1,"Vonage CTI — Roles & Responsibilities",span_cols=4); B.set_banner(ws,2,"R=Responsible, A=Accountable, C=Consulted, I=Informed.",subtitle=True,span_cols=4)
B.set_header_row(ws,4,["Activity","ECS","Customer","Notes"])
raci=[["Vonage tenant configuration (queues, agents, numbers)","C","R","Customer's Vonage admin owns; ECS specifies needs"],
["Vonage API credentials creation","C","R","Customer provides; ECS vaults"],
["OpenFrame configuration in ServiceNow","R","C","ECS-managed"],
["Vonage CTI adapter install from Store","R","I","ECS-managed"],
["Agent provisioning + extension mapping","C","R","Customer maps agents to extensions"],
["Queue -> assignment group mapping","R","C","ECS configures; Customer confirms routing"],
["Interaction -> Incident/Request flow configuration","R","C","ECS configures OOTB Interaction behavior"],
["Network allowlist for the CTI domain","I","R","Customer's network team"],
["Call recording compliance / consent","I","A","Customer owns the policy decision"],
["Test calls + UAT","R","C","ECS executes; Customer validates"],
["Long-term ownership of CTI operations","I","R","Customer owns post-Hypercare"]]
body_block(ws,5,raci,widths=[46,8,10,44])

# 6 Developer Notes
ws=wb.create_sheet("Developer Notes")
B.set_banner(ws,1,"Vonage CTI — Developer Notes",span_cols=1); B.set_banner(ws,2,"Internal reference for the ECS Solution Architect / senior developer.",subtitle=True,span_cols=1)
r=4
def note_section(ws,r,head,lines):
    B.set_section_header(ws,r,head,span_cols=1); r+=1
    for ln in lines:
        c=ws.cell(r,1,("•  "+ln) if not ln.startswith(("Use","Phase","Interactions")) else ln)
        c.font=F(color=B.APColor.SLATE); c.alignment=Alignment(wrap_text=True,vertical="top"); ws.row_dimensions[r].height=max(22,16+6*(len(ln)//90)); r+=1
    return r
r=note_section(ws,r,"OOTB-First Approach",[
 "Use OOTB OpenFrame + the Vonage CTI adapter + the OOTB Interaction record. No custom telephony middleware or scripted softphone.",
 "Interactions is the OOTB omnichannel entry point; Phase 1 implements the voice channel only.",
 "Caller matching and routing context belong in OOTB CTI config / Decision Tables - not Script Includes."])
r+=1
r=note_section(ws,r,"Implementation Pattern",[
 "Activate OpenFrame (com.glide.openframe) and the Interaction features; confirm Agent/SO Workspace.",
 "Install the Vonage CTI/OpenFrame adapter (Vonage-provided or ServiceNow Store).",
 "Configure OpenFrame: name, adapter URL, frame size, and roles (sn_openframe_user) per the Configuration Data tab.",
 "Configure the Vonage connection (API key, region); vault credentials in the Credential store.",
 "Map agents to Vonage extensions; assign OpenFrame + interaction roles.",
 "Configure inbound screen-pop: on call, create/open an Interaction and match the caller by phone (caller number -> sys_user.phone/mobile_phone).",
 "Configure Interaction -> Incident/Request: agents use the OOTB Create Incident / Create Request action; interaction_related_record links them.",
 "Map Vonage queues/skills to assignment groups for routing context.",
 "Validate in sub-prod with a test number + agent: screen-pop, Interaction creation, record linkage, and (if in scope) the recording link.",
 "Promote via update set; re-point the CTI connection to the production Vonage tenant."])
r+=1
r=note_section(ws,r,"Guardrails",[
 "Any custom click-to-dial or screen-pop logic beyond OOTB CTI capability is a customization - route it to the Customization Council.",
 "Do not persist call media in ServiceNow; store in Vonage and link from the Interaction."])
B.set_columns(ws,[120])

# 7 Port from Legacy
ws=wb.create_sheet("Port from Legacy")
B.set_banner(ws,1,"Vonage CTI — Port from Legacy Instance",span_cols=3); B.set_banner(ws,2,"Use the existing Vonage integration as the spec; rebuild on OOTB OpenFrame + connector.",subtitle=True,span_cols=3)
r=4; B.set_section_header(ws,r,"Porting Approach",span_cols=3); r+=1
for ln in ["The existing Vonage integration on the legacy domain-separated instance is the reference. Take the existing setup/config and use it to build the new integration - like SCCM, the OOTB connector replaces bespoke pipelines. Lift the configuration and intent, not the custom code."]:
    ws.cell(r,1,ln).font=F(color=B.APColor.SLATE); ws.cell(r,1).alignment=Alignment(wrap_text=True,vertical="top"); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3); ws.row_dimensions[r].height=42; r+=1
B.set_header_row(ws,r,["#","Step","Disposition / Notes"]); r+=1
steps=[["Inventory the legacy Vonage integration: OpenFrame config, adapter URL, API creds, click-to-dial and screen-pop scripts, call-logging business rules.","Customer + ECS"],
["Capture the legacy caller-match logic and the screen-pop target (what record opened, what fields populated).","Reference for OOTB config"],
["Export Vonage tenant settings: queues/skills, numbers, agent roster, business hours.","Reuse as-is"],
["Categorize each item: OpenFrame/connector config = rebuild OOTB; custom screen-pop = evaluate vs OOTB CTI; bespoke middleware = Council.","Most = config, not code"],
["Rebuild on the new instance using OOTB OpenFrame + Vonage adapter; apply the exported tenant settings.","Do NOT port custom pipelines"],
["Carry forward only customizations with a documented business need, via the two-key decision.","Logged in Triage Log"],
["Validate parity: inbound call -> screen-pop -> Interaction -> Incident/Request matches the legacy experience.","UAT sign-off"]]
for i,row in enumerate(steps,1):
    B.set_body_row(ws,r,[i]+row,alt_shade=(i%2==0)); ws.row_dimensions[r].height=max(28,14+6*(len(row[0])//50)); r+=1
B.set_columns(ws,[4,72,34])

# 8 ServiceNow Mapping
ws=wb.create_sheet("ServiceNow Mapping")
B.set_banner(ws,1,"ServiceNow Mapping (ECS Solution Architect reference)",span_cols=3); B.set_banner(ws,2,"OOTB tables, fields, and features used.",subtitle=True,span_cols=3)
B.set_header_row(ws,4,["Element","Table / Object","Notes"])
maps=[["Interaction record","interaction","The entry point created on an inbound call"],
["Interaction <-> task link","interaction_related_record","Links the Interaction to incident / sc_req_item"],
["Incident","incident","Created/linked from the Interaction"],
["Request item","sc_req_item","Created/linked from the Interaction"],
["OpenFrame configuration","sys_cti_openframe (OpenFrame config)","Hosts the Vonage softphone in the workspace"],
["Agent","sys_user + sn_openframe_user role","CTI-enabled service desk agents"],
["Caller match","sys_user.phone / mobile_phone","Screen-pop identification"],
["Assignment routing","sys_user_group","Vonage queue/skill -> assignment group"]]
body_block(ws,5,maps,widths=[34,40,44])

wb.save(OUT); print("Saved:",OUT,"tabs:",wb.sheetnames)
