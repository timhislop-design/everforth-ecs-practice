# -*- coding: utf-8 -*-
import sys; sys.path.insert(0,"/tmp")
import dd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
OUT="/sessions/eager-epic-davinci/mnt/everforth-ecs-practice/05_Clients/Connection/02_Delivery/Connection_Project_Delivery_Stories.xlsx"
NAVY="0B1F3A";WHITE="FFFFFF";ALT="F8FAFC";BORDER="E2E8F0";SLATE="475569";CYAN="ECFEFF"
thin=Side(style="thin",color=BORDER);box=Border(left=thin,right=thin,top=thin,bottom=thin)
def bf(c="1A1A1A",b=False,sz=10): return Font(name="Calibri",size=sz,bold=b,color=c)
seq={}
for s in dd.STORIES:
    ab=dd.EPICS[s["epic"]]; seq[ab]=seq.get(ab,0)+1
    s["number"]=f"CONN-{ab}-{seq[ab]:03d}"
    s["description"]=f"As a {s['role']}, I want {s['want']}, so that {s['benefit']}."
    s["acceptance_criteria"]="\n".join(f"AC{i}: {a}" for i,a in enumerate(s["ac"],1))
wb=Workbook(); wb.remove(wb.active)
def titlebar(ws,t,sub):
    ws["A1"]=t; ws["A1"].font=Font(name="Calibri",size=14,bold=True,color=NAVY)
    ws["A2"]=sub; ws["A2"].font=Font(name="Calibri",size=9,italic=True,color=SLATE)
def hrow(ws,cols,r=4):
    for c,h in enumerate(cols,1):
        cell=ws.cell(r,c,h); cell.font=Font(name="Calibri",size=11,bold=True,color=WHITE); cell.fill=PatternFill("solid",fgColor=NAVY)
        cell.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True); cell.border=box
    ws.row_dimensions[r].height=26

# README
ws=wb.create_sheet("README - Import Guide"); r=4
titlebar(ws,"CONNECTION - PROJECT DELIVERY STORIES (the work most projects don't track)",
 "Every non-development activity as a story: documentation, training, governance, testing, go-live, hypercare. Same rigor as config stories. Complements the 141-story config backlog -> together = the complete project backlog.")
def sec(t):
    global r; c=ws.cell(r,1,t); c.font=Font(name="Calibri",size=11,bold=True,color=NAVY); c.fill=PatternFill("solid",fgColor=CYAN); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2); r+=1
def ln(a,b):
    global r; ws.cell(r,1,a).font=bf(b=True); c=ws.cell(r,2,b); c.font=bf(); c.alignment=Alignment(wrap_text=True,vertical="top"); ws.row_dimensions[r].height=max(20,14+6*(len(b)//90)); r+=1
sec("Why this exists")
ln("The gap","Most projects only story development. Documentation, training, governance, and PMO work get done but stay invisible to leadership - so burndown shows 'dev on track' while the rest of the team's effort is uncredited and unmanaged.")
ln("The fix","Story ALL the work, same rigor (acceptance criteria + Definition of Done). Leadership sees the whole project in real time and the team's performance shows in the numbers.")
ln("Pairs with","Connection_User_Stories_SN_Agile.xlsx (141 configuration stories). Import both -> one backlog, full picture.")
r+=1; sec("How to import (ServiceNow Agile / rm_story)")
for stp in ["1. Import EPICS first ('Epics' tab) into rm_epic (coalesce on Number).","2. Import 'Delivery Stories' into rm_story via Import Set + Transform Map.","3. Map columns to fields (same mapping as the config workbook; see that README).","4. Set State on import; refine points/owner in ServiceNow."]:
    ws.cell(r,1,stp).font=bf(); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2); ws.cell(r,1).alignment=Alignment(wrap_text=True); ws.row_dimensions[r].height=20; r+=1
ws.column_dimensions["A"].width=22; ws.column_dimensions["B"].width=104; ws.sheet_view.showGridLines=False

# Epics
ws=wb.create_sheet("Epics")
titlebar(ws,"CONNECTION - DELIVERY EPICS (work-streams)","One epic per work-stream. Import first -> rm_epic.")
hrow(ws,["number","short_description","description","state"])
r=5; en=0
for epic,ab in dd.EPICS.items():
    en+=1; cnt=sum(1 for s in dd.STORIES if s["epic"]==epic)
    vals=[f"CONN-DEPIC-{ab}",f"{epic} (Delivery)",f"Project delivery work-stream: {epic}. {cnt} stories spanning the engagement.","Draft"]
    for c,v in enumerate(vals,1):
        cell=ws.cell(r,c,v); cell.font=bf(); cell.alignment=Alignment(wrap_text=True,vertical="top"); cell.border=box
        if en%2==1: cell.fill=PatternFill("solid",fgColor=ALT)
    ws.row_dimensions[r].height=30; r+=1
for col,w in zip("ABCD",[18,34,84,10]): ws.column_dimensions[col].width=w
ws.freeze_panes="A5"; ws.sheet_view.showGridLines=False

# Delivery Stories
ws=wb.create_sheet("Delivery Stories")
titlebar(ws,"CONNECTION - PROJECT DELIVERY STORIES","Non-development work, storied with acceptance criteria + DoD + traceability to the SOW deliverable/artifact.")
cols=["number","epic","work type","short_description","description","acceptance_criteria","definition_of_done","story_points","priority","sprint","owner (role)","traceability","state"]
hrow(ws,cols)
r=5
for i,s in enumerate(dd.STORIES):
    vals=[s["number"],s["epic"],s["wtype"],s["short"],s["description"],s["acceptance_criteria"],"See 'Definition of Done' tab",s["points"],s["priority"],s["sprint"],s["owner"],s["trace"],"Draft"]
    for c,v in enumerate(vals,1):
        cell=ws.cell(r,c,v); cell.font=bf(); cell.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True); cell.border=box
        if i%2==1: cell.fill=PatternFill("solid",fgColor=ALT)
    ws.row_dimensions[r].height=max(58,14+11*(len(s["acceptance_criteria"])//70)); r+=1
last=r-1
dv=DataValidation(type="list",formula1='"Draft,Ready,Work in progress,Complete,Cancelled"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"M5:M{last}")
for col,w in zip("ABCDEFGHIJKLM",[14,26,13,38,50,68,22,8,14,10,22,34,11]): ws.column_dimensions[col].width=w
ws.freeze_panes="A5"; ws.sheet_view.showGridLines=False

# Definition of Done (delivery)
ws=wb.create_sheet("Definition of Done")
titlebar(ws,"DEFINITION OF DONE - delivery work","A delivery story is Done only when all gates pass.")
r=4
for g in ["Produced using the approved ECS brand template/standard (EcsDocument / pptx_brand / pack builder).",
 "Content complete and accurate to the SOW and the agreed decisions.",
 "Peer or lead reviewed (Practice Lead for high-stakes items).",
 "Reviewed and accepted/signed off by the named party (Sponsor / Product Owner / Technical Lead as applicable).",
 "Stored in the engagement folder (05_Clients/Connection) and referenced in the SOW Deliverables Matrix.",
 "Reflected in the weekly status report and the Executive Health Dashboard.",
 "For client-facing items: correct footer (Confidential) and no internal-only content leaked.",
 "For workshops/sessions: signed-off output captured; deviations logged in the Triage Log."]:
    ws.cell(r,1,"DoD").font=bf(b=True,c=NAVY); c=ws.cell(r,2,g); c.font=bf(); c.alignment=Alignment(wrap_text=True,vertical="top"); c.border=box; ws.cell(r,1).border=box
    ws.row_dimensions[r].height=max(22,14+6*(len(g)//80)); r+=1
ws.column_dimensions["A"].width=8; ws.column_dimensions["B"].width=110; ws.sheet_view.showGridLines=False

# Workstream Summary
ws=wb.create_sheet("Workstream Summary")
titlebar(ws,"WORK-STREAM SUMMARY","Story counts and points per work-stream (live COUNTIF/SUMIF over the Delivery Stories tab).")
hrow(ws,["Work-stream","Stories","Story Points"])
r=5; ds="'Delivery Stories'"
for epic in dd.EPICS:
    ws.cell(r,1,epic).font=bf(); ws.cell(r,1).border=box
    ws.cell(r,2,f'=COUNTIF({ds}!B:B,A{r})').font=bf(b=True,c=NAVY); ws.cell(r,2).border=box
    ws.cell(r,3,f'=SUMIF({ds}!B:B,A{r},{ds}!H:H)').font=bf(b=True,c=NAVY); ws.cell(r,3).border=box
    if r%2==0:
        for c in range(1,4): ws.cell(r,c).fill=PatternFill("solid",fgColor=ALT)
    r+=1
ws.cell(r,1,"TOTAL").font=bf(b=True,c=NAVY); ws.cell(r,2,f"=SUM(B5:B{r-1})").font=bf(b=True,c=NAVY); ws.cell(r,3,f"=SUM(C5:C{r-1})").font=bf(b=True,c=NAVY)
for col,w in zip("ABC",[34,12,14]): ws.column_dimensions[col].width=w
ws.sheet_view.showGridLines=False

order=["README - Import Guide","Epics","Delivery Stories","Definition of Done","Workstream Summary"]
wb._sheets.sort(key=lambda x: order.index(x.title) if x.title in order else 99)
wb.save(OUT)
print(f"Saved: {OUT} | epics={en} | delivery stories={len(dd.STORIES)}")
