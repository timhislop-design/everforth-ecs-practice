# -*- coding: utf-8 -*-
import sys; sys.path.insert(0,"/sessions/eager-epic-davinci/mnt/everforth-ecs-practice/05_Clients/Connection/02_Delivery")
import userstories_data as sd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
OUT="/sessions/eager-epic-davinci/mnt/everforth-ecs-practice/05_Clients/Connection/02_Delivery/Connection_UAT_End_to_End_Test_Scripts.xlsx"
NAVY="0B1F3A";WHITE="FFFFFF";ALT="F8FAFC";BORDER="E2E8F0";SLATE="475569";CYAN="ECFEFF";AMBER="FEF3C7";GREEN="DCFCE7";RED="FEE2E2";BLUE="DBEAFE"
thin=Side(style="thin",color=BORDER);box=Border(left=thin,right=thin,top=thin,bottom=thin)
def bf(c="1A1A1A",b=False,sz=10): return Font(name="Calibri",size=sz,bold=b,color=c)
# recompute story IDs + title map
seq={}; TITLE={}; ORDER=[]
for s in sd.STORIES:
    ab=sd.MODMETA[s["module"]]["abbr"]; seq[ab]=seq.get(ab,0)+1
    sid=f"CONN-{ab}-{seq[ab]:03d}"; s["number"]=sid; TITLE[sid]=(s["module"],s["short"]); ORDER.append(sid)

def steps(*a): return "\n".join(f"{i}. {x}" for i,x in enumerate(a,1))
SCRIPTS=[
 dict(suite="1. Service Desk - Incident Lifecycle",sid="UAT-INC-01",name="Report and resolve an issue via Employee Center",
  who="Business Tester (Service Desk User + Agent)",pre="User logged into Employee Center; an agent account available; KB seeded",
  st=steps("As an employee, open Employee Center and search for the issue.","Confirm a relevant knowledge article appears; if it resolves the issue, stop (deflected).","If not resolved, create an incident from the portal and submit.","As an agent, confirm the incident appears with category, priority (Impact x Urgency), and routed to the correct group.","Confirm the SLA timer is visible for the priority.","Work and resolve the incident with resolution notes.","Confirm the requester receives notification and a satisfaction prompt."),
  exp="Incident is created via the portal, correctly categorized/prioritized/routed, SLA tracked, resolved, and the requester is notified and surveyed - end to end.",
  stories=["CONN-INC-001","CONN-INC-002","CONN-INC-003","CONN-INC-004","CONN-INC-005","CONN-INC-006","CONN-INC-008","CONN-INC-009","CONN-EC-001","CONN-EC-004","CONN-EC-005","CONN-KM-006","CONN-PA-001"]),
 dict(suite="1. Service Desk - Incident Lifecycle",sid="UAT-INC-02",name="Report an issue by phone (Vonage CTI)",
  who="Business Tester (Service Desk Agent)",pre="Agent logged into the workspace with the Vonage softphone; test number available",
  st=steps("Place an inbound test call to the service desk number.","Confirm the OpenFrame softphone rings in the agent workspace.","Confirm an Interaction opens with the caller matched (screen-pop).","Create an incident from the Interaction.","Confirm the Interaction links to the incident.","Resolve the incident and end the call."),
  exp="An inbound Vonage call opens a matched Interaction, the agent creates a linked incident, and resolves it - the phone channel works end to end.",
  stories=["CONN-CTI-001","CONN-CTI-002","CONN-CTI-003","CONN-CTI-004","CONN-CTI-005","CONN-INC-004","CONN-INC-006"]),
 dict(suite="1. Service Desk - Incident Lifecycle",sid="UAT-INC-03",name="Major incident declaration and coordination",
  who="Business Tester (Service Desk Manager + Coordinator)",pre="A P1 incident exists or can be created; MIM roles assigned",
  st=steps("Create or open a P1 incident that meets the major-incident criteria.","Promote/declare it as a Major Incident.","Confirm the MIM coordinator and roles engage and the MIM workspace opens.","Send a stakeholder communication from a template.","Drive the incident to resolution.","Confirm a Post-Incident Review (PIR) is generated."),
  exp="A qualifying P1 is declared major, coordination and comms engage via OOTB MIM, it is resolved, and a PIR is created.",
  stories=["CONN-INC-007","CONN-MIM-001","CONN-MIM-002","CONN-MIM-003","CONN-MIM-004","CONN-MIM-005"]),
 dict(suite="2. Request & Catalog Fulfillment",sid="UAT-REQ-01",name="Order a catalog item end to end",
  who="Business Tester (Requester + Approver + Fulfiller)",pre="Catalog items configured; approver and fulfillment group set",
  st=steps("As a requester, open Employee Center and order an in-scope catalog item.","Complete the variables and submit the request.","As the approver, receive the approval notification and approve (via email or portal).","Confirm the request moves to fulfillment and tasks generate to the right group.","As the fulfiller, complete the fulfillment tasks.","Confirm the request closes and the requester is notified; confirm the SLA tracked."),
  exp="A catalog request is ordered, approved, fulfilled, and closed with SLA tracking and notifications - the full request lifecycle.",
  stories=["CONN-SCR-001","CONN-SCR-002","CONN-SCR-003","CONN-SCR-004","CONN-SCR-005","CONN-SCR-006","CONN-EC-003","CONN-PF-006"]),
 dict(suite="2. Request & Catalog Fulfillment",sid="UAT-REQ-02",name="Order a service via Virtual Agent",
  who="Business Tester (Requester)",pre="Virtual Agent active with catalog topic; user authenticated",
  st=steps("Open the Virtual Agent from Employee Center.","Ask to order an in-scope item via conversation.","Provide the requested details in the conversation.","Submit and confirm a request is created.","Verify the request appears in 'my requests'."),
  exp="A user orders an in-scope item conversationally through the Virtual Agent and a correct request is created.",
  stories=["CONN-VA-001","CONN-VA-006","CONN-SCR-004"]),
 dict(suite="2. Request & Catalog Fulfillment",sid="UAT-REQ-03",name="Generic catch-all request routing",
  who="Business Tester (Requester)",pre="Generic catch-all catalog item configured",
  st=steps("Order a service that has no dedicated catalog item using the generic request.","Submit and confirm it routes to a defined service path/group.","Confirm no user community is left without a path."),
  exp="A service without a dedicated item is captured by the generic catch-all and routed correctly.",
  stories=["CONN-SCR-001","CONN-SCR-004","CONN-SCR-006"]),
 dict(suite="3. Change & Release",sid="UAT-CHG-01",name="Normal change with CI impact and CAB approval",
  who="Business Tester (Change Requester + CAB Member)",pre="A configuration item exists in CMDB; CAB configured",
  st=steps("Raise a Normal change and relate it to an in-scope CI.","Complete the risk assessment; confirm CI-driven risk scoring produces a score.","Confirm affected services/CIs display from the service map.","Confirm the change routes to CAB for review.","As CAB, approve the change.","Implement and close the change; confirm conflict/blackout checks ran."),
  exp="A normal change inherits CI/service impact, is risk-scored, reviewed by CAB, approved, implemented, and closed.",
  stories=["CONN-CHG-001","CONN-CHG-002","CONN-CHG-003","CONN-CHG-004","CONN-CHG-006","CONN-CHG-007","CONN-CHG-008","CONN-CMDB-008","CONN-CSDM-008"]),
 dict(suite="3. Change & Release",sid="UAT-CHG-02",name="Standard (pre-approved) change",
  who="Business Tester (Change Requester)",pre="Standard Change Catalog templates configured",
  st=steps("Raise a change from a Standard Change template.","Confirm it follows the pre-approved path (no full CAB).","Implement and close the change."),
  exp="A standard change is raised from a template, follows the pre-approved path, and closes.",
  stories=["CONN-CHG-005","CONN-CHG-001"]),
 dict(suite="3. Change & Release",sid="UAT-CHG-03",name="Emergency change expedited path",
  who="Business Tester (Change Manager)",pre="Emergency change type active",
  st=steps("Raise an Emergency change.","Confirm the expedited approval path engages.","Approve, implement, and close; confirm post-implementation review if required."),
  exp="An emergency change follows the expedited approval path and closes with the required review.",
  stories=["CONN-CHG-001","CONN-CHG-006"]),
 dict(suite="4. Problem to Permanent Fix",sid="UAT-PRB-01",name="Recurring incidents to problem to fix",
  who="Business Tester (Problem Manager)",pre="Several related incidents exist; PI similarity active",
  st=steps("Identify recurring/similar incidents (PI may surface a cluster).","Create a problem from the incidents.","Record a workaround in the KEDB and confirm it is discoverable from related incidents.","Raise/link a change to deliver the permanent fix.","Resolve the problem and confirm linked incidents reflect the resolution."),
  exp="Recurring incidents become a problem with a KEDB workaround, a change delivers the fix, and the problem resolves - closed-loop.",
  stories=["CONN-PRB-001","CONN-PRB-002","CONN-PRB-003","CONN-PRB-004","CONN-PRB-007","CONN-INC-009","CONN-PI-004","CONN-PI-005","CONN-CHG-001"]),
 dict(suite="5. Employee Self-Service & Deflection",sid="UAT-EX-01",name="Self-serve via search, knowledge, and Virtual Agent",
  who="Business Tester (Employee)",pre="Employee Center live with KB and Virtual Agent",
  st=steps("Open Employee Center and search for a common question.","Confirm AI Search returns a relevant knowledge article.","Open the Virtual Agent and attempt the same request conversationally.","Confirm the VA resolves it or offers to create a record.","Provide article feedback (useful/not)."),
  exp="An employee self-serves through search, knowledge, and the Virtual Agent - deflecting a ticket or creating one cleanly.",
  stories=["CONN-EC-001","CONN-EC-003","CONN-EC-004","CONN-EC-005","CONN-EC-007","CONN-KM-001","CONN-KM-006","CONN-KM-008","CONN-VA-001","CONN-VA-002","CONN-VA-005","CONN-INC-004"]),
 dict(suite="5. Employee Self-Service & Deflection",sid="UAT-EX-02",name="Check my ticket status via Virtual Agent",
  who="Business Tester (Employee)",pre="Authenticated VA session; user has an open record",
  st=steps("Open the Virtual Agent as an authenticated user.","Ask for the status of your open incident/request.","Confirm the VA returns your record status correctly."),
  exp="An authenticated user retrieves their ticket status conversationally.",
  stories=["CONN-VA-003","CONN-VA-005","CONN-EC-007"]),
 dict(suite="6. Identity & Access",sid="UAT-IAM-01",name="SSO login and role-based access",
  who="Business Tester (Standard User + Agent)",pre="SSO configured; AD groups mapped to roles",
  st=steps("Log in via SSO with a standard user.","Confirm successful authentication and appropriate menus/records.","Log in as a user in an ITIL group.","Confirm the agent sees agent capabilities per their role.","Confirm a user without a role cannot access restricted areas."),
  exp="SSO authenticates users and role/group membership from AD grants the correct access - no more, no less.",
  stories=["CONN-INT-001","CONN-INT-002","CONN-INT-005","CONN-PF-002","CONN-PF-004"]),
 dict(suite="7. CMDB / CSDM & Change Impact",sid="UAT-CMDB-01",name="CI data flows and drives change impact",
  who="Business Tester (CMDB/Change SME)",pre="SCCM/Intune connectors run; CSDM service maps built",
  st=steps("Confirm a device discovered by SCCM/Intune appears as a CI in the CMDB.","Open the related service map and confirm the CI is in the dependency chain.","Raise a change against that CI.","Confirm affected services and downstream CIs display for impact assessment.","Confirm risk reflects the impact."),
  exp="Connector-sourced CI data populates the CMDB, builds service maps, and drives change impact and risk - data to decision.",
  stories=["CONN-SGC-001","CONN-SGC-002","CONN-SGC-004","CONN-SGC-006","CONN-CMDB-001","CONN-CMDB-003","CONN-CMDB-008","CONN-CSDM-006","CONN-CSDM-008","CONN-DISC-003"]),
 dict(suite="7. CMDB / CSDM & Change Impact",sid="UAT-CMDB-02",name="CMDB health and source-of-record reconciliation",
  who="Business Tester (CMDB SME)",pre="Multiple CI sources active; IRE rules configured",
  st=steps("Open the CMDB Health dashboard and confirm completeness/compliance metrics.","Confirm a CI from overlapping sources resolves to a single record per the source-of-record rules.","Confirm a duplicate candidate is matched, not duplicated."),
  exp="CMDB health is measurable and IRE reconciliation keeps a single, authoritative CI across sources.",
  stories=["CONN-CMDB-004","CONN-CMDB-005","CONN-DISC-006","CONN-DISC-008","CONN-SGC-002","CONN-SGC-008"]),
 dict(suite="8. Asset Lifecycle (HAM)",sid="UAT-HAM-01",name="Asset from intake to retirement",
  who="Business Tester (Asset Manager)",pre="Stockrooms and asset classes configured",
  st=steps("Intake a new hardware asset into a stockroom.","Assign the asset to a user.","Confirm the asset links to its CMDB CI.","Move the asset through its lifecycle states.","Retire and dispose the asset; confirm the record reflects disposal."),
  exp="A hardware asset is tracked from intake through assignment, CI linkage, and retirement/disposal.",
  stories=["CONN-HAM-001","CONN-HAM-002","CONN-HAM-003","CONN-HAM-005","CONN-HAM-006","CONN-HAM-007","CONN-HAM-008","CONN-CMDB-008"]),
 dict(suite="9. Knowledge Management",sid="UAT-KM-01",name="Author, review, publish, and consume an article",
  who="Business Tester (Knowledge Author + Manager + Consumer)",pre="Knowledge bases and workflow configured",
  st=steps("As an author, create an article from a template in the correct KB.","Submit it through the review workflow.","As a manager, review and publish the article.","As a consumer, find the article via Employee Center search and the Virtual Agent.","Submit feedback on the article."),
  exp="An article moves through authoring, review, publish, consumption, and feedback - the knowledge lifecycle.",
  stories=["CONN-KM-001","CONN-KM-002","CONN-KM-003","CONN-KM-004","CONN-KM-005","CONN-KM-006","CONN-KM-008"]),
 dict(suite="10. Reporting & Dashboards",sid="UAT-PA-01",name="Operational activity reflects in dashboards",
  who="Business Tester (Manager / Analytics user)",pre="PA indicators activated; activities performed in prior scripts",
  st=steps("Perform or rely on prior incidents, requests, and changes.","Open the relevant Performance Analytics dashboard/scorecard.","Confirm indicators (MTTR, SLA attainment, request volume, change success) reflect the activity.","Confirm thresholds (green/yellow/red) display.","Confirm access is correct for the role."),
  exp="Operational activity surfaces correctly in PA dashboards and scorecards with thresholds and role-based access.",
  stories=["CONN-PA-001","CONN-PA-002","CONN-PA-003","CONN-PA-004","CONN-PA-006","CONN-INC-005","CONN-CHG-003"]),
]

# build coverage map: story -> [scripts]
cov={}
for sc in SCRIPTS:
    for st in sc["stories"]: cov.setdefault(st,[]).append(sc["sid"])

wb=Workbook(); wb.remove(wb.active)
def titlebar(ws,t,sub):
    ws["A1"]=t; ws["A1"].font=Font(name="Calibri",size=14,bold=True,color=NAVY)
    ws["A2"]=sub; ws["A2"].font=Font(name="Calibri",size=9,italic=True,color=SLATE)
def hrow(ws,cols,r=4):
    for c,h in enumerate(cols,1):
        cell=ws.cell(r,c,h); cell.font=Font(name="Calibri",size=11,bold=True,color=WHITE); cell.fill=PatternFill("solid",fgColor=NAVY)
        cell.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True); cell.border=box
    ws.row_dimensions[r].height=26

# Overview
ws=wb.create_sheet("Overview")
titlebar(ws,"CONNECTION - UAT END-TO-END TEST SCRIPTS","End-to-end business journeys grouped into suites. Each script traces directly to the user stories it validates (Stories Validated column + Story Coverage tab).")
r=4
def sec(t):
    global r; c=ws.cell(r,1,t); c.font=Font(name="Calibri",size=11,bold=True,color=NAVY); c.fill=PatternFill("solid",fgColor=CYAN); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2); r+=1
def ln(a,b):
    global r; ws.cell(r,1,a).font=bf(b=True); c=ws.cell(r,2,b); c.font=bf(); c.alignment=Alignment(wrap_text=True,vertical="top"); ws.row_dimensions[r].height=max(20,14+6*(len(b)//90)); r+=1
sec("What this is")
ln("Purpose","End-to-end UAT scripts for end users (business validation), grouped by functional suite to keep UAT contained. This is NOT story testing - story acceptance criteria are validated by the team during the sprint.")
ln("Traceability","Every script lists the exact story IDs it validates. The Story Coverage tab inverts this so each of the 141 stories shows the UAT script(s) that exercise it (or 'Sprint story-test only').")
ln("When","Stage 3 / Sprint 6-7 (UAT, Weeks 13-16), before Go-Live (Week 16).")
ln("Result codes","PASS = worked as described | FAIL = did not (log a defect) | BLOCKED = could not run | SKIP = not applicable. Do NOT FAIL for a feature that was never in scope.")
r+=1; sec("Suites")
for s in ["1. Service Desk - Incident Lifecycle","2. Request & Catalog Fulfillment","3. Change & Release","4. Problem to Permanent Fix","5. Employee Self-Service & Deflection","6. Identity & Access","7. CMDB / CSDM & Change Impact","8. Asset Lifecycle (HAM)","9. Knowledge Management","10. Reporting & Dashboards"]:
    ws.cell(r,1,s).font=bf(); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2); r+=1
ws.column_dimensions["A"].width=26; ws.column_dimensions["B"].width=104; ws.sheet_view.showGridLines=False

# Test Scripts
ws=wb.create_sheet("Test Scripts")
titlebar(ws,"UAT END-TO-END TEST SCRIPTS","Run each script as a journey. Record Result and (if FAIL) a Defect ID. Stories Validated ties each script to the backlog.")
cols=["Suite","Script ID","Scenario Name","Who Tests","Pre-Conditions","Test Steps","Expected End Result","Stories Validated","Result","Defect ID","Notes"]
hrow(ws,cols)
r=5
for i,sc in enumerate(SCRIPTS):
    vals=[sc["suite"],sc["sid"],sc["name"],sc["who"],sc["pre"],sc["st"],sc["exp"],", ".join(sc["stories"]),"","",""]
    for c,v in enumerate(vals,1):
        cell=ws.cell(r,c,v); cell.font=bf(); cell.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True); cell.border=box
        if i%2==1: cell.fill=PatternFill("solid",fgColor=ALT)
    ws.row_dimensions[r].height=max(90,14+7*(len(sc["st"])//40))
    r+=1
last=r-1
dv=DataValidation(type="list",formula1='"PASS,FAIL,BLOCKED,SKIP"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"I5:I{last}")
for v,col in [("PASS",GREEN),("FAIL",RED),("BLOCKED",AMBER),("SKIP",BLUE)]:
    from openpyxl.formatting.rule import CellIsRule
    ws.conditional_formatting.add(f"I5:I{last}",CellIsRule(operator="equal",formula=[f'"{v}"'],fill=PatternFill("solid",fgColor=col)))
for col,w in zip("ABCDEFGHIJK",[26,12,30,22,30,60,42,30,10,12,26]): ws.column_dimensions[col].width=w
ws.freeze_panes="A5"; ws.sheet_view.showGridLines=False

# Story Coverage
ws=wb.create_sheet("Story Coverage")
titlebar(ws,"STORY COVERAGE - direct correlation to UAT","Every story and where it is validated in UAT. 'Sprint story-test only' = validated against acceptance criteria during the sprint (not an end-user journey).")
hrow(ws,["Story ID","Module","Story","Validated by UAT script(s)","Validation type"])
r=5; e2e=0; sprintonly=0
for sid in ORDER:
    mod,short=TITLE[sid]
    scripts=cov.get(sid,[])
    if scripts: vtype="End-to-end UAT"; e2e+=1; covtxt=", ".join(scripts)
    else: vtype="Sprint story-test only"; sprintonly+=1; covtxt="Sprint story-test only (acceptance criteria)"
    vals=[sid,mod,short,covtxt,vtype]
    for c,v in enumerate(vals,1):
        cell=ws.cell(r,c,v); cell.font=bf(); cell.alignment=Alignment(wrap_text=True,vertical="top"); cell.border=box
        if r%2==0: cell.fill=PatternFill("solid",fgColor=ALT)
    ws.row_dimensions[r].height=22; r+=1
sr=r+1
ws.cell(sr,1,"SUMMARY").font=Font(name="Calibri",size=11,bold=True,color=NAVY)
ws.cell(sr+1,1,"Total stories").font=bf(c=SLATE); ws.cell(sr+1,2,len(ORDER)).font=bf(b=True,c=NAVY)
ws.cell(sr+2,1,"Covered by end-to-end UAT").font=bf(c=SLATE); ws.cell(sr+2,2,e2e).font=bf(b=True,c=NAVY)
ws.cell(sr+3,1,"Sprint story-test only").font=bf(c=SLATE); ws.cell(sr+3,2,sprintonly).font=bf(b=True,c=NAVY)
for col,w in zip("ABCDE",[15,18,46,40,22]): ws.column_dimensions[col].width=w
ws.freeze_panes="A5"; ws.sheet_view.showGridLines=False

# Defect Log
ws=wb.create_sheet("Defect Log")
titlebar(ws,"UAT DEFECT LOG","Log every FAIL here. ECS triages and resolves; testers re-test.")
hrow(ws,["Defect ID","Script ID","Story ID","Severity","Summary","Steps to Reproduce","Expected","Actual","Status","Owner","Resolution"])
for i in range(5,17):
    for c in range(1,12):
        cell=ws.cell(i,c,""); cell.border=box; cell.font=bf()
        if i%2==1: cell.fill=PatternFill("solid",fgColor=ALT)
dv2=DataValidation(type="list",formula1='"P1 - Critical,P2 - High,P3 - Medium,P4 - Low"',allow_blank=True); ws.add_data_validation(dv2); dv2.add("D5:D16")
dv3=DataValidation(type="list",formula1='"Open,In Triage,In Progress,Fixed - Retest,Closed,Deferred"',allow_blank=True); ws.add_data_validation(dv3); dv3.add("I5:I16")
for col,w in zip("ABCDEFGHIJK",[12,12,14,14,34,40,30,30,14,16,34]): ws.column_dimensions[col].width=w
ws.freeze_panes="A5"; ws.sheet_view.showGridLines=False

order=["Overview","Test Scripts","Story Coverage","Defect Log"]
wb._sheets.sort(key=lambda x: order.index(x.title) if x.title in order else 99)
wb.save(OUT)
print(f"Saved: {OUT} | scripts={len(SCRIPTS)} | stories covered E2E={e2e} | sprint-only={sprintonly}")
