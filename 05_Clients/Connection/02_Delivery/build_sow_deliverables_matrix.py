# -*- coding: utf-8 -*-
"""Connection - SOW Deliverables Matrix (EM day-1). Maps every SOW deliverable to its supporting baseline + status."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
OUT="/sessions/eager-epic-davinci/mnt/everforth-ecs-practice/05_Clients/Connection/02_Delivery/Connection_SOW_Deliverables_Matrix.xlsx"
NAVY="0B1F3A";WHITE="FFFFFF";ALT="F8FAFC";BORDER="E2E8F0";SLATE="475569"
GREEN="DCFCE7";BLUE="DBEAFE";AMBER="FEF3C7"
thin=Side(style="thin",color=BORDER);box=Border(left=thin,right=thin,top=thin,bottom=thin)
wb=Workbook();ws=wb.active;ws.title="SOW Deliverables Matrix"
ws["A1"]="CONNECTION - SOW DELIVERABLES MATRIX  (EM Day-1 Reference)"
ws["A1"].font=Font(name="Calibri",size=14,bold=True,color=NAVY)
ws["A2"]="Every deliverable committed in SOW v2.0 Sections 5, 10-11, mapped to its supporting baseline. Status: Ready (built for Connection) | Adapt from library (pull + tailor when the sprint arrives) | GAP - build."
ws["A2"].font=Font(name="Calibri",size=9,italic=True,color=SLATE)
H=["#","Deliverable","Period","Type","Audience","ECS Owner","Supporting Baseline (path / file)","Status","Notes"]
hr=4
for c,h in enumerate(H,1):
    cell=ws.cell(hr,c,h);cell.font=Font(name="Calibri",size=11,bold=True,color=WHITE)
    cell.fill=PatternFill("solid",fgColor=NAVY);cell.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True);cell.border=box
R=[
# Deliverable, Period, Type, Audience, Owner, Baseline, Status, Notes
("Governance Charter Package - Decision Rights + Customization Council Charter","Sprint 0","Doc","Client","EM/SA","01_Onboarding/Client_Facing/Customer_Governance_Charter.docx (+ lib INT-S0-04/05)","Ready","Charter built; lib decision-rights/council templates back it"),
("RACI Matrix (part of Governance Package)","Sprint 0","Doc","Joint","EM","01_Onboarding/Internal_Team/Connection_Role_and_Accountability_QuickRef.docx","Adapt from library","Quick-ref is RACI-style; produce a client-facing RACI cut"),
("Baseline 18-Week Project Plan","Sprint 0","Doc","Joint","EM","02_Delivery/Connection_18Week_Project_Plan.xlsx","Ready","Set start date (B3)"),
("Foundation Data Workbooks (users, locations, groups, SLAs, taxonomy)","Sprint 0","Doc","Client (fill)","PC","02_Delivery/Accelerator_Packs/Foundation_Accelerator_Pack/","Ready","Distribute Sprint 0"),
("Baseline Risk Register","Sprint 0","Doc","Joint","EM","02_Delivery/Connection_Governance_Triage_and_RAID.xlsx (RAID tab) / lib INT-S0-06","Ready","RAID covers it; or use lib risk register template"),
("Plug-ins installed + core configuration","Sprint 0","Platform Config","-","SA/TC","Workshops/Connection_Platform_Foundation_Workshop.pptx + Foundation pack","Ready","Collateral in place"),
("Platform Architecture & CSDM Alignment Document","Month 1","Doc","Client","SA","02_Delivery/Connection_Platform_Architecture_and_CSDM_Alignment.docx","Ready","Template built; SA completes from CSDM/CMDB workshop"),
("Sprint 1-2 Design Documentation (workbooks updated w/ decisions)","Month 1","Doc","Joint","PC/SA","03_Shared/04_Sprint_Workbooks (Platform, CSDM, CMDB, Discovery, Incident)","Adapt from library","Copy sprint workbooks into Connection; update with decisions"),
("Active Governance Triage Log","Month 1+","Doc","Joint","EM","02_Delivery/Connection_Governance_Triage_and_RAID.xlsx","Ready","Living through Go-Live"),
("ITSM Core v1 - Incident + Problem configured","Month 1","Platform Config","-","SA/TC","Workshops: Incident, Problem + ITSM pack","Ready","Collateral in place"),
("Sprint 3-4 Design Documentation","Month 2","Doc","Joint","PC/SA","03_Shared/04_Sprint_Workbooks (Change, Catalog)","Adapt from library","Copy + update with decisions"),
("ITSM Core v2 - Request, Change, CAB, SLA, assignment rules","Month 2","Platform Config","-","SA/TC","Workshops: Change + ITSM pack","Ready","Collateral in place"),
("Service Catalog Phase 1 - 5 items","Month 2","Platform Config","-","PC/TC","Workshops: Service Catalog + Service_Catalog pack","Ready","Collateral in place"),
("Integration Baseline - AD/SSO + SCCM + Intune validated","Month 2","Platform Config","-","SA/TC","Workshops: Integrations, SGC + Integration pack (AD/SSO/SCCM/Intune)","Ready","Vonage CTI pack also built"),
("Employee Experience - EC, VA, AI Search, KM","Month 3","Platform Config","-","PC/TC","Workshops: Employee Center, Virtual Agent, Knowledge + packs","Ready","Collateral in place"),
("HAM Foundations - stockrooms, asset classes, CSDM data model","Month 3","Platform Config","-","SA/TC","Workshops: HAM + ITAM_HAM_Foundations pack","Ready","Collateral in place"),
("Remaining Catalog Items (13-15 total)","Month 3","Platform Config","-","PC/TC","Workshops: Service Catalog + pack","Ready","Collateral in place"),
("Performance Analytics - baseline dashboard set","Month 3","Platform Config","-","SA","Workshops: Performance Analytics + PA pack","Ready","Collateral in place"),
("UAT Execution Report (results + P1/P2 defect log)","Go-Live","Doc","Joint","QA/PC","lib 02_Client/06_UAT_Execution (CLT-UAT-02/03) + 01_Internal/08_UAT_Test_Packs","Adapt from library","Strong library coverage; copy into Connection at UAT"),
("Go-Live Readiness Sign-Off","Go-Live","Doc","Client","EM","lib 02_Client/07_Closeout_and_Hypercare/CLT-CO-02_Go_Live_Checklist.docx","Adapt from library","Tailor checklist to Connection"),
("KT Package - Administrator Guide / Admin KT (4 sessions)","Go-Live","Doc","Client","SA","02_Delivery/Knowledge_Transfer/Connection_Administrator_Guide_and_KT.docx","Ready","Template built - 4-session Admin KT plan + admin reference"),
("KT Package - Train-the-Trainer materials (2 per process area)","Go-Live","Doc","Client","PC","02_Delivery/Knowledge_Transfer/Connection_Train_the_Trainer_Toolkit.docx","Ready","Template built - 2 sessions/area; demo-script refs"),
("KT Package - Sprint workbook set","Go-Live","Doc","Joint","PC","03_Shared/04_Sprint_Workbooks (Connection copies)","Adapt from library","Bundle the populated Connection workbooks"),
("Operational Handoff Pack (ownership matrix, support model, escalation)","Go-Live","Doc","Joint","EM","lib CLT-CO-03_Hypercare_Support_Model.docx + RACI","Adapt from library","Add an ownership matrix cut from RACI"),
("Lessons Learned & Project Closeout","Close","Doc","Joint","EM","lib CLT-CO-01_Engagement_Closeout_Summary.docx + 01_Internal/10_Lessons_Learned","Adapt from library","Copy + complete at close"),
("Hypercare (support model + exit report)","Close","Doc","Joint","EM","lib CLT-CO-03 + CLT-CO-06_Hypercare_Exit_Report.docx","Adapt from library","2-week Hypercare per SOW Sec 11"),
("Production environment live + verified","Go-Live","Platform Config","-","SA/EM","Cutover via project plan Stage 3","Ready","Governed cutover Wk 16"),
]
r=hr+1
for i,row in enumerate(R,1):
    vals=[i,row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7]]
    for c,v in enumerate(vals,1):
        cell=ws.cell(r,c,v);cell.font=Font(name="Calibri",size=10,color="1A1A1A")
        cell.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True);cell.border=box
        if i%2==0: cell.fill=PatternFill("solid",fgColor=ALT)
    r+=1
last=r-1
dv=DataValidation(type="list",formula1='"Ready,Adapt from library,GAP - build"',allow_blank=True);ws.add_data_validation(dv);dv.add(f"H{hr+1}:H{last}")
for val,col in [("Ready",GREEN),("Adapt from library",BLUE),("GAP - build",AMBER)]:
    ws.conditional_formatting.add(f"H{hr+1}:H{last}",CellIsRule(operator="equal",formula=[f'"{val}"'],fill=PatternFill("solid",fgColor=col)))
sr=last+2
ws.cell(sr,1,"SUMMARY").font=Font(name="Calibri",size=11,bold=True,color=NAVY)
for i,(lbl,f) in enumerate([("Total deliverables",f"=COUNTA(B{hr+1}:B{last})"),("Ready",f'=COUNTIF(H{hr+1}:H{last},"Ready")'),("Adapt from library",f'=COUNTIF(H{hr+1}:H{last},"Adapt from library")'),("GAP - build",f'=COUNTIF(H{hr+1}:H{last},"GAP - build")')]):
    ws.cell(sr+1+i,1,lbl).font=Font(name="Calibri",size=10,color=SLATE)
    ws.cell(sr+1+i,2,f).font=Font(name="Calibri",size=10,bold=True,color=NAVY)
for c,w in zip("ABCDEFGHI",[4,46,10,15,12,11,52,18,40]): ws.column_dimensions[c].width=w
ws.freeze_panes=f"A{hr+1}";ws.sheet_view.showGridLines=False
wb.save(OUT);print("Saved:",OUT,"rows:",len(R))
