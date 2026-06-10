# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
NAVY="0B1F3A";WHITE="FFFFFF";ALT="F8FAFC";BORDER="E2E8F0";SLATE="475569"
GREEN="DCFCE7";YELLOW="FEF9C3";RED="FEE2E2";BLUE="DBEAFE";AMBER="FEF3C7"
thin=Side(style="thin",color=BORDER);box=Border(left=thin,right=thin,top=thin,bottom=thin)
def bf(c="1A1A1A",b=False,sz=10): return Font(name="Calibri",size=sz,bold=b,color=c)
def header(ws,cols,hr=4):
    for c,h in enumerate(cols,1):
        cell=ws.cell(hr,c,h);cell.font=Font(name="Calibri",size=11,bold=True,color=WHITE);cell.fill=PatternFill("solid",fgColor=NAVY)
        cell.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True);cell.border=box
    ws.row_dimensions[hr].height=26

# ============ 1) CUSTOMER DEPENDENCY TRACKER ============
OUT1="/sessions/eager-epic-davinci/mnt/everforth-ecs-practice/05_Clients/Connection/02_Delivery/Connection_Customer_Dependency_Tracker.xlsx"
wb=Workbook();ws=wb.active;ws.title="Customer Dependencies"
ws["A1"]="CONNECTION - CUSTOMER RESPONSIBILITIES & DEPENDENCY TRACKER"
ws["A1"].font=Font(name="Calibri",size=14,bold=True,color=NAVY)
ws["A2"]="Operationalizes SOW v2.0 Section 6. These are structural dependencies that GATE ECS's ability to build on schedule. ECS flags each milestone in advance; at-risk items are jointly assessed and logged in the sprint workbook before they become a delay."
ws["A2"].font=Font(name="Calibri",size=9,italic=True,color=SLATE)
header(ws,["ID","Dependency","Type","Connection Owner","Due / Timing Rule","Target Sprint","Status","Impact if Late","Notes / Date"])
D=[
("Foundation Accelerator Pack workbooks completed (users, locations, departments, groups, assignment rules, schedules, SLAs)","Pre-Build","Connection PM + SMEs","Due 1 week before each in-scope execution sprint","Sprint 0-1","Not Started","Build cannot start; sprint slips","ECS provides templates"),
("SSO/AD configuration details + technical SME access for integration review","Pre-Build","Technical Lead","Sprint 0","Sprint 0","Not Started","SSO/AD config blocked; access issues at go-live",""),
("Licenses procured; dev/test environments provisioned with ECS admin access","Pre-Build","IT / Procurement","Sprint 0","Sprint 0","Not Started","No environment to build in","Structural gate"),
("MID Server(s) configured and validated","Pre-Build","Technical Lead","By end of Sprint 0","Sprint 0","Not Started","Discovery/integrations blocked; possible added cost","SOW: avoid delays/costs"),
("Complete IP Range Sets documented for Discovery schedules","Pre-Build","Technical Lead","Sprint 0-2","Sprint 2","Not Started","Discovery cannot run / incomplete CMDB",""),
("Named Product Owner with acceptance authority","Pre-Build","Executive Sponsor","Sprint 0","Sprint 0","Not Started","No one to accept stories; 3-day clock cannot run","Single empowered decision-maker"),
("Named SMEs per process area (ITSM, Catalog, CMDB/Discovery, Integrations, Employee Experience)","Pre-Build","Connection PM","Sprint 0","Sprint 0","Not Started","Workshops cannot be scheduled/decided",""),
("Executive Sponsor confirmed and available for governance cadence (bi-weekly, 45 min)","Pre-Build","Executive Sponsor","Sprint 0 onward","Sprint 0","Not Started","Governance and escalation stall",""),
("SME availability for workshops and validation sessions as scheduled","Ongoing","Connection PM","Per ECS/PM schedule","All sprints","Not Started","Cold/late workshops; rework",""),
("User story acceptance decisions within 3 business days of demo","Ongoing","Product Owner","3 business days post-demo","All sprints","Not Started","Stories pushed to backlog; may trigger a PCR","SOW-defined window"),
("Governance Triage Log review and disposition within 5 business days","Ongoing","Executive Sponsor","5 business days post-submission","All sprints","Not Started","Deviations unresolved; build blocked",""),
("Data quality ownership (accuracy/completeness of Accelerator Pack data)","Ongoing","Connection","Continuous","All sprints","Not Started","Bad data -> rework, CMDB health miss",""),
("CAB members available for Change Management configuration workshops","Ongoing","CAB Members","Sprint 4","Sprint 4","Not Started","Change/CAB config blocked",""),
("UAT testers assigned and available with sufficient platform knowledge","Ongoing","UAT Lead","UAT window (Sprint 6-7)","Sprint 6-7","Not Started","UAT cannot complete; go-live at risk",""),
("End-user communication, change management, and adoption (beyond train-the-trainer)","Ongoing","Connection","Sprint 7-8 onward","Sprint 7-8","Not Started","Low adoption; value not realized","Customer-owned per SOW"),
("Ongoing knowledge content creation and maintenance","Ongoing","Connection","Continuous","All / ongoing","Not Started","Thin KB; weak deflection","Customer-owned per SOW"),
("Vonage-side integration tasks completed","Ongoing","Connection / Vonage Admin","Sprint 4 (with CTI build)","Sprint 4","Not Started","CTI/phone channel not functional","Customer-owned per SOW"),
]
r=5
for i,row in enumerate(D,1):
    vals=[f"DEP-{i:02d}",row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7]]
    for c,v in enumerate(vals,1):
        cell=ws.cell(r,c,v);cell.font=bf();cell.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True);cell.border=box
        if i%2==0: cell.fill=PatternFill("solid",fgColor=ALT)
    ws.row_dimensions[r].height=max(34,14+6*(len(row[0])//40)); r+=1
last=r-1
dv=DataValidation(type="list",formula1='"Not Started,In Progress,Complete,At Risk,Blocked,Overdue"',allow_blank=True);ws.add_data_validation(dv);dv.add(f"G5:G{last}")
for v,col in [("Complete",GREEN),("In Progress",BLUE),("Not Started",YELLOW),("At Risk",AMBER),("Blocked",RED),("Overdue",RED)]:
    ws.conditional_formatting.add(f"G5:G{last}",CellIsRule(operator="equal",formula=[f'"{v}"'],fill=PatternFill("solid",fgColor=col)))
sr=last+2
ws.cell(sr,1,"SUMMARY").font=Font(name="Calibri",size=11,bold=True,color=NAVY)
for i,(lbl,f) in enumerate([("Total dependencies",f"=COUNTA(B5:B{last})"),("Complete",f'=COUNTIF(G5:G{last},"Complete")'),("In Progress",f'=COUNTIF(G5:G{last},"In Progress")'),("At Risk / Blocked / Overdue",f'=COUNTIF(G5:G{last},"At Risk")+COUNTIF(G5:G{last},"Blocked")+COUNTIF(G5:G{last},"Overdue")'),("Not Started",f'=COUNTIF(G5:G{last},"Not Started")')]):
    ws.cell(sr+1+i,1,lbl).font=bf(c=SLATE);ws.cell(sr+1+i,2,f).font=bf(b=True,c=NAVY)
for c,w in zip("ABCDEFGHI",[8,52,12,22,30,14,14,40,26]): ws.column_dimensions[c].width=w
ws.freeze_panes="A5";ws.sheet_view.showGridLines=False
wb.save(OUT1);print("Saved:",OUT1)

# ============ 2) GO-LIVE READINESS CHECKLIST ============
OUT2="/sessions/eager-epic-davinci/mnt/everforth-ecs-practice/05_Clients/Connection/02_Delivery/Connection_Go_Live_Readiness_Checklist.xlsx"
wb=Workbook();ws=wb.active;ws.title="Go-Live Readiness"
ws["A1"]="CONNECTION - GO-LIVE READINESS CHECKLIST (go/no-go gate, Week 16)"
ws["A1"].font=Font(name="Calibri",size=14,bold=True,color=NAVY)
ws["A2"]="Every Gate=Yes criterion must be Met (or have an agreed plan) before go-live. The EM compiles; the Product Owner and Executive Sponsor authorize."
ws["A2"].font=Font(name="Calibri",size=9,italic=True,color=SLATE)
header(ws,["#","Category","Readiness Criterion","Owner","Status","Gate (blocking?)","Evidence / Notes"])
G=[
("Configuration","All in-scope configuration stories Done and demo-accepted by the Product Owner","Solution Architect","Yes"),
("Configuration","No open P1/P2 configuration defects","Solution Architect","Yes"),
("Data","Foundation data loaded and validated; CMDB Health meets the agreed baseline","CMDB SME","Yes"),
("Data","Knowledge baseline content seeded and published","Knowledge Manager","No"),
("Integrations","AD/SSO, SCCM, Intune, and email validated in production-like environment","Integration Engineer","Yes"),
("Integrations","Vonage CTI / Interactions validated (inbound call -> Interaction -> Incident/Request)","CTI Engineer","Yes"),
("Testing","SIT complete with integration touchpoints validated","Technical Consultant","Yes"),
("Testing","UAT executed; pass rate meets the agreed threshold; P1/P2 defects resolved or have an agreed plan","UAT Lead","Yes"),
("Training & KT","Admin KT (4 sessions) delivered; Administrator Guide handed over","Solution Architect","Yes"),
("Training & KT","Train-the-Trainer sessions delivered per process area","Process Consultant","No"),
("Training & KT","Knowledge Transfer Package assembled and accepted","Solution Architect","No"),
("Cutover","Cutover Runbook reviewed and approved; rollback plan ready","Solution Architect","Yes"),
("Cutover","Cutover window scheduled and communicated; freeze in effect","Engagement Manager","Yes"),
("Support","Hypercare support model agreed; ECS L2+ coverage confirmed for the window","Engagement Manager","Yes"),
("Support","Connection Service Desk ready for L0-L1 intake; escalation contacts confirmed","Connection PM","Yes"),
("Governance","All open deviations in the Triage Log dispositioned; no blocking PCRs open","Engagement Manager","Yes"),
("Governance","Risk register has no open go-live blockers (or each has a mitigation)","Engagement Manager","Yes"),
("Sign-offs","Product Owner acceptance of the configured platform","Product Owner","Yes"),
("Sign-offs","Executive Sponsor go-live authorization (signed)","Executive Sponsor","Yes"),
]
r=5
for i,row in enumerate(G,1):
    vals=[i,row[0],row[1],row[2],"Not Started",row[3],""]
    for c,v in enumerate(vals,1):
        cell=ws.cell(r,c,v);cell.font=bf();cell.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True);cell.border=box
        if i%2==0: cell.fill=PatternFill("solid",fgColor=ALT)
    ws.row_dimensions[r].height=max(30,14+6*(len(row[1])//48)); r+=1
last=r-1
dv=DataValidation(type="list",formula1='"Not Started,In Progress,Met,At Risk,Not Met"',allow_blank=True);ws.add_data_validation(dv);dv.add(f"E5:E{last}")
for v,col in [("Met",GREEN),("In Progress",BLUE),("Not Started",YELLOW),("At Risk",AMBER),("Not Met",RED)]:
    ws.conditional_formatting.add(f"E5:E{last}",CellIsRule(operator="equal",formula=[f'"{v}"'],fill=PatternFill("solid",fgColor=col)))
sr=last+2
ws.cell(sr,1,"SUMMARY").font=Font(name="Calibri",size=11,bold=True,color=NAVY)
for i,(lbl,f) in enumerate([("Total criteria",f"=COUNTA(C5:C{last})"),("Met",f'=COUNTIF(E5:E{last},"Met")'),("Blocking gates",f'=COUNTIF(F5:F{last},"Yes")'),("Blocking gates Met",f'=COUNTIFS(F5:F{last},"Yes",E5:E{last},"Met")'),("Blocking gates NOT met",f'=COUNTIFS(F5:F{last},"Yes")-COUNTIFS(F5:F{last},"Yes",E5:E{last},"Met")')]):
    ws.cell(sr+1+i,1,lbl).font=bf(c=SLATE);ws.cell(sr+1+i,2,f).font=bf(b=True,c=NAVY)
ws.cell(sr+6,1,"GO / NO-GO: GO only when all blocking gates are Met.").font=bf(b=True,c=NAVY)
for c,w in zip("ABCDEFG",[4,16,56,22,14,16,34]): ws.column_dimensions[c].width=w
ws.freeze_panes="A5";ws.sheet_view.showGridLines=False
wb.save(OUT2);print("Saved:",OUT2)
