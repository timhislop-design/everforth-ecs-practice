# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
NAVY="0B1F3A";WHITE="FFFFFF";ALT="F8FAFC";BORDER="E2E8F0";SLATE="475569";CYAN="ECFEFF"
GREEN="DCFCE7";YELLOW="FEF9C3";RED="FEE2E2";AMBER="FEF3C7";BLUE="DBEAFE"
thin=Side(style="thin",color=BORDER);box=Border(left=thin,right=thin,top=thin,bottom=thin)
def bf(c="1A1A1A",b=False,sz=10): return Font(name="Calibri",size=sz,bold=b,color=c)
def titlebar(ws,t,sub):
    ws["A1"]=t;ws["A1"].font=Font(name="Calibri",size=14,bold=True,color=NAVY)
    ws["A2"]=sub;ws["A2"].font=Font(name="Calibri",size=9,italic=True,color=SLATE)
def header(ws,cols,hr=4):
    for c,h in enumerate(cols,1):
        cell=ws.cell(hr,c,h);cell.font=Font(name="Calibri",size=11,bold=True,color=WHITE);cell.fill=PatternFill("solid",fgColor=NAVY)
        cell.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True);cell.border=box
    ws.row_dimensions[hr].height=26
def blankrows(ws,hr,n,ncol):
    for i in range(hr+1,hr+1+n):
        for c in range(1,ncol+1):
            cell=ws.cell(i,c,"");cell.border=box;cell.font=bf()
            if i%2==1: cell.fill=PatternFill("solid",fgColor=ALT)
def dv(ws,col,opts,r1,r2):
    d=DataValidation(type="list",formula1='"'+",".join(opts)+'"',allow_blank=True);ws.add_data_validation(d);d.add(f"{col}{r1}:{col}{r2}")

# ===== PROJECT CONTROLS WORKBOOK =====
OUT="/sessions/eager-epic-davinci/mnt/everforth-ecs-practice/05_Clients/Connection/02_Delivery/Connection_Project_Controls.xlsx"
wb=Workbook();wb.remove(wb.active)
# PCR Log
ws=wb.create_sheet("PCR Log")
titlebar(ws,"CONNECTION - PROJECT CHANGE REQUEST (PCR) LOG","Per SOW Sec 9. Any scope/schedule change, or the 6th approved customization, or chronic acceptance/dependency delay, is logged and dispositioned here.")
header(ws,["PCR ID","Date Raised","Trigger","Description","Impact (scope/budget/schedule)","Disposition","Two-Key (Sponsor / ECS Practice)","Status","Notes"])
blankrows(ws,4,12,9)
dv(ws,"C",["Customer-initiated scope change","Schedule/dependency impact","Customization cap (#6)","Acceptance delay (>3 days)"],5,16)
dv(ws,"F",["Approved","Rejected","Deferred"],5,16); dv(ws,"H",["Open","In Review","Decided","Closed"],5,16)
for c,w in zip("ABCDEFGHI",[10,12,24,40,32,14,24,12,30]): ws.column_dimensions[c].width=w
ws.freeze_panes="A5";ws.sheet_view.showGridLines=False
# PCR Form template
ws=wb.create_sheet("PCR Request Form")
titlebar(ws,"PCR REQUEST FORM (template - copy per request)","Complete one form per PCR. The Solution Architect drafts the impact assessment; the two-key decision approves/rejects.")
r=4
fields=[("PCR ID",""),("Date raised",""),("Raised by",""),("Trigger",""),("Description of change",""),
 ("Business need / justification",""),("OOTB alternative considered (SA)",""),("Scope impact",""),("Budget impact",""),
 ("Schedule impact",""),("Upgrade/technical impact (SA)",""),("Sponsor decision (business need)",""),
 ("ECS Practice decision (technical path)",""),("Final disposition",""),("Decision date",""),("Logged in PCR Log? (Y/N)","")]
for lbl,val in fields:
    ws.cell(r,1,lbl).font=bf(b=True,c=NAVY);ws.cell(r,1).border=box
    c=ws.cell(r,2,val);c.font=bf();c.fill=PatternFill("solid",fgColor=AMBER if val=="" else WHITE);c.border=box;c.alignment=Alignment(wrap_text=True,vertical="top")
    ws.row_dimensions[r].height=28;r+=1
ws.column_dimensions["A"].width=34;ws.column_dimensions["B"].width=80;ws.sheet_view.showGridLines=False
# Acceptance & Sign-off Log
ws=wb.create_sheet("Acceptance & Sign-off Log")
titlebar(ws,"ACCEPTANCE & SIGN-OFF LOG","Operationalizes the SOW 3-business-day acceptance window. Late acceptance pushes stories to backlog and may trigger a PCR.")
header(ws,["Item ID","Type","Description","Demo / Delivery Date","Acceptance Due (3 biz days)","Accepted By","Date Accepted","Status","Notes"])
blankrows(ws,4,14,9)
dv(ws,"B",["Story","Sprint","Deliverable","Workshop","Go-Live"],5,18); dv(ws,"H",["Pending","Accepted","Rejected","Overdue"],5,18)
for v,col in [("Accepted",GREEN),("Pending",YELLOW),("Rejected",RED),("Overdue",RED)]:
    ws.conditional_formatting.add("H5:H18",CellIsRule(operator="equal",formula=[f'"{v}"'],fill=PatternFill("solid",fgColor=col)))
for c,w in zip("ABCDEFGHI",[12,12,40,16,18,18,14,12,30]): ws.column_dimensions[c].width=w
ws.freeze_panes="A5";ws.sheet_view.showGridLines=False
# Decision Register
ws=wb.create_sheet("Decision Register")
titlebar(ws,"DECISION REGISTER","Running log of engagement decisions (beyond deviations, which live in the Triage Log).")
header(ws,["Decision ID","Date","Topic / Area","Decision","Made By","Source (workshop / meeting)","Status","Notes"])
blankrows(ws,4,14,8)
dv(ws,"G",["Open","Decided","Revisited","Superseded"],5,18)
for c,w in zip("ABCDEFGH",[12,12,22,44,18,24,12,28]): ws.column_dimensions[c].width=w
ws.freeze_panes="A5";ws.sheet_view.showGridLines=False
# Assumptions & Out-of-Scope
ws=wb.create_sheet("Assumptions & Out-of-Scope")
titlebar(ws,"ASSUMPTIONS & OUT-OF-SCOPE REGISTER","Per SOW Sec 7-8. Track assumptions (and whether they hold) and out-of-scope items (so scope creep is visible).")
header(ws,["ID","Type","Statement","SOW Ref","Owner","Status","Impact if Wrong / Requested","Notes"])
seed=[
 ["A-01","Assumption","Connection provides timely, accurate data via Accelerator Packs","Sec 6/8","Connection PM","Holds","Rework, CMDB health miss",""],
 ["A-02","Assumption","Existing Discovery/Catalog configs are leveraged only where free of technical debt","Sec 2","Solution Architect","Holds","Added scope if debt found",""],
 ["A-03","Assumption","Connection completes Vonage-side integration tasks","Sec 6","Connection","Holds","CTI not functional",""],
 ["O-01","Out-of-Scope","Event Management (EM) - later phase","Sec 7","-","Confirmed","PCR if requested in Phase 1",""],
 ["O-02","Out-of-Scope","Now Assist / GenAI - later phase","Sec 7","-","Confirmed","PCR if requested",""],
 ["O-03","Out-of-Scope","SAM (Software Asset Mgmt) beyond foundations","Sec 7","-","Confirmed","PCR if requested",""],
 ["O-04","Out-of-Scope","End-user training delivery & broader OCM (customer-owned)","Sec 11","Connection","Confirmed","Adoption risk if not done",""],
]
r=5
for i,row in enumerate(seed):
    for c,v in enumerate(row,1):
        cell=ws.cell(r,c,v);cell.font=bf();cell.alignment=Alignment(wrap_text=True,vertical="top");cell.border=box
        if i%2==1: cell.fill=PatternFill("solid",fgColor=ALT)
    r+=1
blankrows(ws,r-1,6,8)
dv(ws,"B",["Assumption","Out-of-Scope"],5,r+5); dv(ws,"F",["Holds","Invalidated","Confirmed","Requested"],5,r+5)
for c,w in zip("ABCDEFGH",[8,14,46,10,18,12,34,24]): ws.column_dimensions[c].width=w
ws.freeze_panes="A5";ws.sheet_view.showGridLines=False
wb._sheets.sort(key=lambda x:["PCR Log","PCR Request Form","Acceptance & Sign-off Log","Decision Register","Assumptions & Out-of-Scope"].index(x.title))
wb.save(OUT);print("Saved:",OUT)

# ===== SIT TEST SCRIPTS & TEST DATA PLAN =====
OUT2="/sessions/eager-epic-davinci/mnt/everforth-ecs-practice/05_Clients/Connection/02_Delivery/Connection_SIT_Test_Scripts.xlsx"
wb=Workbook();wb.remove(wb.active)
ws=wb.create_sheet("SIT Scripts")
titlebar(ws,"CONNECTION - SIT (SYSTEM INTEGRATION TEST) SCRIPTS","ECS-internal integration testing before UAT. Validates each integration touchpoint end-to-end in sub-production.")
header(ws,["SIT ID","Integration","Scenario","Pre-Conditions","Test Steps","Expected Result","Result","Defect ID"])
def steps(*a): return "\n".join(f"{i}. {x}" for i,x in enumerate(a,1))
SIT=[
 ["SIT-01","AD/SSO","SSO login + attribute mapping","IdP configured; test users in AD","1. Initiate SSO login as a test user\n2. Confirm SAML assertion accepted\n3. Verify user record + attributes","User authenticates; profile attributes populate from AD"],
 ["SIT-02","AD groups","Group-to-role sync","AD group mapping configured","1. Add a test user to a mapped AD group\n2. Run/await the group import\n3. Check ServiceNow group/role membership","User receives the mapped group/role in ServiceNow"],
 ["SIT-03","SCCM SGC","CI import + class mapping","SCCM connector configured; MID up","1. Run the SCCM scheduled import\n2. Inspect imported CIs and classes\n3. Validate sample against SCCM","In-scope CIs populate with correct class/attributes; no load errors"],
 ["SIT-04","Intune","Endpoint CI import","Intune connector configured","1. Run the Intune import\n2. Inspect mobile/endpoint CIs","Intune-managed devices appear as CIs"],
 ["SIT-05","IRE","Reconciliation / dedup across sources","SCCM + Intune + Discovery active","1. Ingest a CI present in two sources\n2. Confirm IRE matches to one record","Single authoritative CI per source-of-record rules; no duplicates"],
 ["SIT-06","Email","Outbound + inbound","SMTP relay + inbound mailbox set","1. Trigger a notification\n2. Send an inbound test email","Notification delivered; inbound action creates/updates the correct record"],
 ["SIT-07","Vonage CTI","Inbound call -> Interaction -> Incident","OpenFrame + Vonage adapter configured","1. Place an inbound test call\n2. Confirm screen-pop Interaction with caller match\n3. Create an incident from the Interaction","Call opens a matched Interaction; linked incident is created"],
 ["SIT-08","CMDB->Change","CI impact drives change risk","CMDB populated; CSDM maps built","1. Raise a change against a CI\n2. Confirm affected services/CIs + risk score","Change shows CI/service impact and a CI-driven risk score"],
]
r=5
for i,row in enumerate(SIT):
    vals=[row[0],row[1],row[2],row[3],row[4],row[5],"",""]
    for c,v in enumerate(vals,1):
        cell=ws.cell(r,c,v);cell.font=bf();cell.alignment=Alignment(wrap_text=True,vertical="top");cell.border=box
        if i%2==1: cell.fill=PatternFill("solid",fgColor=ALT)
    ws.row_dimensions[r].height=max(58,14+7*(len(row[4])//30));r+=1
last=r-1
dv(ws,"G",["PASS","FAIL","BLOCKED"],5,last)
for v,col in [("PASS",GREEN),("FAIL",RED),("BLOCKED",AMBER)]:
    ws.conditional_formatting.add(f"G5:G{last}",CellIsRule(operator="equal",formula=[f'"{v}"'],fill=PatternFill("solid",fgColor=col)))
for c,w in zip("ABCDEFGH",[10,14,30,28,52,40,10,12]): ws.column_dimensions[c].width=w
ws.freeze_panes="A5";ws.sheet_view.showGridLines=False
# Test Data Plan
ws=wb.create_sheet("Test Data Plan")
titlebar(ws,"TEST DATA PLAN (SIT + UAT)","What test data is needed, its source, and how it is refreshed/anonymized for testing.")
header(ws,["#","Data set","Purpose","Source","Volume","Refresh / Anonymization","Owner","Status"])
TD=[
 ["Users & groups","SSO/role testing","AD import (subset)","~50 test users","Refresh from AD; no anonymization needed (internal)","Tech Lead"],
 ["CIs (servers/endpoints)","CMDB/Change/Discovery","SCCM/Intune (subset)","~100 CIs","Discovery/connector run in sub-prod","Integration Engineer"],
 ["Catalog items + requests","Request/fulfillment","Configured items","Top 10-15 items","Created during build","Process Consultant"],
 ["Knowledge articles","Search/VA/deflection","Ported baseline","~20 articles","Ported from legacy; reviewed","Knowledge Manager"],
 ["Incidents/Problems (seed)","Incident/Problem/PI","Generated","~30 records","Created for PI similarity + reporting","Process Consultant"],
 ["Phone test number + agent","Vonage CTI","Vonage tenant","1 number, 1 agent","Provided by Connection","Connection / CTI"],
]
r=5
for i,row in enumerate(TD,1):
    vals=[i]+row+["Not Started"]
    for c,v in enumerate(vals,1):
        cell=ws.cell(r,c,v);cell.font=bf();cell.alignment=Alignment(wrap_text=True,vertical="top");cell.border=box
        if i%2==0: cell.fill=PatternFill("solid",fgColor=ALT)
    r+=1
dv(ws,"H",["Not Started","In Progress","Ready"],5,r-1)
for c,w in zip("ABCDEFGH",[4,22,24,22,14,34,18,12]): ws.column_dimensions[c].width=w
ws.freeze_panes="A5";ws.sheet_view.showGridLines=False
wb._sheets.sort(key=lambda x:["SIT Scripts","Test Data Plan"].index(x.title))
wb.save(OUT2);print("Saved:",OUT2)
