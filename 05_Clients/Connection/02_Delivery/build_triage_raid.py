# -*- coding: utf-8 -*-
"""Build: Connection - Governance Triage Log & RAID (xlsx). Shared (both teams).
Consolidates INT-TBV-03 Variance Tracker into a lean Connection log + a RAID tab."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

OUT = "/sessions/eager-epic-davinci/mnt/everforth-ecs-practice/05_Clients/Connection/02_Delivery/Connection_Governance_Triage_and_RAID.xlsx"
NAVY="0B1F3A"; WHITE="FFFFFF"; ALT="F8FAFC"; BORDER="E2E8F0"; SLATE="475569"
GREEN="DCFCE7"; YELLOW="FEF9C3"; RED="FEE2E2"; BLUE="DBEAFE"
thin=Side(style="thin",color=BORDER); box=Border(left=thin,right=thin,top=thin,bottom=thin)

def header_row(ws, headers, r):
    for c,h in enumerate(headers,1):
        cell=ws.cell(r,c,h)
        cell.font=Font(name="Calibri",size=11,bold=True,color=WHITE)
        cell.fill=PatternFill("solid",fgColor=NAVY)
        cell.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
        cell.border=box

def body(ws, rows, r0, ncols):
    for i,row in enumerate(rows):
        for c in range(1,ncols+1):
            cell=ws.cell(r0+i,c,row[c-1] if c-1<len(row) else "")
            cell.font=Font(name="Calibri",size=10,color="1A1A1A")
            cell.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
            cell.border=box
            if i%2==0: cell.fill=PatternFill("solid",fgColor=ALT)

wb=Workbook()

# ---- Tab 1: Triage Log ----
ws=wb.active; ws.title="Triage Log"
ws["A1"]="CONNECTION - GOVERNANCE TRIAGE LOG"; ws["A1"].font=Font(name="Calibri",size=14,bold=True,color=NAVY)
ws["A2"]="System of record for every deviation request - visible to both teams. Council-approved + built customizations add to variance (cap 5). Rejected do not. Log within 24 hrs of decision."
ws["A2"].font=Font(name="Calibri",size=9,italic=True,color=SLATE)
h=["Req ID","Sprint Raised","Requester","Description of Request","OOTB Alternative (SA Analysis)","Classification","Rule of 3 Result","Council Decision","Decision Date","Two-Key (Sponsor / Practice)","PCR?","Effort (hrs)","Notes"]
hr=4; header_row(ws,h,hr)
rows=[[f"CVT-{i:03d}","","","","","","","","","","",""] for i in range(1,13)]
body(ws,rows,hr+1,len(h)); last=hr+len(rows)
# dropdowns
def dv(ws,col,opts,r1,r2):
    d=DataValidation(type="list",formula1='"'+",".join(opts)+'"',allow_blank=True); ws.add_data_validation(d); d.add(f"{col}{r1}:{col}{r2}")
dv(ws,"F",["Config (within OOTB)","Customization","PCR"],hr+1,last)
dv(ws,"G",["Pass - OOTB","Fail - Customization"],hr+1,last)
dv(ws,"H",["Approved","Rejected","Deferred","In Triage"],hr+1,last)
dv(ws,"K",["Yes","No"],hr+1,last)
for val,color in [("Approved",GREEN),("Rejected",RED),("Deferred",YELLOW),("In Triage",BLUE)]:
    ws.conditional_formatting.add(f"H{hr+1}:H{last}",CellIsRule(operator="equal",formula=[f'"{val}"'],fill=PatternFill("solid",fgColor=color)))
# summary
sr=last+2
ws.cell(sr,1,"SUMMARY").font=Font(name="Calibri",size=11,bold=True,color=NAVY)
summ=[("Approved customizations (cap 5)",f'=COUNTIFS(F{hr+1}:F{last},"Customization",H{hr+1}:H{last},"Approved")'),
      ("Remaining before PCR conversation",f'=5-COUNTIFS(F{hr+1}:F{last},"Customization",H{hr+1}:H{last},"Approved")'),
      ("Open (In Triage)",f'=COUNTIF(H{hr+1}:H{last},"In Triage")'),
      ("Rejected",f'=COUNTIF(H{hr+1}:H{last},"Rejected")'),
      ("PCRs triggered",f'=COUNTIF(K{hr+1}:K{last},"Yes")')]
for i,(lbl,f) in enumerate(summ):
    ws.cell(sr+1+i,1,lbl).font=Font(name="Calibri",size=10,color=SLATE)
    ws.cell(sr+1+i,2,f).font=Font(name="Calibri",size=10,bold=True,color=NAVY)
for c,w in zip("ABCDEFGHIJKLM",[9,11,14,30,30,18,16,14,12,20,7,10,24]): ws.column_dimensions[c].width=w
ws.freeze_panes=f"A{hr+1}"; ws.sheet_view.showGridLines=False

# ---- Tab 2: RAID ----
r=wb.create_sheet("RAID")
r["A1"]="CONNECTION - RAID LOG"; r["A1"].font=Font(name="Calibri",size=14,bold=True,color=NAVY)
r["A2"]="Risks, Assumptions, Issues, Dependencies. Reviewed weekly; material items surface in the Weekly Status Report and Executive Health Dashboard."
r["A2"].font=Font(name="Calibri",size=9,italic=True,color=SLATE)
h2=["Type","ID","Description","Impact","Likelihood","Owner","Status","Due","Mitigation / Action"]
hr2=4; header_row(r,h2,hr2)
rr=[[t,f"{p}-{i:02d}","","","","","","",""] for t,p in [("Risk","R"),("Assumption","A"),("Issue","I"),("Dependency","D")] for i in range(1,4)]
body(r,rr,hr2+1,len(h2)); last2=hr2+len(rr)
dv(r,"A",["Risk","Assumption","Issue","Dependency"],hr2+1,last2)
dv(r,"D",["High","Medium","Low"],hr2+1,last2)
dv(r,"E",["High","Medium","Low"],hr2+1,last2)
dv(r,"G",["Open","In Progress","Mitigated","Closed","Blocked"],hr2+1,last2)
for val,color in [("Closed",GREEN),("Mitigated",GREEN),("In Progress",BLUE),("Blocked",RED),("Open",YELLOW)]:
    r.conditional_formatting.add(f"G{hr2+1}:G{last2}",CellIsRule(operator="equal",formula=[f'"{val}"'],fill=PatternFill("solid",fgColor=color)))
for c,w in zip("ABCDEFGHI",[13,8,40,11,12,16,14,12,34]): r.column_dimensions[c].width=w
r.freeze_panes=f"A{hr2+1}"; r.sheet_view.showGridLines=False

wb.save(OUT); print("Saved:",OUT)
