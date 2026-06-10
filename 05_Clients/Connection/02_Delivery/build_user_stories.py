# -*- coding: utf-8 -*-
import sys; sys.path.insert(0,"/tmp")
import sd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
OUT="/sessions/eager-epic-davinci/mnt/everforth-ecs-practice/05_Clients/Connection/02_Delivery/Connection_User_Stories_SN_Agile.xlsx"
NAVY="0B1F3A";WHITE="FFFFFF";ALT="F8FAFC";BORDER="E2E8F0";SLATE="475569";CYAN="ECFEFF"
thin=Side(style="thin",color=BORDER);box=Border(left=thin,right=thin,top=thin,bottom=thin)
def hfont(b=True,c=WHITE,sz=11): return Font(name="Calibri",size=sz,bold=b,color=c)
def bfont(c="1A1A1A",b=False,sz=10): return Font(name="Calibri",size=sz,bold=b,color=c)
wb=Workbook(); wb.remove(wb.active)

# assign IDs per module
seq={}; 
for s in sd.STORIES:
    ab=sd.MODMETA[s["module"]]["abbr"]; seq[ab]=seq.get(ab,0)+1
    s["number"]=f"CONN-{ab}-{seq[ab]:03d}"
    s["description"]=f"As a {s['role']}, I want {s['want']}, so that {s['benefit']}."
    s["acceptance_criteria"]="\n".join(f"AC{i}: {a}" for i,a in enumerate(s["ac"],1))
    s["epic"]=f"{s['module']} - Phase 1 Configuration"

def title(ws,t,sub,span):
    ws["A1"]=t; ws["A1"].font=Font(name="Calibri",size=14,bold=True,color=NAVY)
    ws["A2"]=sub; ws["A2"].font=Font(name="Calibri",size=9,italic=True,color=SLATE)
def header(ws,cols,r=4):
    for c,h in enumerate(cols,1):
        cell=ws.cell(r,c,h); cell.font=hfont(); cell.fill=PatternFill("solid",fgColor=NAVY)
        cell.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True); cell.border=box
    ws.row_dimensions[r].height=26

STORY_COLS=["number","epic","short_description","description","acceptance_criteria","definition_of_done","story_points","priority","sprint","assignment (role)","state","notes / OOTB reference"]
def story_sheet(name, sheetkind):
    ws=wb.create_sheet(name)
    n="Application" if sheetkind=="app" else "Integration"
    title(ws,f"CONNECTION - {n.upper()} USER STORIES (SN Agile / rm_story import)",
          "Default Phase 1 configuration stories for the contract-configurable decisions. Given/When/Then acceptance criteria; shared Definition of Done on its own tab. See README for import steps.","")
    header(ws,STORY_COLS)
    r=5
    rows=[s for s in sd.STORIES if sd.MODMETA[s["module"]]["sheet"]==sheetkind]
    for i,s in enumerate(rows):
        meta=sd.MODMETA[s["module"]]
        vals=[s["number"],s["epic"],s["short"],s["description"],s["acceptance_criteria"],
              "See 'Definition of Done' tab",s["points"],s["priority"],meta["sprint"],meta["role"],"Draft",s["note"]]
        for c,v in enumerate(vals,1):
            cell=ws.cell(r,c,v); cell.font=bfont(); cell.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True); cell.border=box
            if i%2==1: cell.fill=PatternFill("solid",fgColor=ALT)
        ws.row_dimensions[r].height=max(54,14+11*(len(s["acceptance_criteria"])//70))
        r+=1
    last=r-1
    dv=DataValidation(type="list",formula1='"Draft,Ready,Work in progress,Complete,Cancelled"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"K5:K{last}")
    dvp=DataValidation(type="list",formula1='"1 - Critical,2 - High,3 - Moderate,4 - Low"',allow_blank=True); ws.add_data_validation(dvp); dvp.add(f"H5:H{last}")
    for col,w in zip("ABCDEFGHIJKL",[15,26,40,52,72,22,8,14,10,26,11,44]): ws.column_dimensions[col].width=w
    ws.freeze_panes="A5"; ws.sheet_view.showGridLines=False
    return len(rows)

# README
ws=wb.create_sheet("README - Import Guide")
title(ws,"CONNECTION - USER STORY BACKLOG (ServiceNow Agile import)",
 "Default Phase 1 user stories for the configurable decisions defined by the SOW. NOT exhaustive per process - one story per contract-configurable decision.","")
r=4
def sec(t):
    global r; c=ws.cell(r,1,t); c.font=Font(name="Calibri",size=11,bold=True,color=NAVY); c.fill=PatternFill("solid",fgColor=CYAN)
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2); r+=1
def line(a,b=""):
    global r; ws.cell(r,1,a).font=bfont(b=True); c=ws.cell(r,2,b); c.font=bfont(); c.alignment=Alignment(wrap_text=True,vertical="top"); ws.row_dimensions[r].height=max(20,14+6*(len(b)//80)); r+=1
sec("What this is")
line("Purpose","Seed backlog of default configuration user stories for Connection Phase 1. Each story maps to a contract-configurable decision from the workshop decks / SOW. Teams refine estimates and add detail per engagement.")
line("Tabs","Epics (import first) | Application User Stories | Integration User Stories | Definition of Done | this README.")
line("Acceptance criteria","Given/When/Then per story. The shared Definition of Done lives on its own tab and applies to every story.")
r+=1; sec("How to import to ServiceNow Agile (rm_story / rm_epic)")
for stp in [
 "1. Import EPICS first: Agile Development > Epics, or System Import Sets > Load Data from the 'Epics' tab into target table rm_epic. Coalesce on Number (short_description).",
 "2. Import STORIES: Load the 'Application User Stories' and 'Integration User Stories' tabs into target table rm_story via Import Set + Transform Map.",
 "3. Map columns to fields (see mapping below). Set Epic by display name/number; Sprint is optional until sprints exist.",
 "4. Set State = Draft on import; refine Story Points and Assignment Group in ServiceNow.",
 "5. Validate a sample, then commit. Re-run is safe if you coalesce on Number."]:
    ws.cell(r,1,stp).font=bfont(); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2); ws.cell(r,1).alignment=Alignment(wrap_text=True,vertical="top"); ws.row_dimensions[r].height=max(20,14+6*(len(stp)//90)); r+=1
r+=1; sec("Field mapping (workbook column -> rm_story field)")
for a,b in [("number","Number (or u_external_id)"),("epic","Epic [reference rm_epic]"),("short_description","Short description"),("description","Description"),("acceptance_criteria","Acceptance criteria"),("story_points","Story points"),("priority","Priority"),("sprint","Sprint [reference rm_sprint]"),("assignment (role)","Assignment group / Assigned to"),("state","State"),("notes / OOTB reference","Work notes / comments")]:
    line(a,b)
r+=1; sec("Conventions")
line("Story points","Fibonacci (1,2,3,5,8). Estimates are starting points - refine in sprint planning.")
line("Priority","1 - Critical / 2 - High / 3 - Moderate / 4 - Low.")
line("State","Draft on import; progress in ServiceNow (Ready > Work in progress > Complete).")
line("Numbering","CONN-<MODULE>-NNN (local reference; map to Number or an external-id field).")
ws.column_dimensions["A"].width=26; ws.column_dimensions["B"].width=96; ws.sheet_view.showGridLines=False

# Definition of Done
ws=wb.create_sheet("Definition of Done")
title(ws,"CONNECTION - DEFINITION OF DONE (applies to every story)",
 "From the Engagement Delivery Guidelines + SOW. A story is Done only when all gates pass.","")
r=4
for g in [
 "Acceptance criteria were reviewed and approved by the Product Owner BEFORE build began.",
 "All acceptance criteria are met and validated by the Product Owner (or delegate) at close.",
 "Built OOTB-first within the Rule of Three (Configuration / UI Policy / Flow Designer); any deviation cleared the Customization Council via the two-key decision.",
 "Peer configuration/code review complete.",
 "Integration touchpoints exercised in sub-production (if the story touches integrations).",
 "No P1 defects open against the story; P2s triaged.",
 "Configuration captured in an update set and promotable; no development left in production.",
 "Any deviation logged in the Governance Triage Log with scope/budget/upgrade impact and PCR status.",
 "Demo-ready: the story can be shown against real Connection data in the sprint demo."]:
    ws.cell(r,1,"DoD").font=bfont(b=True,c=NAVY); c=ws.cell(r,2,g); c.font=bfont(); c.alignment=Alignment(wrap_text=True,vertical="top"); c.border=box; ws.cell(r,1).border=box
    ws.row_dimensions[r].height=max(22,14+6*(len(g)//80)); r+=1
ws.column_dimensions["A"].width=8; ws.column_dimensions["B"].width=110; ws.sheet_view.showGridLines=False

# Epics
ws=wb.create_sheet("Epics")
title(ws,"CONNECTION - EPICS (import first -> rm_epic)","One epic per in-scope module. Stories reference these.","")
header(ws,["number","short_description","description","state"])
r=5; en=0
for mod,meta in sd.MODMETA.items():
    en+=1; cnt=sum(1 for s in sd.STORIES if s["module"]==mod)
    vals=[f"CONN-EPIC-{meta['abbr']}",f"{mod} - Phase 1 Configuration",
          f"Phase 1 OOTB configuration of {mod} for Connection - {cnt} default stories covering the contract-configurable decisions. {meta['sprint']} (primary).","Draft"]
    for c,v in enumerate(vals,1):
        cell=ws.cell(r,c,v); cell.font=bfont(); cell.alignment=Alignment(wrap_text=True,vertical="top"); cell.border=box
        if en%2==1: cell.fill=PatternFill("solid",fgColor=ALT)
    ws.row_dimensions[r].height=34; r+=1
for col,w in zip("ABCD",[16,40,80,10]): ws.column_dimensions[col].width=w
ws.freeze_panes="A5"; ws.sheet_view.showGridLines=False

na=story_sheet("Application User Stories","app")
ni=story_sheet("Integration User Stories","int")
# reorder: README, Epics, App, Int, DoD
order=["README - Import Guide","Epics","Application User Stories","Integration User Stories","Definition of Done"]
wb._sheets.sort(key=lambda x: order.index(x.title) if x.title in order else 99)
wb.save(OUT)
print(f"Saved: {OUT} | epics={en} | app stories={na} | integration stories={ni} | total={na+ni}")
