# -*- coding: utf-8 -*-
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
OUT="/sessions/eager-epic-davinci/mnt/everforth-ecs-practice/05_Clients/Connection/02_Delivery/Connection_Delivery_Readiness_Audit.xlsx"
NAVY="0B1F3A";WHITE="FFFFFF";ALT="F8FAFC";BORDER="E2E8F0";SLATE="475569"
GREEN="DCFCE7";YELLOW="FEF9C3";RED="FEE2E2";AMBER="FEF3C7"
thin=Side(style="thin",color=BORDER);box=Border(left=thin,right=thin,top=thin,bottom=thin)
def bf(c="1A1A1A",b=False,sz=10): return Font(name="Calibri",size=sz,bold=b,color=c)
wb=Workbook();ws=wb.active;ws.title="Delivery Readiness Audit"
ws["A1"]="CONNECTION - DELIVERY READINESS AUDIT (start to finish; ECS + customer lenses)"
ws["A1"].font=Font(name="Calibri",size=14,bold=True,color=NAVY)
ws["A2"]="Audited against SOW v2.0 across the full lifecycle. Status: Have | Partial | Gap. Priority: P1 (close before/at the relevant phase) | P2 | P3. The biggest delivery risk is customer dependencies (SOW Sec 6)."
ws["A2"].font=Font(name="Calibri",size=9,italic=True,color=SLATE)
H=["#","Phase","Lens","Capability / Artifact","Status","What exists today","Gap / what's missing","Delivery risk if unaddressed","Recommendation","Priority"]
hr=4
for c,h in enumerate(H,1):
    cell=ws.cell(hr,c,h);cell.font=Font(name="Calibri",size=11,bold=True,color=WHITE);cell.fill=PatternFill("solid",fgColor=NAVY)
    cell.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True);cell.border=box
R=[
# Phase, Lens, Capability, Status, What exists, Gap, Risk, Recommendation, Priority
("Setup / Sprint 0","Both","Team & client onboarding","Have","Onboarding guides, vision, role quick-ref, kickoff deck, checklist","-","-","Maintain","P3"),
("Setup / Sprint 0","Both","Governance operating model","Have","Governance Charter, Delivery Guidelines, Triage/RAID","-","-","Maintain","P3"),
("Setup / Sprint 0","ECS","Project plan & schedule","Have","18-week plan (Sprints 0-8); resource buildup tab","Start date (B3) not set","Dates won't flow until set","Set start date at kickoff","P3"),
("Setup / Sprint 0","Customer","Customer Responsibilities & Dependency Tracker (full engagement)","Have","Connection_Customer_Dependency_Tracker.xlsx (all SOW Sec 6 deps, owner/due/status/impact)","-","-","Maintain weekly; flag at-risk early","P1"),
("Setup / Sprint 0","Both","RACI matrix","Partial","Role & Accountability Quick-Ref (RACI-style)","No formal per-deliverable/per-activity RACI","Ownership ambiguity on specific deliverables","Derive a deliverable-level RACI cut","P2"),
("Setup / Sprint 0","Both","Communication plan","Partial","Cadence defined in setup story + guidelines","No standalone comms plan artifact","Minor - cadence is known","Optional one-pager","P3"),
("Build / Sprints 1-6","Customer","Workshop pre-reads (per module)","Gap","18 workshop decks + scope notes; library WP_01-17 exist","Connection-tailored client pre-reads not created","Cold workshops -> weak/slow decisions -> rework (facilitation model assumes pre-warmed decisions)","ADAPT from library WP_01-17 for in-scope modules","P2"),
("Build / Sprints 1-6","ECS","Workshop decks & facilitation","Have","18 client decks, scope notes, facilitation guide","-","-","Maintain","P3"),
("Build / Sprints 1-6","Customer","Accelerator data packs","Have","13 packs incl. Vonage; index README","-","-","Distribute on schedule","P3"),
("Build / Sprints 1-6","ECS","Configuration backlog (stories)","Have","141 config stories, SN Agile-ready, AC + DoD","-","-","Import to SN Agile","P2"),
("Build / Sprints 1-6","ECS","Project delivery backlog (non-dev)","Have","55 delivery stories across 9 work-streams","-","-","Import; track on dashboard","P3"),
("Build / Sprints 1-6","ECS","Sprint plan & capacity model","Gap","Stories carry target sprints; resource buildup tab","No capacity-vs-load check (do ~196 stories fit team velocity over 6 sprints?)","Hidden overload -> mid-engagement slippage","BUILD - sprint-by-sprint capacity vs story-point load","P2"),
("Build / Sprints 1-6","ECS","Definition of Ready","Gap","Definition of Done exists","No DoR (when a story may enter a sprint)","Half-baked stories enter sprints -> churn","Add a short DoR (1 page / tab)","P3"),
("Build / Sprints 1-6","ECS","Sprint ceremonies (planning, retro)","Partial","Sprint Demo template exists","No planning agenda or retrospective template","Inconsistent ceremonies","Add planning + retro templates","P3"),
("Build / Sprints 1-6","ECS","Per-module demo scripts","Gap","Sprint Demo shell; library INT-DS demo scripts exist","Connection demo scripts not created","Inconsistent demos; weak sign-off moments","ADAPT library demo scripts for in-scope modules","P3"),
("Build / Sprints 1-6","Both","Decision register (running)","Partial","Triage Log (deviations) + sprint workbooks","No general open-decisions register beyond deviations","Decisions lost between workshops","Add a decisions log (or extend RAID)","P3"),
("Build / Sprints 1-6","Both","Story/sprint acceptance & sign-off","Partial","DoD requires PO sign-off; SOW = 3-day acceptance window","No acceptance/sign-off log operationalizing the window","Acceptance drift -> PCR triggers missed","BUILD a lightweight acceptance/sign-off log","P2"),
("Build / Sprints 1-6","ECS","Design documentation (per sprint)","Partial","Delivery story + library sprint workbooks","Connection sprint-workbook copies not yet created","Decisions not captured -> KT gaps","Copy sprint workbooks into Connection as decisions land","P3"),
("Testing","ECS","UAT end-to-end scripts + guidebook","Have","18 E2E scripts, story traceability, guidebook, defect log","-","-","Assign testers at UAT","P2"),
("Testing","ECS","SIT / integration test scripts & test-data plan","Gap","SIT delivery story exists","No SIT script pack or test-data strategy","Integration defects surface late (AD/SSO, SCCM, Intune, CTI)","BUILD SIT scripts + a test-data plan","P2"),
("Go-Live / Cutover","ECS","Cutover Runbook","Have","Connection_Cutover_Runbook.docx (sequence, owners, validation, rollback, comms)","-","-","Finalize specifics in Sprint 7","P1"),
("Go-Live / Cutover","Both","Go-Live Readiness Checklist (Connection)","Have","Connection_Go_Live_Readiness_Checklist.xlsx (gated go/no-go criteria)","-","-","Work it through Stage 3","P1"),
("Go-Live / Cutover","Both","Go-Live readiness sign-off","Partial","Delivery story + DoD","No sign-off form","Authorization not formally captured","Add a sign-off form","P2"),
("Hypercare / Close","Both","Operational Handoff Pack","Gap","Library CLT-CO-03; delivery story","Ownership matrix, support model, escalation not built for Connection","Unclear steady-state ownership post-Hypercare","BUILD from library + RACI","P2"),
("Hypercare / Close","Both","Hypercare support model & exit report","Partial","Library CLT-CO-03/06; delivery story","Connection copies not created","Handover ambiguity","Copy + tailor near Go-Live","P3"),
("Hypercare / Close","ECS","KT Package (Admin + Train-the-Trainer)","Have","Admin Guide & KT, Train-the-Trainer Toolkit","-","-","Deliver at Go-Live","P3"),
("Hypercare / Close","Both","Lessons Learned & Closeout + 12-mo roadmap","Partial","Library + delivery stories","Connection copies not created","Close not documented","Produce at close","P3"),
("Governance / Scope","Both","PCR (Project Change Request) template","Gap","SOW Sec 9 PCR process referenced in guidelines","No PCR request form/template","Scope changes uncontrolled; SOW ties acceptance delays & customization #6 to PCR","BUILD a PCR form + log","P2"),
("Governance / Scope","ECS","Assumptions & Out-of-Scope register","Gap","SOW Sec 7-8 text","No tracker to manage assumptions/out-of-scope","Scope creep; assumptions unverified","Add a register (or tab in RAID)","P3"),
("Governance / Scope","ECS","Customization variance tracking","Partial","Triage/RAID has Triage Log + cap counter","Library INT-TBV-03 variance tracker (effort %) not copied","Variance % not tracked to capacity","Optional - add variance tab","P3"),
("Governance / Scope","Both","Status report, exec dashboard, deliverables matrix","Have","Weekly status, Exec Dashboard v2 (delivery metrics), SOW matrix","-","-","Run the cadence","P3"),
]
r=hr+1
for i,row in enumerate(R,1):
    vals=[i]+list(row)
    for c,v in enumerate(vals,1):
        cell=ws.cell(r,c,v);cell.font=bf();cell.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True);cell.border=box
        if i%2==0: cell.fill=PatternFill("solid",fgColor=ALT)
    ws.row_dimensions[r].height=max(40,14+6*(max(len(str(row[5])),len(str(row[6])),len(str(row[7])))//40))
    r+=1
last=r-1
# status col E (5), priority col J (10)
dvS=DataValidation(type="list",formula1='"Have,Partial,Gap"',allow_blank=True);ws.add_data_validation(dvS);dvS.add(f"E{hr+1}:E{last}")
dvP=DataValidation(type="list",formula1='"P1,P2,P3"',allow_blank=True);ws.add_data_validation(dvP);dvP.add(f"J{hr+1}:J{last}")
for v,col in [("Have",GREEN),("Partial",YELLOW),("Gap",RED)]:
    ws.conditional_formatting.add(f"E{hr+1}:E{last}",CellIsRule(operator="equal",formula=[f'"{v}"'],fill=PatternFill("solid",fgColor=col)))
ws.conditional_formatting.add(f"J{hr+1}:J{last}",CellIsRule(operator="equal",formula=['"P1"'],fill=PatternFill("solid",fgColor=RED)))
ws.conditional_formatting.add(f"J{hr+1}:J{last}",CellIsRule(operator="equal",formula=['"P2"'],fill=PatternFill("solid",fgColor=AMBER)))
# summary
sr=last+2
ws.cell(sr,1,"SUMMARY").font=Font(name="Calibri",size=11,bold=True,color=NAVY)
sums=[("Total items",f"=COUNTA(D{hr+1}:D{last})"),("Have",f'=COUNTIF(E{hr+1}:E{last},"Have")'),("Partial",f'=COUNTIF(E{hr+1}:E{last},"Partial")'),("Gap",f'=COUNTIF(E{hr+1}:E{last},"Gap")'),("P1 (close first)",f'=COUNTIF(J{hr+1}:J{last},"P1")'),("P2",f'=COUNTIF(J{hr+1}:J{last},"P2")')]
for i,(lbl,f) in enumerate(sums):
    ws.cell(sr+1+i,1,lbl).font=bf(c=SLATE);ws.cell(sr+1+i,2,f).font=bf(b=True,c=NAVY)
for c,w in zip("ABCDEFGHIJ",[4,16,10,30,9,34,40,40,34,9]): ws.column_dimensions[c].width=w
ws.freeze_panes=f"A{hr+1}";ws.sheet_view.showGridLines=False
wb.save(OUT);print("Saved:",OUT,"items:",len(R))
