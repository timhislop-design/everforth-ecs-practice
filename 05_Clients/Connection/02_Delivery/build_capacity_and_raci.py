# -*- coding: utf-8 -*-
import sys; sys.path.insert(0,".")
import userstories_data as cfg, deliverystories_data as dlv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
NAVY="0B1F3A";WHITE="FFFFFF";ALT="F8FAFC";BORDER="E2E8F0";SLATE="475569";CYAN="ECFEFF"
GREEN="DCFCE7";YELLOW="FEF9C3";RED="FEE2E2";AMBER="FEF3C7";BLUE="DBEAFE"
thin=Side(style="thin",color=BORDER);box=Border(left=thin,right=thin,top=thin,bottom=thin)
def bf(c="1A1A1A",b=False,sz=10): return Font(name="Calibri",size=sz,bold=b,color=c)
def header(ws,cols,hr=4):
    for c,h in enumerate(cols,1):
        cell=ws.cell(hr,c,h);cell.font=Font(name="Calibri",size=11,bold=True,color=WHITE);cell.fill=PatternFill("solid",fgColor=NAVY)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True);cell.border=box
    ws.row_dimensions[hr].height=28

# aggregate story points per sprint
def normsprint(s):
    s=str(s)
    if s.startswith("Sprint "): return s
    return "Ongoing / All"
load={}
for s in cfg.STORIES:
    sp=normsprint(cfg.MODMETA[s["module"]]["sprint"]); load.setdefault(sp,[0,0]); load[sp][0]+=s["points"]
for s in dlv.STORIES:
    sp=normsprint(s["sprint"]); load.setdefault(sp,[0,0]); load[sp][1]+=s["points"]
order=["Sprint 0","Sprint 1","Sprint 2","Sprint 3","Sprint 4","Sprint 5","Sprint 6","Sprint 7","Sprint 8","Ongoing / All"]

# ===== Sprint Plan & Capacity Model =====
OUT="/sessions/eager-epic-davinci/mnt/everforth-ecs-practice/05_Clients/Connection/02_Delivery/Connection_Sprint_Plan_and_Capacity.xlsx"
wb=Workbook();ws=wb.active;ws.title="Sprint Capacity"
ws["A1"]="CONNECTION - SPRINT PLAN & CAPACITY MODEL"
ws["A1"].font=Font(name="Calibri",size=14,bold=True,color=NAVY)
ws["A2"]="Load = story points planned per sprint (config + delivery backlog). Set Capacity (pts) per sprint from team velocity; the model flags over-loaded sprints. Adjust scope/sequence before a sprint goes red."
ws["A2"].font=Font(name="Calibri",size=9,italic=True,color=SLATE)
header(ws,["Sprint","Stage","Config Load (pts)","Delivery Load (pts)","Total Load (pts)","Capacity (pts)","Load vs Capacity","Status"])
stage={"Sprint 0":"Stage 1","Sprint 1":"Stage 1","Sprint 2":"Stage 1","Sprint 3":"Stage 2","Sprint 4":"Stage 2","Sprint 5":"Stage 2","Sprint 6":"Stage 3","Sprint 7":"Stage 3","Sprint 8":"Stage 4","Ongoing / All":"All"}
r=5
for sp in order:
    cf,dl=load.get(sp,[0,0])
    ws.cell(r,1,sp).font=bf(b=True); ws.cell(r,2,stage[sp]).font=bf()
    ws.cell(r,3,cf).font=bf(); ws.cell(r,4,dl).font=bf()
    ws.cell(r,5,f"=C{r}+D{r}").font=bf(b=True,c=NAVY)
    ws.cell(r,6,30 if sp.startswith("Sprint") and sp not in("Sprint 0","Sprint 8") else (15 if sp in("Sprint 0","Sprint 8") else 0)).font=bf(c="0000FF")  # editable capacity (blue=input)
    ws.cell(r,7,f'=IF(F{r}=0,"-",E{r}/F{r})').font=bf(); ws.cell(r,7).number_format="0%"
    ws.cell(r,8,f'=IF(F{r}=0,"n/a",IF(E{r}>F{r},"OVER",IF(E{r}>0.85*F{r},"TIGHT","OK")))').font=bf(b=True)
    for c in range(1,9):
        ws.cell(r,c).border=box; ws.cell(r,c).alignment=Alignment(horizontal="center",vertical="center")
        if (r%2)==0: ws.cell(r,c).fill=PatternFill("solid",fgColor=ALT)
    r+=1
last=r-1
ws.cell(r,1,"TOTAL").font=bf(b=True,c=NAVY)
ws.cell(r,3,f"=SUM(C5:C{last})").font=bf(b=True,c=NAVY); ws.cell(r,4,f"=SUM(D5:D{last})").font=bf(b=True,c=NAVY); ws.cell(r,5,f"=SUM(E5:E{last})").font=bf(b=True,c=NAVY); ws.cell(r,6,f"=SUM(F5:F{last})").font=bf(b=True,c=NAVY)
for v,col in [("OVER",RED),("TIGHT",AMBER),("OK",GREEN)]:
    ws.conditional_formatting.add(f"H5:H{last}",CellIsRule(operator="equal",formula=[f'"{v}"'],fill=PatternFill("solid",fgColor=col)))
ws.cell(r+2,1,"Capacity (blue) is editable: set from team size x velocity. OVER = load exceeds capacity (rebalance); TIGHT = >85%.").font=bf(c=SLATE)
for c,w in zip("ABCDEFGH",[14,10,16,16,16,14,16,10]): ws.column_dimensions[c].width=w
ws.freeze_panes="A5"; ws.sheet_view.showGridLines=False
wb.save(OUT); print("Saved:",OUT)

# ===== RACI Matrix =====
OUT2="/sessions/eager-epic-davinci/mnt/everforth-ecs-practice/05_Clients/Connection/02_Delivery/Connection_RACI_Matrix.xlsx"
wb=Workbook();ws=wb.active;ws.title="RACI"
ws["A1"]="CONNECTION - RACI MATRIX (deliverable & activity level)"
ws["A1"].font=Font(name="Calibri",size=14,bold=True,color=NAVY)
ws["A2"]="R=Responsible (does it), A=Accountable (owns the outcome, one per row), C=Consulted, I=Informed. ECS roles + Connection roles."
ws["A2"].font=Font(name="Calibri",size=9,italic=True,color=SLATE)
roles=["EM","SA","PC","TC","Prac.Lead","Sponsor","Prod.Owner","Cust.PM","Tech.Lead","SMEs"]
header(ws,["Activity / Deliverable"]+roles)
RA=[
("Engagement governance & status reporting","A","I","I","I","C","I","I","C","I","I"),
("Customization Council / two-key decisions","A","R","C","I","R","R","C","I","I","I"),
("Project plan & schedule","A","C","C","I","I","I","C","R","I","I"),
("Customer dependency management","A","I","I","I","I","C","C","R","C","C"),
("Workshops facilitation & sign-off","C","C","R","I","I","I","A","C","C","R"),
("OOTB vs customization decisions (Rule of Three)","C","R","C","C","A","C","C","I","I","I"),
("Platform architecture & CSDM alignment","I","A","C","C","C","I","C","I","R","C"),
("Configuration build (per story)","I","C","C","R","I","I","A","I","C","C"),
("Integrations & SGC build","I","A","I","R","C","I","I","I","R","C"),
("Vonage CTI / Interactions build","I","A","I","R","C","I","I","I","C","R"),
("Accelerator pack data completion","I","C","C","I","I","I","A","R","C","R"),
("Data quality & validation","I","C","C","C","I","I","A","R","C","R"),
("Story acceptance (3-day window)","C","C","C","I","I","I","A","C","I","C"),
("Governance Triage Log disposition","R","C","I","I","C","A","C","I","I","I"),
("SIT execution","I","C","I","R","A","I","I","I","C","I"),
("UAT coordination & execution","C","C","C","I","A","I","C","R","I","R"),
("Defect triage & resolution","C","A","C","R","C","I","C","I","I","I"),
("Go-Live readiness & sign-off","R","C","C","I","C","A","R","C","C","I"),
("Cutover execution","A","R","C","R","C","I","I","C","R","I"),
("Admin KT & Train-the-Trainer","I","R","R","C","I","I","C","C","C","R"),
("Knowledge content creation (ongoing)","I","I","C","I","I","I","A","R","I","R"),
("End-user comms & adoption","I","I","C","I","I","C","A","R","I","C"),
("Hypercare support","A","C","C","R","C","I","C","C","R","I"),
("Operational handoff & ownership","A","C","I","I","C","R","C","R","R","I"),
("Lessons learned & closeout","A","C","C","I","C","C","C","C","I","I"),
]
r=5
fillmap={"R":BLUE,"A":GREEN,"C":YELLOW,"I":ALT}
for i,row in enumerate(RA,1):
    ws.cell(r,1,row[0]).font=bf(); ws.cell(r,1).alignment=Alignment(wrap_text=True,vertical="center"); ws.cell(r,1).border=box
    for c in range(2,12):
        val=row[c-1]; cell=ws.cell(r,c,val); cell.font=bf(b=True,c=NAVY); cell.alignment=Alignment(horizontal="center",vertical="center"); cell.border=box
        cell.fill=PatternFill("solid",fgColor=fillmap.get(val,WHITE))
    r+=1
ws.column_dimensions["A"].width=42
for c in "BCDEFGHIJK": ws.column_dimensions[c].width=9
ws.freeze_panes="B5"; ws.sheet_view.showGridLines=False
# legend
lr=r+1
ws.cell(lr,1,"Legend:  R = Responsible   A = Accountable   C = Consulted   I = Informed").font=bf(b=True,c=NAVY)
ws.cell(lr+1,1,"Roles: EM=Engagement Mgr, SA=Solution Architect, PC=Process Consultant, TC=Technical Consultant, Prac.Lead=ECS Practice Lead; Sponsor/Prod.Owner/Cust.PM/Tech.Lead/SMEs = Connection").font=bf(c=SLATE)
wb.save(OUT2); print("Saved:",OUT2)
