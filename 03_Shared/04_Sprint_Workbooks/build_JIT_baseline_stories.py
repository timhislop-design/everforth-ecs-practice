"""
Build ECS_JIT_Baseline_Stories.xlsx
Single workbook, one tab per process area, covering all 18-week sprint modules.

Stories are DECISION-POINT DRIVEN — one story per key workshop decision,
NOT basic configuration stories (no "as a user I want to log in" stories).

Each story has:
  Story ID | Decision Point | User Story | JIT Trigger | Acceptance Criteria |
  OOTB Defense Language | Story Points | Status

Tabs: README + one per module (15 modules across 6 sprints)
"""
import sys, os
HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(HERE, "..", ".."))
TEMPLATES = os.path.join(REPO, "00_Templates_and_Branding")
sys.path.insert(0, os.path.join(REPO, "03_Shared", "00_Templates_and_Branding"))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(HERE, "ECS_JIT_Baseline_Stories.xlsx")

# ── Brand palette ──────────────────────────────────────────────────────────────
NAVY    = "0B1F3A"
WHITE   = "FFFFFF"
TEAL    = "14B8A6"
AMBER   = "FFF3CD"
AMBER_T = "92400E"
SLATE   = "475569"
ALT     = "F8FAFC"
BORDER  = "E2E8F0"
CYAN_BG = "ECFEFF"
BODY    = "1A1A1A"
FONT    = "Arial"

def _font(color=BODY, bold=False, size=10):
    return Font(name=FONT, color=color, bold=bold, size=size)

def _fill(hex_color):
    return PatternFill(fill_type="solid", start_color=hex_color, end_color=hex_color)

THIN = Border(
    left=Side(style="thin", color=BORDER), right=Side(style="thin", color=BORDER),
    top=Side(style="thin", color=BORDER), bottom=Side(style="thin", color=BORDER),
)

def banner(ws, row, text, sub=False, cols=8):
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(NAVY)
    c.font = _font(WHITE, bold=not sub, size=13 if not sub else 9)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28 if not sub else 18
    for col in range(2, cols+1):
        x = ws.cell(row=row, column=col)
        x.fill = _fill(NAVY)

def header_row(ws, row, cols):
    for i, h in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = _fill(NAVY)
        c.font = _font(WHITE, bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN
    ws.row_dimensions[row].height = 32

def story_row(ws, row, values, alt=False):
    bg = ALT if alt else WHITE
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.fill = _fill(bg)
        c.font = _font(BODY, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = THIN
    # Story ID col — bold navy
    ws.cell(row=row, column=1).font = _font(NAVY, bold=True, size=10)
    # Status col (last) — center
    ws.cell(row=row, column=8).alignment = Alignment(horizontal="center", vertical="top")
    ws.row_dimensions[row].height = max(60, 14 * max(1, max(len(str(v)) for v in values) // 60))

COLS = ["Story ID", "Decision Point", "User Story (JIT — Decision-Driven)",
        "JIT Trigger (activate when…)", "Acceptance Criteria",
        "OOTB Defense Language", "Points", "Status"]

COL_WIDTHS = [10, 28, 42, 30, 42, 42, 7, 10]

def set_cols(ws, widths=COL_WIDTHS):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, row=5):
    ws.freeze_panes = ws.cell(row=row, column=1)


# ── Story data ─────────────────────────────────────────────────────────────────
# Format per story tuple:
# (id, decision_point, user_story, jit_trigger, acceptance_criteria, ootb_defense, points, status)

STORIES = {

  "PLT — Sprint 1\nPlatform Foundation": [
    ("PLT-001", "4.1 Environment Strategy",
     "As a ServiceNow Admin, I want the Dev/Test/Prod three-environment model configured so that changes are promoted safely without exposing production data to development risk.",
     "Customer confirms 3-environment split (not 2).",
     "• Three distinct instances registered and linked in the Update Set promotion chain.\n• Dev → Test → Prod promotion path is documented and validated with a test Update Set.\n• Cross-environment data bleed test passes (no prod data visible in Dev).",
     "'A 2-environment model saves licensing cost but collapses the safety net. Every customer who has gone 2-env has asked for a 3rd within 6 months after a bad test promotion hits prod. We front-load that cost now.'",
     3, "Baseline"),

    ("PLT-002", "4.2 Access & Security",
     "As a Security Admin, I want OOTB role groups (itil, admin, approver_user) mapped to customer AD groups so that access is governed by HR lifecycle, not manual ServiceNow user management.",
     "Customer confirms AD group naming convention for role mapping.",
     "• itil, admin, itil_admin, approver_user roles assigned only via AD group membership.\n• No local ServiceNow role assignments for named individuals (admin accounts excepted).\n• User added to AD group → ServiceNow access granted within 1 SSO session.",
     "'Manual role assignment in ServiceNow creates a shadow access list that HR offboarding never touches. AD-driven roles mean when someone leaves the company, their ServiceNow access is revoked automatically.'",
     3, "Baseline"),

    ("PLT-003", "4.3 User & Group Structure",
     "As a Process Owner, I want OOTB group-based assignment (not individual assignment) configured as the default so that ticket routing survives staff turnover.",
     "Customer confirms assignment group taxonomy (at least Tier 1/Tier 2/Tier 3 support structure).",
     "• All assignment_group fields on Incident, Request, Change are mandatory — individual assigned_to is optional.\n• At least 3 assignment groups created with named managers.\n• Test incident routed to correct group without requiring a specific assignee.",
     "'Individual assignment models break every time someone goes on leave or leaves the company. Group assignment is the OOTB default for a reason — it makes the queue the unit of work, not the person.'",
     2, "Baseline"),

    ("PLT-004", "4.4 Authentication (SSO)",
     "As a Security Admin, I want OOTB SAML 2.0 SSO configured with the customer's IdP so that users never hold a separate ServiceNow credential.",
     "Customer confirms IdP metadata URL and SSO admin contact available for configuration session.",
     "• SAML SSO login path is the only active login method for non-admin accounts.\n• Local login disabled for all non-emergency admin accounts.\n• Test user authenticates via IdP without being prompted for a ServiceNow password.\n• SSO fallback (break-glass admin) documented and access-controlled.",
     "'Local credentials in ServiceNow are a separate attack surface and a separate offboarding action. SSO means one credential, one offboarding step, one audit trail.'",
     3, "Baseline"),

    ("PLT-005", "4.5 Deployment Model",
     "As a ServiceNow Admin, I want OOTB Update Sets as the sole change deployment mechanism so that all platform changes are tracked, reversible, and promotable.",
     "Customer agrees no direct production changes are permitted post-go-live.",
     "• Update Set capture is enforced in Dev (Update Set auto-creation enabled).\n• Promotion from Dev → Test → Prod requires a named approver sign-off in the Update Set log.\n• At least one test Update Set promoted end-to-end before go-live.\n• Rollback procedure documented for failed promotions.",
     "'Direct production changes are invisible to the audit trail and unrecoverable when something goes wrong. Update Sets are the OOTB mechanism for exactly this — they are not overhead, they are the safety net.'",
     2, "Baseline"),

    ("PLT-006", "4.6 Notifications",
     "As a Process Owner, I want the OOTB notification baseline (assignment, escalation, resolution) configured with no custom templates so that notification maintenance stays with the platform team, not development.",
     "Customer confirms which notification triggers are in scope for MVP (assignment / escalation / resolution / SLA breach).",
     "• OOTB Notification records used without HTML customization beyond logo/signature.\n• Test notification sent for each trigger type and confirmed received.\n• No custom notification scripts or Flow Designer notification actions created at MVP.",
     "'Custom notification templates require a developer every time the wording needs to change. OOTB notifications with a clean subject line and signature are what 90% of customers actually need — and they are maintainable by the admin, not a developer.'",
     2, "Baseline"),

    ("PLT-007", "4.7 Naming & Governance",
     "As a ServiceNow Admin, I want a documented naming standard enforced for all OOTB configuration objects (groups, categories, update sets, notification names) so that the instance remains navigable after the engagement ends.",
     "Customer confirms prefix convention and governance owner before any configuration objects are created.",
     "• Naming standard document signed off by customer admin before Sprint 1 config begins.\n• At least one naming review checkpoint per sprint included in the sprint retro.\n• All Update Sets, groups, and notification records created in Sprint 1 conform to the agreed standard.",
     "'An uncontrolled naming convention in Sprint 1 becomes 18 sprints of inconsistency. The standard takes 30 minutes to agree and saves weeks of cleanup.'",
     1, "Baseline"),
  ],

  "INC — Sprint 1\nIncident Management": [
    ("INC-001", "4.1 Incident Intake",
     "As a Service Desk Agent, I want OOTB intake channels (portal self-service, email-to-incident, phone-logged manual entry) configured so that all intake paths feed the same incident queue without custom routing logic.",
     "Customer confirms which intake channels are in scope at MVP.",
     "• Portal self-service form creates incident with correct category, contact, and urgency.\n• Email-to-incident rule creates incident from designated mailbox with no custom parsing.\n• Manually logged incidents require only mandatory OOTB fields.\n• All three paths create incidents with identical mandatory field set.",
     "'Custom intake parsing scripts break every time the email format changes. OOTB email-to-incident is designed for exactly this — it captures the essentials and lets the agent enrich the record.'",
     3, "Baseline"),

    ("INC-002", "4.2 Lifecycle",
     "As a Service Desk Manager, I want the OOTB incident lifecycle (New → In Progress → Resolved → Closed) enforced without custom states so that reporting on MTTR is accurate and consistent.",
     "Customer confirms they will not add custom states (e.g., 'Pending Vendor', 'On Hold') without a documented business justification.",
     "• State field uses only OOTB values: New, In Progress, On Hold, Resolved, Closed.\n• On Hold requires a valid hold reason and sets a resume date.\n• Resolved incidents auto-close after the agreed closure window (default 3 days).\n• MTTR report is accurate and matches the state transitions in the audit log.",
     "'Every custom state is a reporting hole. OOTB uses On Hold with a reason code for everything customers typically want a custom state for — vendor dependency, awaiting customer response, scheduled maintenance.'",
     2, "Baseline"),

    ("INC-003", "4.3 Priority Model",
     "As a Service Desk Manager, I want the OOTB priority matrix (Impact × Urgency = Priority) configured and enforced so that priority is calculated consistently, not subjectively set by agents.",
     "Customer confirms Impact and Urgency value definitions before configuration.",
     "• Priority field is read-only and auto-calculated from Impact × Urgency.\n• Impact and Urgency definitions are documented and visible on the incident form.\n• Sample priority matrix test: High Impact + High Urgency = Priority 1; confirmed in 5 test incidents.\n• Agents cannot override the calculated priority without Change Manager approval.",
     "'Agent-set priority creates a priority inflation problem within weeks — everyone's incident becomes P1. The matrix removes subjectivity and makes the priority queue trustworthy.'",
     2, "Baseline"),

    ("INC-004", "4.4 Assignment Model",
     "As a Service Desk Manager, I want OOTB assignment rules to route incidents to the correct group based on category/subcategory so that dispatchers are not needed for standard routing.",
     "Customer confirms category/subcategory taxonomy and group mapping before assignment rules are configured.",
     "• Assignment rules configured for top 10 category/subcategory combinations.\n• Test incidents routed correctly to the right group without dispatcher intervention.\n• Unmatched categories route to a named default 'Service Desk' group, not left unassigned.\n• Assignment rule audit report shows >80% auto-assignment rate after 2-week pilot.",
     "'Manual dispatching is the most expensive way to route a ticket. Assignment rules are an OOTB feature that pays back within the first sprint if the category taxonomy is clean.'",
     3, "Baseline"),

    ("INC-005", "4.5 SLAs",
     "As a Service Owner, I want OOTB SLA definitions configured for each priority tier so that breach risk is visible in real time without custom dashboard development.",
     "Customer confirms SLA target times per priority tier (P1 response/resolution, P2 response/resolution, etc.).",
     "• SLA Definition records created for P1, P2, P3, P4 response and resolution.\n• SLA timer starts on incident creation; pause conditions configured for On Hold state.\n• SLA breach indicator visible on incident form and in the service desk queue view.\n• OOTB SLA report shows compliance % by priority for the test pilot period.",
     "'Custom SLA timers are a development project. OOTB SLA definitions with pause/stop conditions handle every scenario customers typically want to customize — they just need the right conditions configured.'",
     3, "Baseline"),

    ("INC-006", "4.6 Notifications",
     "As an Incident Assignee, I want OOTB incident notifications (assigned, escalated, resolved, reopened) configured without custom email templates so that the team is informed without a notification sprawl problem.",
     "Customer confirms which notification events are in scope and who receives each.",
     "• Assigned notification sent to assignee within 60 seconds of assignment.\n• Escalation notification sent to group manager when SLA is at 75% elapsed.\n• Resolution confirmation sent to caller with satisfaction survey link.\n• No more than 6 notification triggers configured at MVP — excess requests deferred to Phase 2 review.",
     "'Notification sprawl is the fastest way to train people to ignore notifications. Six well-timed notifications beat thirty that no one reads. We add more in Phase 2 based on what operators actually want.'",
     2, "Baseline"),
  ],

  "CAT — Sprint 2\nService Catalog & Request": [
    ("CAT-001", "4.1 Catalog Structure",
     "As a Portal User, I want the service catalog organized into no more than 8 top-level categories using OOTB category records so that users can find what they need without a search.",
     "Customer confirms consolidated category list (max 8) before any catalog items are created.",
     "• Catalog has ≤ 8 top-level categories.\n• No category has fewer than 3 items (merge or remove).\n• Category names are role-agnostic (e.g., 'IT Equipment' not 'IMAC Requests from IT').\n• All categories visible and correctly sorted in the Employee Center portal.",
     "'Most customers start with 20+ categories because every team wanted their own. Users don't browse by team — they browse by need. Eight categories is the usability threshold where portal adoption stays healthy.'",
     2, "Baseline"),

    ("CAT-002", "4.2 Catalog Items (Top 10–15)",
     "As a Service Owner, I want the top 10–15 highest-volume request types built as OOTB catalog items (not custom forms) so that the catalog delivers immediate value without development debt.",
     "Customer confirms the list of top 10–15 items by volume (pull from legacy ITSM or service desk ticket data).",
     "• Each item uses OOTB Record Producer or Catalog Item — no custom form builders.\n• Each item has: clear name, description, icon, category assignment, and at least one variable.\n• Submitted request creates an RITM linked to a parent REQ.\n• All 10–15 items tested end-to-end (submit → fulfillment group receives → resolved) before go-live.",
     "'Building 50 catalog items before go-live is how catalog projects fail. The top 15 by volume cover 80% of requests. Build those right, prove the model works, then add the rest.'",
     5, "Baseline"),

    ("CAT-003", "4.3 Request Data Capture",
     "As a Fulfillment Agent, I want catalog item variables to capture only the data needed to fulfill the request (not every possible field) so that the requester experience is fast and the fulfillment data is complete.",
     "Customer confirms the minimum required fields per item with the fulfillment group, not just the requester team.",
     "• Each catalog item has ≤ 10 variables (if more are needed, escalate to an approval for data collection).\n• Mandatory variables are limited to the minimum required to route and fulfill — all others optional.\n• Form tested with 5 requesters who are not IT staff; completion rate ≥ 90% without help.",
     "'Long catalog forms kill portal adoption. If fulfillment needs more data, the fulfillment agent calls the requester — that is more efficient than a 20-field form that 40% of users abandon.'",
     2, "Baseline"),

    ("CAT-004", "4.4 Approval Model",
     "As a Manager, I want OOTB catalog approval (Manager approval → Department Head for high-value items) configured without custom approval scripts so that approvals are auditable and maintainable.",
     "Customer confirms which items require approval and who approves at each tier.",
     "• Items requiring approval use OOTB Approval Action in the catalog flow.\n• Manager approver is derived from the requester's manager field (sys_user.manager) — not hardcoded.\n• High-value items (>$500 or defined threshold) route to Department Head via second approval stage.\n• Approved/Rejected notification sent within 60 seconds of approval decision.",
     "'Hardcoded approver names in catalog flows break every time someone changes roles. Manager-derived approvals from the user record are maintainable by HR, not by a developer.'",
     3, "Baseline"),

    ("CAT-005", "4.5 Fulfillment Routing",
     "As a Fulfillment Manager, I want catalog items to route automatically to the correct fulfillment group based on item category so that no dispatcher is needed for standard catalog routing.",
     "Customer confirms fulfillment group per catalog item category.",
     "• Every catalog item has an assignment_group pre-filled at submission — no unassigned RITMs.\n• Routing is category-based via OOTB catalog flow, not scripted per-item.\n• Fulfillment group manager receives a queue notification when a new RITM arrives.\n• 100% of test RITMs reach the correct group within 2 minutes of submission.",
     "'Per-item routing scripts are the most common catalog technical debt pattern. Category-level routing means adding a new item to a category automatically inherits the routing — no developer required.'",
     2, "Baseline"),

    ("CAT-006", "4.6 User Experience",
     "As a Portal User, I want the Employee Center catalog integrated as the single entry point (not a separate catalog URL) so that users find requests and knowledge from the same place.",
     "Customer confirms Employee Center is the primary portal (not the legacy Service Portal or a custom portal).",
     "• Catalog is accessed only via Employee Center — no separate /esc or /sp bookmark promoted to users.\n• Employee Center home page shows the top 5 most-requested items as quick links.\n• Search in Employee Center returns catalog items alongside knowledge articles.\n• Mobile access tested and confirmed functional for top 5 catalog items.",
     "'Two portals split user adoption and double the maintenance burden. Employee Center is the investment platform — it is where Now Assist, Virtual Agent, and Knowledge converge. One portal.'",
     2, "Baseline"),
  ],

  "EC — Sprint 2\nEmployee Center": [
    ("EC-001", "4.1 Homepage Design",
     "As a Portal User, I want the Employee Center homepage configured with the OOTB layout (hero banner, quick actions, announcements, top topics) without custom widgets so that the portal is maintainable by the comms team, not developers.",
     "Customer confirms homepage content owners (who writes the announcements, who manages top topics).",
     "• Homepage uses only OOTB Employee Center widgets — no custom HTML widgets.\n• Hero banner is configured with customer branding (logo, colors) via the OOTB theme builder.\n• Top 5 quick actions are the 5 highest-volume catalog items.\n• Content owner can update announcements without ECS or developer involvement after go-live.",
     "'Custom homepage widgets are a support liability. The OOTB widget library covers the homepage needs of 95% of customers — and ServiceNow invests in them with every release. Custom widgets get orphaned.'",
     3, "Baseline"),

    ("EC-002", "4.2 Navigation Structure",
     "As a Portal User, I want Employee Center navigation organized by topic (not by department) so that users find content based on their need, not their knowledge of IT org structure.",
     "Customer confirms topic taxonomy (max 6 top-level topics) before any content is created.",
     "• Navigation has ≤ 6 top-level topics.\n• Topics are user-need-centric (e.g., 'My Equipment', 'HR & Benefits', 'IT Help') not org-centric.\n• No topic has fewer than 3 content items.\n• Navigation tested with 5 non-IT users; all locate a target item within 2 clicks.",
     "'Department-based navigation means users need to know which team owns what before they can find it. Topic-based navigation works for users who just know what they need.'",
     2, "Baseline"),

    ("EC-003", "4.3 Content Strategy",
     "As a Knowledge Manager, I want Employee Center content limited to knowledge articles and catalog items (no custom static pages) at MVP so that all content is searchable, version-controlled, and governed.",
     "Customer confirms they will not request custom HTML pages or embedded iframes at MVP.",
     "• All content in Employee Center at MVP is either a KB article or a catalog item.\n• No custom page types or embedded external content at MVP.\n• Content governance owner named and assigned before go-live.\n• Stale content review process (quarterly) documented and assigned.",
     "'Custom static pages in the portal become orphaned documentation that no one maintains. Knowledge articles have workflows, version history, and feedback loops. They are the right vehicle for portal content.'",
     2, "Baseline"),

    ("EC-004", "4.4 Search Behavior",
     "As a Portal User, I want unified search across catalog items and knowledge articles configured so that users get relevant results without switching between search contexts.",
     "Customer confirms which content types are included in search scope at MVP.",
     "• Employee Center search returns KB articles and catalog items in a unified result set.\n• Top 20 search terms tested and confirmed returning relevant first-page results.\n• 'Did you mean' and spell-check enabled (OOTB Search Suggestions).\n• Catalog items with zero views after 30 days flagged for relevance review.",
     "'Split search — one for catalog, one for knowledge — is the single biggest usability complaint in portal implementations. Unified search is OOTB; it requires no custom development.'",
     2, "Baseline"),

    ("EC-005", "4.5 Branding",
     "As a Communications Manager, I want Employee Center branded using the OOTB Theme Builder (logo, colors, font) without custom CSS so that branding updates can be made by the comms team without developer involvement.",
     "Customer provides brand hex colors, logo (SVG/PNG), and confirms no custom CSS is needed for the approved design.",
     "• Logo, primary color, and secondary color configured via OOTB Theme Builder.\n• No custom CSS injected into the portal at MVP.\n• Theme updated test confirms comms team can change the banner image without IT assistance.\n• Mobile rendering tested on iOS and Android for the branded theme.",
     "'Custom CSS breaks on every Employee Center release. Theme Builder is the supported path — it is what ServiceNow tests against. One rogue CSS file has caused portal outages on platform upgrades.'",
     1, "Baseline"),

    ("EC-006", "4.6 Entry Point Consolidation",
     "As a Change Manager, I want the legacy Service Portal (/sp) decommissioned or redirected to Employee Center after go-live so that users have one URL, not two.",
     "Customer confirms go-live date for Employee Center and agrees on the /sp redirect plan.",
     "• /sp redirect to /esc configured and tested before go-live announcement.\n• Legacy portal bookmarks and email links updated in all external communications.\n• Help desk briefed on the portal change and redirect before user communications go out.\n• Legacy /sp access disabled (or redirect-only) 30 days after Employee Center go-live.",
     "'Leaving /sp active creates a split user base — some users on the old portal, some on the new one. Support costs double, content diverges. The redirect is a 30-minute configuration.'",
     1, "Baseline"),
  ],

  "KM — Sprint 3\nKnowledge Management": [
    ("KM-001", "4.1 Knowledge Base Structure",
     "As a Knowledge Manager, I want no more than 3 knowledge bases (IT, HR, General) configured at MVP so that search relevance stays high and governance overhead stays manageable.",
     "Customer confirms knowledge base count and audience per KB before any articles are created.",
     "• ≤ 3 knowledge bases at MVP.\n• Each KB has a named manager and a defined target audience.\n• KB visibility rules configured (IT KB visible to all; HR KB visible to HR and employees; internal tech KB restricted to agents).\n• Articles cannot be created without a KB assignment — no orphan articles.",
     "'One KB per team is a common request that produces a fragmented search experience. Three KBs with clear audience rules cover every access scenario. Consolidate first, split later only if there is a genuine access control reason.'",
     2, "Baseline"),

    ("KM-002", "4.2 Article Template",
     "As a Knowledge Author, I want a standardized OOTB knowledge article template (symptom / cause / resolution / related items) enforced so that all articles are usable as deflection content.",
     "Customer confirms article template structure before authors are given KB write access.",
     "• Article template configured with mandatory sections: Symptom, Cause, Resolution, Related CIs (optional).\n• Template enforced via OOTB Article Template record — authors cannot publish without all mandatory sections.\n• 5 pilot articles created using the template and rated by 3 non-author reviewers for clarity.\n• Deflection rate measured from portal search result click-throughs to article (baseline established).",
     "'Unstructured articles look like a wiki dump and fail as deflection content. The Symptom/Cause/Resolution format is what the Virtual Agent and search ranking algorithms are tuned to surface.'",
     2, "Baseline"),

    ("KM-003", "4.3 Ownership & Governance",
     "As a Knowledge Manager, I want every article assigned to a named owner (not a group) so that stale content has a single accountable person, not a committee.",
     "Customer confirms named ownership policy before KB launch.",
     "• Every published article has a named kb_knowledge.author who is still active in the organization.\n• Stale article report runs monthly: articles not reviewed in 12 months are flagged for the named owner.\n• Retired articles are moved to 'Retired' state, not deleted — preserve audit trail.\n• Knowledge Manager receives monthly stale article summary via OOTB scheduled report.",
     "'Group-owned articles are no one's responsibility. Named ownership with a stale content report creates the accountability loop that keeps the KB useful past the 6-month mark.'",
     2, "Baseline"),

    ("KM-004", "4.4 Approval Workflow",
     "As a Knowledge Reviewer, I want OOTB article approval workflow (Draft → Review → Published) enforced for all articles so that unreviewed content cannot reach end users.",
     "Customer confirms who the reviewers are per KB and whether one reviewer or two is required for publication.",
     "• OOTB knowledge workflow active for all KBs (Draft → Awaiting Approval → Published).\n• Reviewer assigned automatically from the KB's designated reviewer group.\n• Article cannot transition to Published without at least one explicit approval.\n• Author notified of reviewer feedback within the article record (no external email chains).",
     "'Direct publish access for authors is how 'my workaround from 2019' becomes the top search result for three years. The approval step is one click for the reviewer — it is not a bottleneck; it is a quality gate.'",
     2, "Baseline"),

    ("KM-005", "4.5 Content Strategy",
     "As a Knowledge Manager, I want the initial KB seeded with the top 25 agent-known resolutions (converted from resolved incidents) so that deflection value is measurable from day one.",
     "Customer confirms the top 25 most-resolved incident categories as the article seed list.",
     "• 25 pilot articles created from top resolved incident categories before KB launch.\n• Each article linked to at least one resolved incident from which it was derived.\n• Portal search for each article's symptom keyword returns the article on page 1.\n• Deflection rate baseline measured at 30 days post-launch (target: ≥ 10% of portal searches end without ticket creation).",
     "'Launching an empty KB and hoping agents will populate it produces an empty KB. Seeding from the 25 most-resolved incidents gives users immediate value and gives authors a template to follow.'",
     3, "Baseline"),

    ("KM-006", "4.6 Visibility & Access",
     "As a Portal User, I want knowledge articles surfaced in incident and request forms (related articles widget) so that agents and requesters can resolve issues without creating a ticket.",
     "Customer confirms related articles widget is enabled on Incident and the portal request intake form.",
     "• Related articles widget active on Incident form — triggered by short_description keyword match.\n• Portal request intake form shows article suggestions before submission.\n• Agent clicks article → article view logged and linked to the incident record.\n• Incident resolved by agent using article → kb_use flag set for deflection reporting.",
     "'The KB that lives only in a portal tab and not in the incident form is a missed deflection opportunity. The related articles widget is OOTB — it takes 15 minutes to enable and directly reduces ticket volume.'",
     1, "Baseline"),
  ],

  "VA — Sprint 3\nVirtual Agent": [
    ("VA-001", "4.1 Use Case Scope",
     "As a Service Desk Manager, I want the Virtual Agent MVP limited to 5–8 OOTB topic categories so that the agent is reliable and well-tested before scope expands.",
     "Customer confirms the top 5–8 use cases by volume (pull from top catalog items and incident categories).",
     "• ≤ 8 topics in the VA at go-live.\n• All 8 topics are derived from OOTB Now Platform topic blocks — no custom NLU topics at MVP.\n• Each topic tested with 10 varied utterances before go-live.\n• Out-of-scope utterances fall back to live agent gracefully (no dead-ends).",
     "'A virtual agent with 50 half-built topics is worse than one with 8 well-built ones. Users who hit a dead-end once rarely return. Eight topics that work perfectly build the trust that lets us expand to 50.'",
     3, "Baseline"),

    ("VA-002", "4.2 Core Topics",
     "As a Portal User, I want the Virtual Agent to handle password reset, software request, incident status check, and IT equipment request without agent involvement so that these high-volume requests are deflected 24/7.",
     "Customer confirms these 4 topics are in scope and the back-end fulfillment automation is in place for each.",
     "• Password reset topic integrated with Active Directory password reset API.\n• Software request topic submits catalog item directly without human intervention.\n• Incident status topic returns real-time status from the incident table.\n• Equipment request topic creates a catalog RITM and confirms submission to the user.\n• All 4 topics tested end-to-end with real users in UAT.",
     "'Password reset alone deflects 15–25% of service desk volume in most environments. These 4 topics are the ROI core — they are OOTB Now Platform capabilities, not custom chatbot development.'",
     5, "Baseline"),

    ("VA-003", "4.3 Conversation Design",
     "As a Virtual Agent Conversation Designer, I want OOTB conversation flow templates used as the starting point for all topics so that conversation design is consistent and maintainable without a dedicated chatbot developer.",
     "Customer confirms who owns conversation design post-go-live (cannot be ECS).",
     "• All topics built using OOTB Topic Designer conversation blocks — no custom script includes at MVP.\n• Conversation owner named before go-live and has completed Virtual Agent Topic Designer training.\n• Conversation review checklist completed for each topic (clear prompts, graceful fallback, no dead-ends).\n• Mean conversation length ≤ 5 turns for all MVP topics.",
     "'Custom conversation scripts require a developer every time the wording needs to change. OOTB Topic Designer lets a business analyst update the conversation without touching code.'",
     3, "Baseline"),

    ("VA-004", "4.4 Escalation Strategy",
     "As a Service Desk Agent, I want clear escalation from Virtual Agent to live agent (with full conversation context transferred) so that users never have to repeat themselves after escalating.",
     "Customer confirms whether live agent handoff target is ServiceNow Agent Chat or a third-party (Teams, Slack).",
     "• Escalation to live agent is available on every topic at every point in the conversation.\n• Full conversation transcript is included in the live agent handoff record.\n• Live agent receives context: user identity, topic attempted, conversation summary.\n• Escalation SLA for live agent pickup configured (target: 2-minute response for business hours).",
     "'A virtual agent without a graceful escalation path creates user frustration that is worse than no virtual agent. The escalation experience is the safety net — it must be as well-designed as the topics themselves.'",
     3, "Baseline"),

    ("VA-005", "4.5 Knowledge Integration",
     "As a Portal User, I want the Virtual Agent to surface knowledge articles before creating tickets so that deflection is measurable and article gaps are identified through VA usage data.",
     "Customer confirms the KB is populated (≥ 25 articles) before VA launch.",
     "• VA presents matching KB articles before offering to create an incident for all applicable topics.\n• Article click-through logged in the VA conversation analytics.\n• Topics where KB articles deflect >30% of conversations flagged for article quality improvement.\n• Incidents created from VA tracked with source = 'Virtual Agent' for deflection reporting.",
     "'A virtual agent that creates tickets without checking the KB is just a bot-based ticketing form. The deflection step is what makes the business case. It requires a populated KB — which is why knowledge always precedes virtual agent in the delivery sequence.'",
     2, "Baseline"),

    ("VA-006", "4.6 Channel Strategy",
     "As an IT Leader, I want the Virtual Agent accessible via Employee Center (web) and Microsoft Teams so that users can access support where they already work.",
     "Customer confirms Teams is the primary channel and confirms Azure App Registration access for the Teams bot.",
     "• Virtual Agent embedded in Employee Center portal (OOTB web widget).\n• Teams bot deployed via OOTB Microsoft Teams integration (no custom bot framework).\n• Teams channel tested with 10 users before go-live.\n• Single conversation history visible regardless of channel used.",
     "'Building a custom Teams bot when ServiceNow has a certified native Teams integration is writing code that ServiceNow already wrote. The OOTB integration is the supported, upgradeable path.'",
     3, "Baseline"),
  ],

  "CHG — Sprint 4\nChange Management": [
    ("CHG-001", "4.1 Change Types Usage",
     "As a Change Manager, I want only Normal, Standard, and Emergency change types configured (no custom types) so that every change follows a predictable, auditable path.",
     "Customer confirms they will retire any custom change types from the legacy tool.",
     "• Change type field limited to Normal, Standard, Emergency — no additional values.\n• Change type selection drives the approval and CAB workflow automatically.\n• Legacy custom change types documented and mapped to one of the three OOTB types.\n• Change type selection cannot be changed after approval begins.",
     "'Custom change types are how 'we just need a fast-track option' turns into eight change types with overlapping rules. Three types, three workflows, one governance model.'",
     2, "Baseline"),

    ("CHG-002", "4.2 Approval Model",
     "As a Change Approver, I want approval levels derived from change risk (Normal Low = manager only; Normal High = CAB; Emergency = CAB Chair) so that approval overhead matches actual risk.",
     "Customer confirms risk assessment criteria and approval mapping before any change records are created.",
     "• Risk assessment drives approval routing automatically — no manual override of approval path.\n• Low-risk Normal changes require only implementation manager approval.\n• High-risk Normal changes route to CAB automatically.\n• Emergency changes route to CAB Chair with a 2-hour response SLA.",
     "'Flat approval models (everything goes to CAB) kill velocity. Risk-based routing is OOTB; it requires a risk assessment matrix to be agreed, not custom code.'",
     3, "Baseline"),

    ("CHG-003", "4.3 CAB Structure",
     "As a CAB Chair, I want virtual CAB configured using OOTB CAB Workbench so that CAB review is asynchronous and auditable, not dependent on a weekly meeting.",
     "Customer confirms CAB membership and confirms asynchronous approval is acceptable for low/medium risk changes.",
     "• CAB Workbench enabled and configured for the change advisory group.\n• CAB members can approve/reject asynchronously via email action or ServiceNow mobile.\n• CAB meeting record created automatically for high-risk changes requiring synchronous review.\n• CAB decision captured in the change record with named approver and timestamp.",
     "'Weekly CAB meetings are a bottleneck that pushes teams toward Emergency changes to bypass the process. Asynchronous CAB via the Workbench is OOTB and reduces CAB meeting frequency while improving coverage.'",
     3, "Baseline"),

    ("CHG-004", "4.4 Change Scheduling",
     "As a Change Manager, I want OOTB change conflict detection configured so that changes to the same CI in the same window are flagged automatically before approval.",
     "Customer confirms maintenance window calendar and freeze periods before configuration.",
     "• Change schedule conflict detection enabled for all Normal changes.\n• Maintenance windows and freeze periods configured in the Change Schedule calendar.\n• Conflict warning displayed when a change affects a CI already in an approved change within the same window.\n• Freeze period blocks Normal change approvals (Emergency changes require CAB Chair override).",
     "'Manual conflict checking is how two changes hit the same server in the same weekend and cause an outage. OOTB conflict detection is automated — it requires a populated CI relationship map and a maintenance calendar.'",
     2, "Baseline"),

    ("CHG-005", "4.5 Standard Change Templates",
     "As a Change Implementer, I want pre-approved Standard Change templates for the top 15 routine changes so that routine work bypasses the approval queue without bypassing governance.",
     "Customer confirms the top 15 routine change types (e.g., patch Tuesday, password resets, user account changes).",
     "• 15 Standard Change templates created and pre-approved before go-live.\n• Templates have mandatory implementation tasks with acceptance criteria.\n• Standard Change created from a template requires no additional approvals.\n• Templates reviewed quarterly by CAB Chair for continued pre-approval validity.",
     "'Standard changes without templates mean every implementer re-invents the wheel for routine work. Templates standardize the implementation steps, create institutional knowledge, and the pre-approval is the governance control.'",
     3, "Baseline"),

    ("CHG-006", "4.6 Change ↔ Incident Relationship",
     "As an Incident Manager, I want incidents caused by changes automatically linked to the causative change record so that root cause analysis and PIR data is complete.",
     "Customer confirms this linkage is required for their post-implementation review process.",
     "• OOTB 'caused by change' field on Incident is visible and used by agents.\n• Agents prompted to check for recent changes on the affected CI when creating a P1/P2 incident.\n• Change-caused incidents visible on the change record's related incidents tab.\n• Monthly PIR report includes change-caused incident rate per change type.",
     "'Manual change-incident linkage is rarely done under pressure. Making the field visible and prompting agents at incident creation takes 5 minutes to configure and makes PIR data complete.'",
     2, "Baseline"),
  ],

  "CSDM — Sprint 4\nCSDM": [
    ("CSDM-001", "4.1 Service Definition Scope",
     "As a Service Owner, I want only Business Services and Technical Services defined at MVP (not the full CSDM stack) so that the service model is accurate before it is complex.",
     "Customer confirms they will not attempt to populate all CSDM layers at MVP.",
     "• Business Service and Technical Service records created for the top 10 services.\n• No Offering, Business Capability, or Application Service records at MVP unless already populated.\n• Each Business Service has a named Service Owner and an associated SLA.\n• Service model reviewed and approved by the Service Owner before go-live.",
     "'Attempting the full CSDM stack in Sprint 4 produces a service model that is 80% wrong and 100% orphaned. Two layers — Business Service and Technical Service — give you the ITSM routing and reporting value without the governance debt.'",
     3, "Baseline"),

    ("CSDM-002", "4.2 Service Layers",
     "As a CMDB Manager, I want Business Services linked to their supporting Technical Services and CIs so that service impact analysis in Event Management and Incident Management is accurate.",
     "Customer confirms the CI-to-Technical Service relationships for the top 10 services.",
     "• Top 10 Business Services each have at least one Technical Service mapped.\n• Each Technical Service has at least one CI relationship (cmdb_rel_ci).\n• Service Impact Analysis test: simulate a CI outage and confirm the correct Business Service is flagged.\n• Relationship map reviewed by each Service Owner before go-live.",
     "'A Business Service with no CI relationships is a label, not a model. The CI relationship is what powers service impact analysis, event correlation, and ITSM routing. It is the minimum viable CSDM deliverable.'",
     3, "Baseline"),

    ("CSDM-003", "4.3 Naming Standards",
     "As a CMDB Manager, I want service naming standards enforced (Business Service = business-facing name; Technical Service = technical system name) so that service records are navigable by both business and IT audiences.",
     "Customer confirms naming convention before any service records are created.",
     "• Naming standard documented and signed off before Sprint 4 begins.\n• Business Services named from the business user's perspective (e.g., 'Email', not 'Exchange 2019 Cluster').\n• Technical Services named from the IT system perspective (e.g., 'Exchange Online' not 'Email Backend').\n• Naming review checkpoint included in Sprint 4 retro.",
     "'Business Services named by the IT team end up unrecognizable to the business audience they are meant to represent. The naming standard session takes 30 minutes and prevents months of confusion.'",
     1, "Baseline"),

    ("CSDM-004", "4.4 Ownership",
     "As a Service Owner, I want every Business Service to have a named owner and backup owner so that service catalog, SLA, and incident routing decisions always have an accountable stakeholder.",
     "Customer confirms named owners for each Business Service before records are created.",
     "• Business Service ownership field populated with a named individual (not a group) for all 10 services.\n• Backup owner named for each service.\n• Service Owner receives a weekly digest of incidents, changes, and SLA performance for their service.\n• Owner list reviewed quarterly — orphaned services (owner left org) escalated immediately.",
     "'Service ownership is not a ServiceNow configuration — it is an organizational commitment. We can configure the field, but the customer must name the owners. Without a named owner, the service model becomes a CMDB artifact that no one maintains.'",
     1, "Baseline"),

    ("CSDM-005", "4.5 Relationships",
     "As a CMDB Manager, I want service-to-CI relationships maintained by Discovery (automated) not by manual data entry so that the service model stays accurate beyond go-live.",
     "Customer confirms Discovery is in scope and will cover the CIs that support the top 10 services.",
     "• CI records under each Technical Service are populated by Discovery, not manual import.\n• Discovery schedule set to refresh CI-to-service relationships at least weekly.\n• Manual relationship overrides documented and flagged for Discovery team review.\n• CMDB Health Dashboard shows relationship completeness % for each Technical Service.",
     "'Manually maintained service-to-CI relationships degrade within 90 days as infrastructure changes. Discovery-maintained relationships are the only sustainable model at scale.'",
     2, "Baseline"),

    ("CSDM-006", "4.6 Integration with ITSM",
     "As an Incident Manager, I want the 'Affected Service' field on Incident to use the CSDM Business Service list so that incident routing and reporting are aligned to the service model.",
     "Customer confirms the CSDM Business Services list will be used as the Incident affected_service reference.",
     "• Incident form's affected_service field references cmdb_ci_service (Business Services only).\n• Incident assignment group populated from the affected service's support group when CI is unknown.\n• Service-level incident report available OOTB using the affected_service field.\n• Agents trained to populate affected_service at incident creation (not just at resolution).",
     "'Incidents that do not reference a service cannot generate service-level reporting. The CSDM-ITSM connection is a field reference change, not a development project.'",
     2, "Baseline"),
  ],

  "CMDB — Sprint 4\nCMDB": [
    ("CMDB-001", "4.1 CI Scope",
     "As a CMDB Manager, I want CI scope at MVP limited to the classes that directly support incident routing and SLA (servers, network devices, applications) so that the CMDB is accurate before it is comprehensive.",
     "Customer confirms in-scope CI classes before any CI data is loaded.",
     "• CI scope document signed off before any data load.\n• In-scope CI classes: cmdb_ci_server, cmdb_ci_netgear, cmdb_ci_appl — maximum 5 CI classes at MVP.\n• Out-of-scope CI classes documented with a Phase 2 onboarding date.\n• All in-scope CIs have a named owner and a support group populated.",
     "'Loading every CI class at MVP produces a CMDB that is large and inaccurate. Accurate data for 5 critical classes is worth more than 50% complete data for 20 classes.'",
     2, "Baseline"),

    ("CMDB-002", "4.2 CI Classes to Use",
     "As a Discovery Engineer, I want only OOTB CI classes used (no custom classes created) so that Discovery population, CSDM mapping, and Event Management correlation all work without custom integration.",
     "Customer confirms that any legacy custom CI classes will be retired or mapped to OOTB equivalents.",
     "• Zero custom CI classes created at MVP.\n• Legacy custom CI types mapped to their closest OOTB equivalent (documented in CI class mapping worksheet).\n• Discovery population confirmed for all in-scope OOTB CI classes.\n• CI class mapping reviewed and approved by Discovery, ITSM, and Event Management leads jointly.",
     "'Custom CI classes break Discovery population, require custom identification rules, and are not recognized by Event Management correlation. The OOTB class hierarchy covers every CI type a federal IT environment uses.'",
     2, "Baseline"),

    ("CMDB-003", "4.3 Data Quality Standards",
     "As a CMDB Manager, I want CMDB Health Dashboard targets defined before go-live (Completeness ≥ 80%, Correctness ≥ 85%, Compliance ≥ 80%) so that data quality is measured and actionable from day one.",
     "Customer confirms health targets and names the owner of the monthly CMDB Health review.",
     "• CMDB Health Dashboard enabled and visible to the CMDB Manager at go-live.\n• Baseline health scores captured at end of Sprint 4.\n• Monthly health review meeting scheduled with named participants.\n• CIs below threshold escalated to the CI class owner — not left in the general CMDB queue.",
     "'CMDB quality without a measurement baseline is unmeasurable. The health dashboard is OOTB — it requires health rules to be configured, which is a half-day activity that pays dividends across the entire engagement.'",
     2, "Baseline"),

    ("CMDB-004", "4.4 Ownership",
     "As a CMDB Manager, I want every CI class to have a named class owner so that data quality accountability is assigned and the health review has a decision-maker per domain.",
     "Customer confirms named owners per CI class before data load begins.",
     "• Class owner field populated on all in-scope CI class records.\n• Class owner receives weekly health report for their CI class.\n• Data quality issues escalated to class owner — not to a generic IT email.\n• Owner list reviewed at Sprint 6 retro and updated for go-live.",
     "'A CMDB without class owners is a CMDB without anyone accountable for data quality. The class owner role takes 30 minutes to assign and is the single most important governance decision in the CMDB workstream.'",
     1, "Baseline"),

    ("CMDB-005", "4.5 Relationships",
     "As an Incident Manager, I want CI-to-CI relationships (server hosts application, application depends on database) populated for all in-scope CIs so that service impact analysis produces meaningful results.",
     "Customer confirms relationship scope (which relationships must be in place for ITSM to function correctly).",
     "• Relationship types limited to OOTB rel_type values — no custom relationship types.\n• Server-to-application and application-to-database relationships populated for all in-scope CIs.\n• Relationship completeness target: ≥ 70% of in-scope CIs have at least one relationship.\n• Impact analysis test: CI marked offline → all related applications flagged correctly.",
     "'Isolated CI records with no relationships cannot support service impact analysis, change conflict detection, or Event Management service correlation. Relationships are the data that makes CMDB valuable — not the CI records themselves.'",
     3, "Baseline"),

    ("CMDB-006", "4.6 Integration with ITSM",
     "As a Service Desk Agent, I want the Affected CI field on Incident auto-populated from Event Management alerts so that CI context is captured without manual agent lookup.",
     "Customer confirms Event Management is in scope and alert-to-incident promotion is configured.",
     "• Incidents promoted from Event Management alerts inherit cmdb_ci from the alert.\n• Manually created incidents have a CI lookup widget on the form (search by name, IP, or tag).\n• CI selection filters by environment (production CIs shown by default).\n• Incident-CI link captured in the cmdb_ci field — not in a free-text notes field.",
     "'Free-text CI entry in incident notes is not CMDB integration. The cmdb_ci field link is what powers service impact reporting, change conflict detection, and CMDB-driven routing.'",
     2, "Baseline"),
  ],

  "DISC — Sprint 4\nDiscovery": [
    ("DISC-001", "4.1 Discovery Scope",
     "As a Discovery Engineer, I want IP range scope limited to production infrastructure at MVP so that Discovery resources are focused on the CIs that matter for ITSM and Event Management.",
     "Customer confirms production IP ranges and confirms non-production ranges are excluded at MVP.",
     "• Discovery schedule configured for production IP ranges only.\n• Non-production IP ranges explicitly excluded from the Discovery schedule.\n• Estimated CI count per IP range documented for MID Server sizing.\n• Out-of-scope ranges listed in the Discovery scope document with Phase 2 date.",
     "'Discovering 10,000 development VMs in Sprint 4 fills the CMDB with low-value CIs and obscures the production CIs that ITSM depends on. Scope production first — the rest follows in Phase 2.'",
     2, "Baseline"),

    ("DISC-002", "4.2 MID Server Placement",
     "As a Discovery Engineer, I want MID Servers placed in each network segment that contains production infrastructure so that Discovery reaches all CIs without firewall exceptions that bypass security policy.",
     "Customer confirms network segment layout and the firewall policy for MID Server communication.",
     "• One MID Server per network segment with production CIs that cannot be reached from a central MID Server.\n• MID Server placement approved by customer network security team before installation.\n• MID Server validation test: Discovery triggered from each MID Server; at least 10 CIs discovered per segment.\n• MID Server status monitored via ServiceNow MID Server Health Dashboard.",
     "'A single centralized MID Server that requires firewall exceptions to every segment is a security risk. The right architecture is MID Servers in-segment — it is more MID Servers but fewer firewall holes.'",
     3, "Baseline"),

    ("DISC-003", "4.3 Credential Strategy",
     "As a Security Admin, I want Discovery credentials stored in ServiceNow Credential Store (not in MID Server config files) and scoped per CI class so that credentials are auditable and follow least-privilege.",
     "Customer confirms credential types needed (Windows domain, SSH, SNMP) and approves storage in ServiceNow.",
     "• All credentials stored in ServiceNow Credential Store — no credentials in MID Server config files.\n• Credential scope set per CI class (Windows credentials for servers; SNMP for network devices).\n• Credential rotation tested: rotating a credential in Credential Store takes effect on next Discovery run without MID Server restart.\n• Credential access audit log reviewed and confirmed with customer security team.",
     "'Credentials in MID Server config files are plaintext-adjacent and outside the ServiceNow audit trail. The Credential Store is encrypted, audited, and scoped — it is the only acceptable credential management approach.'",
     3, "Baseline"),

    ("DISC-004", "4.4 Discovery Scheduling",
     "As a CMDB Manager, I want Discovery scheduled to run at least weekly during off-peak hours so that CMDB data is never more than 7 days stale for production CIs.",
     "Customer confirms acceptable Discovery window (day/time) for production impact.",
     "• Weekly Discovery schedule configured for each production IP range.\n• Discovery window confirmed as off-peak (e.g., Sunday 2am–6am).\n• Discovery completion notification sent to CMDB Manager after each run.\n• Discovery run duration baseline measured; alert configured if run exceeds 150% of baseline.",
     "'Monthly Discovery produces a CMDB that is 30 days out of date by definition. Weekly is the minimum viable cadence for production CIs that are actively being changed. Critical infrastructure can run daily.'",
     1, "Baseline"),

    ("DISC-005", "4.5 CI Class Scope",
     "As a CMDB Manager, I want Discovery configured to populate only the in-scope CI classes (agreed in the CMDB workbook) so that Discovery does not auto-create CI types that the CMDB team is not ready to govern.",
     "Customer confirms CI class scope aligned with the CMDB CI class decision (CMDB-002).",
     "• Discovery patterns activated only for in-scope CI classes.\n• Discovery-created CIs land in the correct OOTB CI class — no 'Unknown' or generic cmdb_ci records.\n• New CI class auto-creation disabled — Discovery engineer must explicitly activate a new pattern.\n• CI class scope reviewed at Sprint 6 and Phase 2 scope updated.",
     "'Discovery without CI class governance creates 'Unknown' records that pollute the CMDB faster than the CMDB team can clean them up. Activate patterns deliberately, not by default.'",
     2, "Baseline"),

    ("DISC-006", "4.6 Data Quality & Validation",
     "As a CMDB Manager, I want Discovery results validated against a known CI inventory before go-live so that Discovery accuracy is measured before it becomes the CMDB's primary data source.",
     "Customer provides a reference CI list (from existing CMDB, SCCM, or network scan) for validation.",
     "• Discovery run output compared to reference CI list: ≥ 85% of reference CIs discovered.\n• Discrepancies (CIs in reference but not in Discovery) investigated and explained.\n• Discovery-created CI attributes (name, IP, OS version) validated against the reference for 20 sample CIs.\n• Validation report provided to customer CMDB Manager before Discovery is set as the authoritative source.",
     "'Going live with Discovery as the authoritative source without validating its output is a trust-without-verify moment that creates CMDB debt in the first 30 days. The validation is a one-time effort that earns permanent confidence.'",
     2, "Baseline"),
  ],

  "SGC — Sprint 5\nService Graph Connectors": [
    ("SGC-001", "4.1 Source Systems",
     "As a CMDB Manager, I want only source systems that have a certified ServiceNow Service Graph Connector used at MVP so that CI population is supported, upgradeable, and auditable.",
     "Customer confirms which source systems (e.g., Azure, AWS, Intune, SCCM) have certified SGC connectors.",
     "• Source systems limited to those with a certified SGC connector from the ServiceNow Store.\n• Each connector version validated against the customer's ServiceNow instance version.\n• Non-certified source systems documented with Phase 2 assessment date.\n• SGC connector health dashboard enabled for each active connector.",
     "'Custom integration for CI population bypasses the reconciliation engine, requires developer maintenance, and breaks on every platform upgrade. Certified SGC connectors are the supported path — if one doesn't exist for a source, the source goes in Phase 2.'",
     2, "Baseline"),

    ("SGC-002", "4.2 Data Scope",
     "As a CMDB Manager, I want each SGC configured to import only the CI attributes needed for ITSM and Event Management (not every available field from the source) so that CMDB data stays clean and import performance stays acceptable.",
     "Customer confirms minimum required CI attributes per source system.",
     "• SGC field mapping limited to: name, IP, OS, serial/tag, environment, support group (where available).\n• Source-specific attributes not in the CMDB schema stored in additional_attributes — not as custom columns.\n• Import volume tested: SGC run completes within the agreed maintenance window.\n• Attribute mapping reviewed by CMDB Manager before SGC goes live.",
     "'Importing every field a source system offers produces a CI record that is 80% noise. The CMDB is a governance tool, not a data warehouse. Import the minimum needed for routing and analysis.'",
     2, "Baseline"),

    ("SGC-003", "4.3 CI Class Mapping",
     "As a CMDB Manager, I want SGC CI class mapping reviewed and approved before the first production import so that source system objects land in the correct OOTB CI class.",
     "Customer confirms CI class mapping between source system object types and ServiceNow CI classes.",
     "• CI class mapping document reviewed and approved by CMDB Manager before first import.\n• No 'Unknown' CI class records created after the first SGC run.\n• Mapping conflicts (source object maps to multiple possible CI classes) resolved before import.\n• Mapping reviewed after first run: all imported CIs in expected classes.",
     "'Default SGC mappings are a starting point, not a finished product. Every environment has source system object types that need to be mapped to the right CI class — skipping this review produces CMDB pollution on the first run.'",
     2, "Baseline"),

    ("SGC-004", "4.4 Reconciliation Rules",
     "As a CMDB Manager, I want OOTB Identification and Reconciliation Engine (IRE) rules configured so that duplicate CI creation is prevented when multiple sources report the same CI.",
     "Customer confirms which source is authoritative per CI attribute (e.g., Intune for serial number; Discovery for IP).",
     "• IRE rules configured for all in-scope CI classes before first SGC import.\n• Authority rules defined: which source wins per attribute when sources disagree.\n• Duplicate CI prevention tested: same CI imported from two sources produces one CI record.\n• IRE conflict report reviewed weekly in first 30 days post-launch.",
     "'Without IRE rules, two SGCs that both report the same server create two CI records. The IRE is the OOTB reconciliation engine — it requires authority rules to be configured, which is a one-time configuration activity.'",
     3, "Baseline"),

    ("SGC-005", "4.5 Data Ownership",
     "As a CMDB Manager, I want each SGC-managed CI class to have a named data owner so that changes to SGC mapping and reconciliation rules have an accountable decision-maker.",
     "Customer confirms data owner per SGC source system.",
     "• Data owner named per SGC source system.\n• Data owner approves all changes to CI class mapping and IRE rules.\n• Monthly data owner review meeting scheduled for first 3 months.\n• Orphaned ownership (owner left org) escalated to CMDB Manager immediately.",
     "'SGC configuration without a named owner becomes ECS-dependent for every change. The data owner is the person who can say 'yes, update the mapping' without involving the implementation team.'",
     1, "Baseline"),

    ("SGC-006", "4.6 Sync Frequency",
     "As a CMDB Manager, I want SGC sync schedules aligned with source system change frequency so that CMDB data latency is appropriate to the volatility of each source.",
     "Customer confirms acceptable sync latency per source system.",
     "• Azure/AWS SGCs: daily sync (cloud infrastructure changes daily).\n• Intune/SCCM SGCs: weekly sync (endpoint changes are less frequent).\n• Sync schedule documented and confirmed by the source system owner.\n• Sync failure alert configured: CMDB Manager notified if a scheduled sync fails to complete.",
     "'One-size-fits-all sync frequency wastes resources on stable sources and lets volatile sources drift. Daily for cloud, weekly for on-premises is the right default for most environments.'",
     1, "Baseline"),
  ],

  "HAM — Sprint 5\nHardware Asset Management": [
    ("HAM-001", "4.1 Asset Scope",
     "As a HAM Process Owner, I want asset scope at MVP limited to capital assets (laptops, desktops, servers, network equipment) and excluded consumables so that the asset record has business value without data entry overhead.",
     "Customer confirms asset class scope and the capitalization threshold below which items are consumables.",
     "• Asset classes at MVP: cmdb_ci_computer, cmdb_ci_netgear, cmdb_ci_server.\n• Consumables (cables, mice, keyboards under $X threshold) excluded from asset tracking.\n• Capitalization threshold documented and approved by Finance contact.\n• Excluded items documented in the HAM scope decision record.",
     "'Tracking every cable and power strip in HAM produces an asset database with 50,000 records, 40,000 of which have zero financial or compliance value. Capital assets first — consumables are an inventory problem, not an asset problem.'",
     2, "Baseline"),

    ("HAM-002", "4.2 Asset Classes",
     "As a CMDB Manager, I want OOTB asset CI classes used (no custom hardware classes) so that Discovery population, HAM workflows, and ITSM integration work without custom mapping.",
     "Customer confirms all legacy custom asset types will be mapped to OOTB CI classes.",
     "• Zero custom CI classes created for HAM.\n• Legacy custom asset types mapped to OOTB equivalents (documented in asset class mapping sheet).\n• Discovery populates all in-scope asset CI classes automatically.\n• HAM Workspace dashboard shows all asset classes without custom configuration.",
     "'Custom asset classes require custom Discovery patterns, custom HAM workflows, and custom ITSM field mappings. OOTB classes are pre-integrated across all three — the mapping work is worth it to avoid the integration debt.'",
     2, "Baseline"),

    ("HAM-003", "4.3 Stockroom Strategy",
     "As a Stockroom Manager, I want OOTB stockroom records configured for each physical location where assets are received, staged, or stored so that asset assignment and receiving workflows are location-aware.",
     "Customer confirms physical stockroom locations and the named manager per stockroom.",
     "• Stockroom records created for each physical location (minimum: HQ + each remote site with >10 assets).\n• Named manager assigned per stockroom.\n• Receiving workflow: all new assets received into a stockroom before assignment — never directly to a user.\n• Stockroom inventory report available to each stockroom manager via OOTB HAM Workspace.",
     "'Assets received directly to users bypass the stockroom record and create asset tracking gaps from the first day. The receiving workflow is the OOTB control that keeps the asset record accurate.'",
     2, "Baseline"),

    ("HAM-004", "4.4 Asset Assignment",
     "As a HAM Process Owner, I want asset assignment to users triggered only via a catalog request or an OOTB IMAC task so that asset movements are always documented and auditable.",
     "Customer confirms they will not allow direct edits to the assigned_to field on asset records.",
     "• assigned_to field on asset CI records is read-only — changes only via OOTB IMAC workflow.\n• IMAC catalog item creates a task that drives the assignment change.\n• Asset assignment history auditable from the CI record's activity log.\n• Unauthorized assignment change alert configured (if assigned_to changes without a linked IMAC task).",
     "'Direct edits to assigned_to produce an asset register that nobody trusts within 60 days. The IMAC workflow is the audit trail — without it, asset tracking is a spreadsheet with a ServiceNow UI.'",
     2, "Baseline"),

    ("HAM-005", "4.5 Lifecycle Management",
     "As a HAM Process Owner, I want OOTB asset lifecycle states (In Stock, In Use, In Maintenance, Retired) enforced so that asset disposition is trackable and disposal is documented.",
     "Customer confirms lifecycle state definitions match their asset management policy.",
     "• Asset lifecycle states limited to OOTB values: In Stock, In Use, In Maintenance, Retired.\n• State transitions driven by OOTB IMAC tasks — not direct field edits.\n• Retired assets moved to Retired state with a named disposal method (sold, recycled, destroyed) before removal from active inventory.\n• Monthly asset lifecycle report shows count by state and flags assets stuck in 'In Maintenance' > 90 days.",
     "'Custom lifecycle states are how 'Pending Disposal' sits as a state for three years with no one accountable for the physical asset. OOTB states with IMAC tasks create the audit trail that disposition requires.'",
     2, "Baseline"),

    ("HAM-006", "4.6 Data Source",
     "As a CMDB Manager, I want Discovery as the authoritative source for asset attributes (make, model, serial number, OS) with HAM as the authoritative source for financial and assignment data so that neither system duplicates the other's work.",
     "Customer confirms the data authority split before HAM go-live.",
     "• Discovery populates: make, model, serial, OS version, IP address.\n• HAM/manual processes populate: cost, purchase date, warranty expiry, assigned_to, stockroom.\n• IRE rules configured to preserve HAM-owned attributes when Discovery runs.\n• Data authority document signed off by both CMDB Manager and HAM Process Owner.",
     "'Two systems trying to own the same field create data conflicts and manual reconciliation. The authority split is an organizational agreement, not a technical configuration — but the IRE rules enforce it.'",
     2, "Baseline"),
  ],

  "INT — Sprint 5\nIntegrations": [
    ("INT-001", "4.1 Integration Scope",
     "As an Integration Lead, I want only integrations with a ServiceNow-certified connector used at MVP so that integration maintenance is owned by the connector vendor, not ECS or the customer.",
     "Customer confirms integration list and validates each against the ServiceNow Store for a certified connector.",
     "• Each in-scope integration has a certified ServiceNow Store connector or a native ServiceNow integration (e.g., ITSM Spoke).\n• Custom REST integrations limited to sources with no certified connector and a documented business justification.\n• Integration inventory document lists each integration, its connector, and its data flow direction.\n• Non-certified integrations documented with Phase 2 replacement assessment date.",
     "'Custom integrations are the most common source of upgrade blockers. If a certified connector exists, it is the only acceptable path at MVP.'",
     2, "Baseline"),

    ("INT-002", "4.2 Email Integration",
     "As a Service Desk Manager, I want inbound email-to-incident configured via OOTB Email Client (not a custom email parser) so that email intake is maintainable without developer involvement.",
     "Customer confirms the inbound email address and confirms no custom email parsing logic is required.",
     "• OOTB Email Client configured for the designated service desk inbox.\n• Inbound email creates an incident with: caller (from email), short_description (from subject), description (from body).\n• No custom email parsing scripts at MVP — OOTB parsing handles standard email formats.\n• Email-created incidents tagged with source = 'Email' for channel reporting.",
     "'Custom email parsers break when email formats change — which happens every time a monitoring tool or ITSM client is updated. OOTB parsing covers 95% of email intake scenarios without code.'",
     2, "Baseline"),

    ("INT-003", "4.3 Monitoring Integration",
     "As an Event Management Engineer, I want the monitoring tool integration configured via the certified MID Server connector so that events flow into Event Management without a custom REST endpoint.",
     "Customer confirms monitoring tool names and validates against the certified connector list.",
     "• Monitoring tool connected via certified connector (Dynatrace, SolarWinds, Nagios — as applicable).\n• Test event from each monitoring tool confirmed in em_event table within 60 seconds.\n• No custom REST endpoints created for monitoring event intake at MVP.\n• Event volume baseline captured per source for MID Server capacity validation.",
     "'A custom REST endpoint for monitoring events bypasses the Event Management rules engine and the CMDB correlation layer. The certified connectors feed directly into the rules pipeline — use them.'",
     3, "Baseline"),

    ("INT-004", "4.4 Outbound Integrations",
     "As a Process Owner, I want outbound integrations (email escalation, ITSM notifications to third-party tools) configured via OOTB Flow Designer Spokes so that outbound actions are visible, auditable, and maintainable.",
     "Customer confirms which outbound integration targets are in scope (Teams, Slack, PagerDuty, etc.).",
     "• Outbound integrations use OOTB Flow Designer Spokes (Teams Spoke, Slack Spoke, PagerDuty Spoke).\n• No custom REST callouts from Business Rules at MVP — all outbound actions via Flow Designer.\n• Each outbound integration tested with a test trigger before go-live.\n• Flow Designer execution history used as the audit trail for all outbound events.",
     "'Business Rule REST callouts for outbound notifications are invisible in the admin UI and break silently when the target API changes. Flow Designer Spokes are testable, auditable, and maintainable by a non-developer admin.'",
     3, "Baseline"),

    ("INT-005", "4.5 Data Flow Direction",
     "As an Integration Architect, I want every integration documented with a clear data flow direction and field authority so that data conflicts between ServiceNow and source systems are resolved without manual intervention.",
     "Customer confirms data flow and field authority for each integration before configuration begins.",
     "• Integration data flow matrix completed: source system, target system, fields exchanged, direction (inbound/outbound/bi-directional), authority (who wins on conflict).\n• All bi-directional integrations have conflict resolution rules configured.\n• Field authority documented in the Integration Inventory workbook.\n• IRE rules enforce authority for bi-directional CMDB integrations.",
     "'Undocumented bi-directional integrations produce data fights where ServiceNow and the source system overwrite each other continuously. Document the authority first — then build the integration.'",
     2, "Baseline"),

    ("INT-006", "4.6 Error Handling & Logging",
     "As an Integration Admin, I want all integration errors logged to a ServiceNow Integration Error Log and alerted to a named owner so that failed integrations are detected within 1 hour without manual monitoring.",
     "Customer confirms the named integration owner and the acceptable error detection window.",
     "• OOTB Integration Error Log (or Flow Designer error log) captures all integration failures.\n• Alert configured: integration owner notified within 60 minutes of a failed integration run.\n• Error log reviewed weekly in the first 30 days post-launch.\n• Integration health dashboard shows success/fail rate per integration for the last 7 days.",
     "'Silent integration failures are the most dangerous failure mode — data stops flowing and no one knows for days. Error logging and alerting is a 30-minute configuration that prevents weeks of data drift.'",
     2, "Baseline"),
  ],

  "PA — Sprint 6\nPerformance Analytics": [
    ("PA-001", "4.1 KPI Selection",
     "As an IT Leader, I want MVP KPIs limited to the 10 metrics that directly measure the outcomes promised to the customer in the SOW so that the dashboard drives decisions, not vanity reporting.",
     "Customer confirms the 10 priority KPIs aligned to the engagement's stated outcomes.",
     "• ≤ 10 KPIs configured at MVP.\n• Each KPI maps to a stated engagement outcome or SLA commitment.\n• KPI owner named for each metric (who is accountable for the number).\n• KPI baseline captured on go-live day — no retroactive baselining.",
     "'Dashboards with 50 KPIs produce no actionable decisions. 10 KPIs that are owned and reviewed weekly change behavior. Start narrow and add based on what the customer actually looks at.'",
     2, "Baseline"),

    ("PA-002", "4.2 Core KPIs (Recommended)",
     "As a Service Desk Manager, I want OOTB PA indicators for the 5 core ITSM metrics (MTTR, First Contact Resolution, SLA Compliance, Incident Volume by Priority, Change Success Rate) activated so that baseline performance is visible from day one.",
     "Customer confirms these 5 KPIs are in scope and the data is available to calculate them.",
     "• OOTB PA indicators activated for MTTR, FCR, SLA Compliance %, Incident Volume, Change Success Rate.\n• Each indicator has a target value agreed with the customer before go-live.\n• Dashboard accessible to Service Desk Manager and IT Director from day one.\n• Week-1 actuals compared to target in the Sprint 6 retro.",
     "'These 5 KPIs are the ones every ServiceNow customer should know on day one. They are OOTB — activating them takes one afternoon. Not activating them means flying blind on the metrics that matter most.'",
     3, "Baseline"),

    ("PA-003", "4.3 Data Collection Strategy",
     "As a PA Admin, I want PA data collected from OOTB PA Jobs (not custom scripts) so that data collection is maintained automatically on platform upgrades.",
     "Customer confirms PA Premium is licensed or that PA Standard KPIs cover the agreed metric set.",
     "• PA data collection via OOTB PA Jobs — no custom scripts or scheduled jobs for metric calculation.\n• PA Job schedule set to collect daily data at off-peak hours.\n• Data collection validated: PA indicator values match manual calculation from the source table for 3 KPIs.\n• PA Job failure alert configured: PA Admin notified within 2 hours of a failed collection job.",
     "'Custom scripts for PA data collection break on platform upgrades and are undetectable when they fail silently. OOTB PA Jobs are upgrade-safe and monitored by the platform health dashboard.'",
     2, "Baseline"),

    ("PA-004", "4.4 Breakdown Strategy",
     "As a Service Desk Manager, I want PA KPIs broken down by assignment group and priority (not by individual agent) so that reporting drives team performance improvement, not individual monitoring.",
     "Customer confirms breakdown dimensions agreed with HR and Service Desk leadership.",
     "• KPI breakdowns configured by: assignment group, priority, category, time period.\n• Individual agent breakdown not configured at MVP (HR approval required if added later).\n• Breakdown dimensions reviewed and approved by Service Desk Manager and HR contact.\n• Drilldown from dashboard to underlying incidents available via OOTB PA Analytics Hub.",
     "'Individual-level performance metrics without an HR framework create legal and morale risks. Group-level breakdowns give the Service Desk Manager the actionable data without the HR exposure.'",
     2, "Baseline"),

    ("PA-005", "4.5 Dashboard Design",
     "As an IT Director, I want a single executive dashboard (≤ 6 tiles) showing the top-line KPIs in traffic-light format so that the health of the IT service delivery is visible at a glance without a weekly report.",
     "Customer confirms the 6 tiles and the traffic-light thresholds before dashboard build.",
     "• Executive dashboard has ≤ 6 tiles.\n• Each tile shows current value vs. target in traffic-light format (green/amber/red).\n• Dashboard auto-refreshes daily — no manual refresh required.\n• IT Director can access the dashboard without ServiceNow training (self-service).",
     "'An executive dashboard with 20 tiles is a report, not a dashboard. Six tiles that answer 'is IT performing well or not' is what executives actually need. The detail is one click away.'",
     2, "Baseline"),

    ("PA-006", "4.6 Audience & Usage",
     "As a PA Admin, I want each dashboard published only to the roles that need it so that executives see the executive view, managers see the operational view, and agents see their queue.",
     "Customer confirms dashboard audience per role group.",
     "• Executive dashboard: IT Director role group only.\n• Operations dashboard: Service Desk Manager and Team Lead roles.\n• Agent dashboard (if in scope): individual agent + team lead.\n• Dashboard access reviewed at 30 days — unused dashboards retired or reassigned.",
     "'Publishing all dashboards to all roles creates noise and dilutes focus. Role-appropriate dashboards ensure each audience sees only what they need to act on.'",
     1, "Baseline"),
  ],

  "RPT — Sprint 6\nReporting & Stabilization": [
    ("RPT-001", "4.1 Testing Scope",
     "As a QA Lead, I want UAT focused exclusively on the agreed acceptance criteria from each sprint workbook so that UAT is bounded, measurable, and completable within Sprint 6.",
     "Customer confirms UAT scope is limited to sprint workbook acceptance criteria — no scope creep during UAT.",
     "• UAT test cases derived directly from sprint workbook acceptance criteria — no additional test cases added during UAT.\n• UAT scope document signed off by customer and ECS before testing begins.\n• UAT defects classified: blocking (must fix before go-live) vs. non-blocking (defer to Phase 2).\n• UAT sign-off required from each process area owner before go-live gate review.",
     "'Open-ended UAT is how go-live dates slip by months. Acceptance criteria from the sprint workbooks are the definition of done — UAT confirms we met them, it does not expand them.'",
     3, "Baseline"),

    ("RPT-002", "4.2 Reporting Needs",
     "As an IT Leader, I want a standard report library (top 10 operational reports per module) activated from OOTB report templates so that operational reporting is available at go-live without custom report development.",
     "Customer confirms the top 10 reports needed at go-live and validates they are available as OOTB templates.",
     "• Top 10 operational reports activated from OOTB report templates.\n• Custom reports limited to 2 at MVP — all other requests deferred to Phase 2.\n• Each report scheduled for weekly delivery to the named report owner.\n• Report access controlled by role — service desk agents do not see executive-level financial reports.",
     "'Custom reports in Sprint 6 are scope creep. OOTB report templates cover the operational reports every ServiceNow implementation needs. Document custom report requests for Phase 2.'",
     2, "Baseline"),

    ("RPT-003", "4.3 SLA Tuning",
     "As a Service Owner, I want SLA targets reviewed against Sprint 6 actual performance data and adjusted once (at go-live + 30 days) if targets are materially unreachable so that SLA compliance is meaningful, not aspirational.",
     "Customer confirms a 30-day post-go-live SLA review window is acceptable before any SLA changes.",
     "• SLA performance report generated at go-live + 30 days.\n• SLA targets flagged as unreachable if actual compliance < 70% in the first 30 days.\n• SLA adjustment requires Service Owner approval and Change Manager sign-off.\n• SLA history preserved — adjusted targets are not applied retroactively.",
     "'SLA targets set in Sprint 1 without 30 days of real performance data are often wrong. A structured review at 30 days is the right correction mechanism — it is not scope creep, it is the planned tuning step.'",
     2, "Baseline"),

    ("RPT-004", "4.4 Assignment Optimization",
     "As a Service Desk Manager, I want assignment rule effectiveness reviewed at go-live + 30 days and rules adjusted for the top 5 misrouted categories so that auto-assignment accuracy improves without a development cycle.",
     "Customer confirms a 30-day assignment accuracy review is part of the stabilization scope.",
     "• Assignment accuracy report generated at go-live + 30 days (% incidents auto-routed correctly).\n• Top 5 misrouted categories identified and assignment rules corrected.\n• Assignment rule changes documented in an Update Set.\n• Target: auto-assignment accuracy ≥ 85% at 60 days post-go-live.",
     "'Assignment rules are empirical — they are designed on assumed patterns and corrected on actual patterns. The 30-day tuning session is built into the delivery model for exactly this reason.'",
     2, "Baseline"),

    ("RPT-005", "4.5 Data Quality Standards",
     "As a CMDB Manager, I want a data quality baseline report generated at go-live covering all in-scope modules (CMDB, HAM, Catalog) so that data quality degradation is detectable in Phase 2.",
     "Customer confirms data quality reporting is in scope for Sprint 6.",
     "• CMDB Health Score captured at go-live.\n• HAM asset record completeness % captured at go-live.\n• Catalog item usage report (items with zero submissions in 30 days) generated at go-live + 30 days.\n• Baseline reports archived and used as Phase 2 comparison point.",
     "'Data quality without a baseline is unmeasurable. The go-live snapshot is the reference point that makes Phase 2 health reviews meaningful — it is a 2-hour activity with permanent value.'",
     2, "Baseline"),

    ("RPT-006", "4.6 Handoff & Support Model",
     "As an IT Leader, I want the post-go-live support model (customer admin role, escalation to ECS, ServiceNow support tiers) documented and tested before go-live so that the customer can operate independently from day one.",
     "Customer confirms named customer admin and their readiness to operate the platform independently.",
     "• Named customer ServiceNow Admin confirmed and has completed ServiceNow Administration training.\n• Admin runbook covers: user management, group management, assignment rules, notification updates, Update Set promotion.\n• ECS hypercare engagement window defined (standard: 30 days post-go-live).\n• Escalation path documented: customer admin → ECS hypercare → ServiceNow Support.",
     "'Go-live without a ready customer admin means ECS is the de-facto admin for months after the engagement ends. The admin readiness check is the gate — if the named admin is not ready, go-live should be deferred.'",
     2, "Baseline"),
  ],
}


# ── Build the workbook ─────────────────────────────────────────────────────────

wb = Workbook()
default = wb.active
wb.remove(default)

# ── README tab ─────────────────────────────────────────────────────────────────
readme = wb.create_sheet("README")
banner(readme, 1, "ECS Federal — JIT Baseline Story Library", cols=4)
banner(readme, 2, "Decision-point-driven user stories for all 18-week OOTB delivery sprints. One tab per process area.", sub=True, cols=4)

readme.row_dimensions[4].height = 18
readme.cell(row=4, column=1, value="How to use this workbook").font = _font(NAVY, bold=True, size=11)

readme.row_dimensions[5].height = 72
readme.cell(row=5, column=1, value=(
    "WHAT: Baseline user stories tied to the JIT (Just-in-Time) delivery model. Each story is activated at the moment the customer makes the corresponding workshop decision — not upfront as a waterfall backlog.\n\n"
    "WHAT IS NOT HERE: Basic configuration stories ('configure the platform', 'activate the plugin'). Those are ECS delivery tasks, not decision-driven stories. These stories are activated by customer decisions.\n\n"
    "HOW TO USE: When a customer confirms a decision in the workshop, find the matching story, move Status to Active, refine acceptance criteria to the customer's specific context, and add to the sprint board."
)).alignment = Alignment(wrap_text=True, vertical="top")
readme.cell(row=5, column=1).font = _font(BODY, size=10)

readme.row_dimensions[7].height = 18
readme.cell(row=7, column=1, value="Tab Index").font = _font(NAVY, bold=True, size=11)

header_row(readme, 8, ["Tab", "Process Area", "Sprint", "Story Count"])
readme.column_dimensions["A"].width = 28
readme.column_dimensions["B"].width = 32
readme.column_dimensions["C"].width = 16
readme.column_dimensions["D"].width = 14

tab_index = [
    ("PLT — Sprint 1\nPlatform Foundation",   "Platform Foundation",       "Sprint 1"),
    ("INC — Sprint 1\nIncident Management",   "Incident Management",       "Sprint 1"),
    ("CAT — Sprint 2\nService Catalog & Request", "Service Catalog & Request","Sprint 2"),
    ("EC — Sprint 2\nEmployee Center",         "Employee Center",           "Sprint 2"),
    ("KM — Sprint 3\nKnowledge Management",    "Knowledge Management",      "Sprint 3"),
    ("VA — Sprint 3\nVirtual Agent",           "Virtual Agent",             "Sprint 3"),
    ("CHG — Sprint 4\nChange Management",      "Change Management",         "Sprint 4"),
    ("CSDM — Sprint 4\nCSDM",                  "CSDM",                      "Sprint 4"),
    ("CMDB — Sprint 4\nCMDB",                  "CMDB",                      "Sprint 4"),
    ("DISC — Sprint 4\nDiscovery",             "Discovery",                 "Sprint 4"),
    ("SGC — Sprint 5\nService Graph Connectors","Service Graph Connectors",  "Sprint 5"),
    ("HAM — Sprint 5\nHardware Asset Management","Hardware Asset Management","Sprint 5"),
    ("INT — Sprint 5\nIntegrations",           "Integrations",              "Sprint 5"),
    ("PA — Sprint 6\nPerformance Analytics",   "Performance Analytics",     "Sprint 6"),
    ("RPT — Sprint 6\nReporting & Stabilization","Reporting & Stabilization","Sprint 6"),
]

for i, (tab, area, sprint) in enumerate(tab_index, 9):
    count = len(STORIES.get(tab, []))
    alt = (i % 2 == 0)
    for col, val in enumerate([tab, area, sprint, count], 1):
        c = readme.cell(row=i, column=col, value=val)
        c.fill = _fill(ALT if alt else WHITE)
        c.font = _font(BODY, size=10)
        c.border = THIN
        c.alignment = Alignment(wrap_text=True, vertical="top")
    readme.row_dimensions[i].height = 18

readme.freeze_panes = "A9"


# ── Process area tabs ──────────────────────────────────────────────────────────
for tab_name, stories in STORIES.items():
    ws = wb.create_sheet(tab_name[:31])  # Excel tab name limit = 31 chars

    # Derive sprint label from tab name
    sprint_label = tab_name.split("\n")[0]

    banner(ws, 1, f"JIT Baseline Stories — {tab_name.split(chr(10))[-1]}")
    banner(ws, 2,
           f"{sprint_label}  ·  Decision-point-driven stories  ·  Activate on workshop decision confirmation",
           sub=True)

    header_row(ws, 4, COLS)
    set_cols(ws)
    freeze(ws, row=5)

    for i, story in enumerate(stories, 5):
        story_row(ws, i, list(story), alt=(i % 2 == 0))

total_stories = sum(len(v) for v in STORIES.values())
print(f"Workbook built: {len(STORIES)} tabs, {total_stories} stories total")
wb.save(OUT)
print(f"Saved → {OUT}")
