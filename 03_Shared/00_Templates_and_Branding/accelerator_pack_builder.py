"""
accelerator_pack_builder.py — canonical xlsx workbook builder for ECS Accelerator Packs.

Each AP workbook follows the 8-tab architecture established in AP-04 (HAM Foundations)
and documented in ECS_Accelerator_Pack_Blueprint.docx. This module enforces consistent
styling, font, color palette, and tab structure across every Accelerator Pack in the library.

Standard 8 tabs (in order):
  1. Instructions             (customer-facing — purpose, who fills, sprint window)
  2. Process Decisions        (customer-fillable — questions with ECS OOTB recommendations)
  3. Dependencies             (customer-facing — prerequisites)
  4. Configuration Data       (ECS-owned — final OOTB-aligned values; yellow cells = customer)
  5. R&R                       (joint — RACI matrix ECS vs Customer)
  6. Consultant Guide         (ECS-internal)
  7. Adoption vs Re-engineering (ECS-internal — pushback scenarios + OOTB defense language)
  8. ServiceNow Mapping       (ECS-internal — tables, fields, OOTB features used)

Brand palette (matches AP-04):
  Navy banner fill            #0B1F3A
  White banner text           #FFFFFF
  Cyan section header fill    #ECFEFF
  Amber customer-fill bg      #FEF3C7    (yellow cells = customer to complete)
  Amber customer-fill text    #92400E
  Slate body text             #1E293B
  Navy label text             #0B1F3A
  Light gray alt-row          #F1F5F9
  Font                        Arial throughout
"""
from __future__ import annotations
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# Palette
# =============================================================================
class APColor:
    NAVY        = "0B1F3A"
    WHITE       = "FFFFFF"
    CYAN_BG     = "ECFEFF"
    AMBER_BG    = "FEF3C7"
    AMBER_FG    = "92400E"
    SLATE       = "1E293B"
    ALT_GRAY    = "F1F5F9"
    BORDER      = "E2E8F0"
    BODY_FONT   = "Arial"


# =============================================================================
# Pre-built styles (constructed once at import time)
# =============================================================================
def _font(*, color=APColor.SLATE, bold=False, size=11):
    return Font(name=APColor.BODY_FONT, color=color, bold=bold, size=size)

def _fill(hex_color):
    return PatternFill(fill_type="solid", start_color=hex_color, end_color=hex_color)

THIN_BORDER = Border(
    left=Side(style="thin", color=APColor.BORDER),
    right=Side(style="thin", color=APColor.BORDER),
    top=Side(style="thin", color=APColor.BORDER),
    bottom=Side(style="thin", color=APColor.BORDER),
)


# =============================================================================
# Cell helpers
# =============================================================================
def set_banner(ws, row, text, *, subtitle=False, span_cols=5):
    """Navy banner row (used for sheet title and subtitle at top)."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = _fill(APColor.NAVY)
    cell.font = _font(color=APColor.WHITE, bold=not subtitle, size=14 if not subtitle else 10)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    # Fill across columns
    for c in range(2, span_cols + 1):
        b = ws.cell(row=row, column=c)
        b.fill = _fill(APColor.NAVY)
    if span_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    ws.row_dimensions[row].height = 28 if not subtitle else 20

def set_header_row(ws, row, headers):
    """Navy header row with white bold text. Use for the data table header line."""
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = _fill(APColor.NAVY)
        c.font = _font(color=APColor.WHITE, bold=True)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = THIN_BORDER
    ws.row_dimensions[row].height = 24

def set_section_header(ws, row, text, span_cols=5):
    """Cyan-tinted section header (e.g., '1. Stockroom Inventory')."""
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(APColor.CYAN_BG)
    c.font = _font(color=APColor.NAVY, bold=True)
    c.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(2, span_cols + 1):
        b = ws.cell(row=row, column=col)
        b.fill = _fill(APColor.CYAN_BG)
    if span_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    ws.row_dimensions[row].height = 22

def set_body_row(ws, row, values, *, label_col=1, customer_fill_cols=None,
                  alt_shade=False, alt_color=APColor.ALT_GRAY):
    """Write a body row. customer_fill_cols is a list of 1-based column indexes that
    should be amber-shaded (customer fills in)."""
    customer_fill_cols = customer_fill_cols or []
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v if v is not None else "")
        is_label = (i == label_col)
        c.font = _font(color=APColor.NAVY if is_label else APColor.SLATE, bold=is_label)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = THIN_BORDER
        if i in customer_fill_cols:
            c.fill = _fill(APColor.AMBER_BG)
            c.font = _font(color=APColor.AMBER_FG)
        elif alt_shade:
            c.fill = _fill(alt_color)

def set_columns(ws, widths):
    """Set column widths (in Excel character units)."""
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze_top(ws, rows=4):
    """Freeze the top N rows so the header is sticky on scroll."""
    ws.freeze_panes = ws.cell(row=rows + 1, column=1)


# =============================================================================
# Tab builders — one for each of the 8 standard tabs
# =============================================================================
@dataclass
class TabContent:
    """Container for a single workbook's content across all 8 tabs."""
    workbook_title: str              # e.g. "Software Publishers"
    pack_name: str                   # e.g. "SAM Foundations Accelerator Pack"

    # Tab 1: Instructions
    purpose: str = ""
    who_fills: str = ""
    sprint_window: str = ""
    estimated_effort: str = ""
    related_workbooks: List[str] = field(default_factory=list)
    success_criteria: List[str] = field(default_factory=list)

    # Tab 2: Process Decisions — list of (decision, ootb_recommendation, rationale)
    process_decisions: List[Tuple[str, str, str]] = field(default_factory=list)

    # Tab 3: Dependencies — list of (dependency, status, owner, due, notes)
    dependencies: List[Tuple[str, str, str, str, str]] = field(default_factory=list)

    # Tab 4: Configuration Data — list of sections, each (section_title, [(field, value, notes, is_customer_fill)])
    config_sections: List[Tuple[str, List[Tuple[str, str, str, bool]]]] = field(default_factory=list)

    # Tab 5: R&R — list of (activity, ecs, customer, notes)
    raci_rows: List[Tuple[str, str, str, str]] = field(default_factory=list)

    # Tab 6: Consultant Guide — list of (subhead, prose) tuples; pass empty subhead for plain prose
    consultant_guide_sections: List[Tuple[str, str]] = field(default_factory=list)

    # Tab 7: Adoption vs Re-engineering — list of (request, ootb_pattern, why_ootb, pushback_language, when_to_customize)
    adoption_rows: List[Tuple[str, str, str, str, str]] = field(default_factory=list)

    # Tab 8: ServiceNow Mapping — list of sections, each (section_title, [(label, value, notes)])
    snmap_sections: List[Tuple[str, List[Tuple[str, str, str]]]] = field(default_factory=list)


def build_instructions_tab(wb, content: TabContent):
    ws = wb.create_sheet("Instructions")
    set_banner(ws, 1, f"{content.workbook_title}", span_cols=2)
    set_banner(ws, 2, f"{content.pack_name} · ECS Federal · ServiceNow Practice", subtitle=True, span_cols=2)

    rows = [
        ("Purpose", content.purpose),
        ("Who fills this out", content.who_fills),
        ("Sprint window", content.sprint_window),
        ("Estimated effort", content.estimated_effort),
        ("Related workbooks", " · ".join(content.related_workbooks)),
    ]
    r = 4
    for label, val in rows:
        ws.cell(row=r, column=1, value=label).font = _font(color=APColor.NAVY, bold=True)
        ws.cell(row=r, column=2, value=val).font = _font(color=APColor.SLATE)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = max(22, 16 + 6 * (len(val) // 70))
        r += 1
    r += 1
    if content.success_criteria:
        set_section_header(ws, r, "Success criteria", span_cols=2); r += 1
        for crit in content.success_criteria:
            ws.cell(row=r, column=1, value="•").font = _font(color=APColor.NAVY, bold=True)
            ws.cell(row=r, column=2, value=crit).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=r, column=2).font = _font(color=APColor.SLATE)
            ws.row_dimensions[r].height = 22
            r += 1
    set_columns(ws, [22, 90])

def build_process_decisions_tab(wb, content: TabContent):
    ws = wb.create_sheet("Process Decisions")
    set_banner(ws, 1, f"{content.workbook_title} — Process Decisions")
    set_banner(ws, 2, "Decisions to make collectively. ECS recommendation pre-filled.", subtitle=True)
    set_header_row(ws, 4, ["#", "Decision", "ECS Recommendation (OOTB)", "Customer Decision", "Rationale / Notes"])
    r = 5
    for i, (decision, reco, rationale) in enumerate(content.process_decisions, start=1):
        alt = (i % 2 == 0)
        set_body_row(ws, r, [i, decision, reco, "", rationale],
                     customer_fill_cols=[4], alt_shade=alt)
        # Row height accommodates the longest cell
        max_chars = max(len(decision), len(reco), len(rationale), 30)
        ws.row_dimensions[r].height = max(28, 14 + 5 * (max_chars // 35))
        r += 1
    set_columns(ws, [4, 36, 42, 24, 42])
    freeze_top(ws, 4)

def build_dependencies_tab(wb, content: TabContent):
    ws = wb.create_sheet("Dependencies")
    set_banner(ws, 1, f"{content.workbook_title} — Dependencies", span_cols=6)
    set_banner(ws, 2, "Other Packs and source data that must be in place before configuration.", subtitle=True, span_cols=6)
    set_header_row(ws, 4, ["#", "Dependency", "Status", "Owner", "Due", "Notes"])
    r = 5
    for i, (dep, status, owner, due, notes) in enumerate(content.dependencies, start=1):
        alt = (i % 2 == 0)
        set_body_row(ws, r, [i, dep, status, owner, due, notes], alt_shade=alt)
        max_chars = max(len(dep), len(notes), 20)
        ws.row_dimensions[r].height = max(22, 14 + 5 * (max_chars // 40))
        r += 1
    set_columns(ws, [4, 44, 14, 14, 16, 40])
    freeze_top(ws, 4)

def build_config_data_tab(wb, content: TabContent):
    ws = wb.create_sheet("Configuration Data")
    set_banner(ws, 1, f"{content.workbook_title} — Configuration Data", span_cols=3)
    set_banner(ws, 2, "Final OOTB-aligned configuration. Yellow-shaded values are customer-confirmed inputs.", subtitle=True, span_cols=3)
    set_header_row(ws, 4, ["Field / Setting", "Value", "Notes"])
    r = 5
    for sec_idx, (sec_title, fields) in enumerate(content.config_sections, start=1):
        set_section_header(ws, r, f"{sec_idx}. {sec_title}", span_cols=3); r += 1
        for field_name, value, notes, is_customer_fill in fields:
            fill_cols = [2] if is_customer_fill else []
            set_body_row(ws, r, [field_name, value, notes], customer_fill_cols=fill_cols)
            ws.row_dimensions[r].height = max(22, 14 + 5 * (max(len(str(value)), len(notes)) // 35))
            r += 1
    set_columns(ws, [36, 28, 46])
    freeze_top(ws, 4)

def build_raci_tab(wb, content: TabContent):
    ws = wb.create_sheet("R&R")
    set_banner(ws, 1, f"{content.workbook_title} — Roles & Responsibilities", span_cols=4)
    set_banner(ws, 2, "RACI: R = Responsible, A = Accountable, C = Consulted, I = Informed.", subtitle=True, span_cols=4)
    set_header_row(ws, 4, ["Activity", "ECS", "Customer", "Notes"])
    r = 5
    for i, (activity, ecs, cust, notes) in enumerate(content.raci_rows, start=1):
        alt = (i % 2 == 0)
        set_body_row(ws, r, [activity, ecs, cust, notes], alt_shade=alt)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = max(22, 14 + 5 * (max(len(activity), len(notes)) // 35))
        r += 1
    set_columns(ws, [50, 8, 10, 48])
    freeze_top(ws, 4)

def build_consultant_guide_tab(wb, content: TabContent):
    ws = wb.create_sheet("Consultant Guide")
    set_banner(ws, 1, f"{content.workbook_title} — Consultant Implementation Guide")
    set_banner(ws, 2, "Internal reference for the ECS Solution Architect / Process Consultant.", subtitle=True)
    r = 4
    for subhead, prose in content.consultant_guide_sections:
        if subhead:
            ws.cell(row=r, column=1, value=subhead).font = _font(color=APColor.NAVY, bold=True, size=12)
            ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 22
            r += 1
        if prose:
            ws.cell(row=r, column=1, value=prose).font = _font(color=APColor.SLATE)
            ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = max(40, 14 + 4 * (len(prose) // 80))
            r += 1
        r += 1  # spacer
    set_columns(ws, [110])

def build_adoption_tab(wb, content: TabContent):
    ws = wb.create_sheet("Adoption vs Re-engineering")
    set_banner(ws, 1, f"{content.workbook_title} — Adoption vs Re-engineering", span_cols=5)
    set_banner(ws, 2, "When the customer says 'but our old tool did X.' Use this language.", subtitle=True, span_cols=5)
    set_header_row(ws, 4, ["Customer Request / Legacy Pattern", "OOTB Pattern", "Why OOTB Wins", "Suggested Pushback Language", "When to Customize Anyway"])
    r = 5
    for i, row_data in enumerate(content.adoption_rows, start=1):
        alt = (i % 2 == 0)
        set_body_row(ws, r, list(row_data), alt_shade=alt)
        max_chars = max(len(x) for x in row_data)
        ws.row_dimensions[r].height = max(48, 14 + 4 * (max_chars // 35))
        r += 1
    set_columns(ws, [34, 30, 32, 32, 26])
    freeze_top(ws, 4)

def build_snmap_tab(wb, content: TabContent):
    ws = wb.create_sheet("ServiceNow Mapping")
    set_banner(ws, 1, f"{content.workbook_title} — ServiceNow Mapping", span_cols=3)
    set_banner(ws, 2, "Internal reference — target tables, OOTB features used, integrations.", subtitle=True, span_cols=3)
    r = 4
    for sec_idx, (sec_title, fields) in enumerate(content.snmap_sections, start=1):
        set_section_header(ws, r, f"{sec_idx}. {sec_title}", span_cols=3); r += 1
        for label, value, notes in fields:
            set_body_row(ws, r, [label, value, notes])
            ws.row_dimensions[r].height = max(20, 14 + 4 * (max(len(str(value)), len(notes)) // 40))
            r += 1
    set_columns(ws, [30, 36, 44])
    freeze_top(ws, 4)


# =============================================================================
# Main builder
# =============================================================================
def build_workbook(content: TabContent, output_path: str):
    """Build a full 8-tab workbook and save to output_path."""
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)
    # Build tabs in canonical order
    build_instructions_tab(wb, content)
    build_process_decisions_tab(wb, content)
    build_dependencies_tab(wb, content)
    build_config_data_tab(wb, content)
    build_raci_tab(wb, content)
    build_consultant_guide_tab(wb, content)
    build_adoption_tab(wb, content)
    build_snmap_tab(wb, content)
    wb.save(output_path)
    return output_path
